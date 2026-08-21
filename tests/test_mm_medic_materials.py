"""Unit tests for the MM Medic materials integration (IPRM side).

Covers the XLSX template round-trip and the outgoing client's request signing,
which must match what the MM Medic partner API verifies:
    HMAC-SHA256(secret, "<timestamp>.<raw_body>")
Also covers `mrs.kits_for_instance` (Task 5), which needs the database --
everything else in this file does not.
"""
import hashlib
import hmac
import json
import re
import tempfile
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.admin import routes_materials as routes
from app.extensions import db
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.material_kit import MaterialKit, MaterialKitItem
from app.models.user import User
from app.services import material_reservation_service as mrs
from app.services import xlsx_io
from app.services.mm_medic_client import MMMedicClient, _sign, MMConfigError


# ----------------------------- XLSX round-trip -----------------------------

def test_export_then_parse_ignores_blank_and_zero_rows(monkeypatch):
    # Avoid any network fetch for embedded thumbnails in the unit test.
    monkeypatch.setattr(xlsx_io, '_download_thumb', lambda *a, **k: None)
    catalog = [
        {'sku': 'AAA-1', 'name': 'Голка', 'available': 50, 'image': 'https://x/y.jpg', 'is_consumable': True},
        {'sku': 'BBB-2', 'name': 'Пробірка', 'available': 12, 'image': None, 'is_consumable': True},
        {'sku': 'CCC-3', 'name': 'Марля', 'available': 5, 'image': None, 'is_consumable': False},
    ]
    bio = xlsx_io.export_materials_template_xlsx(catalog)
    tmp = Path(tempfile.gettempdir()) / 'iprm_materials_unit.xlsx'
    tmp.write_bytes(bio.getvalue())
    try:
        wb = load_workbook(tmp)
        ws = wb.active
        # header order: image, sku, name, available, quantity -> quantity col = 5
        ws.cell(row=2, column=5, value=10)   # AAA-1 -> 10
        ws.cell(row=3, column=5, value=0)    # BBB-2 -> 0 (ignored)
        # CCC-3 left blank (ignored)
        wb.save(tmp)

        parsed = xlsx_io.parse_materials_xlsx(tmp)
        assert parsed == {'AAA-1': 10}
    finally:
        tmp.unlink(missing_ok=True)


def test_export_empty_catalog_produces_valid_workbook():
    bio = xlsx_io.export_materials_template_xlsx([])
    tmp = Path(tempfile.gettempdir()) / 'iprm_materials_empty.xlsx'
    tmp.write_bytes(bio.getvalue())
    try:
        assert xlsx_io.parse_materials_xlsx(tmp) == {}
    finally:
        tmp.unlink(missing_ok=True)


def test_export_embeds_image_when_available(monkeypatch):
    import io as _io
    from PIL import Image as _PILImage

    def _fake_thumb(url, max_px=40):
        if not url:
            return None
        buf = _io.BytesIO()
        _PILImage.new('RGB', (max_px, max_px), (200, 0, 0)).save(buf, format='PNG')
        buf.seek(0)
        return buf

    monkeypatch.setattr(xlsx_io, '_download_thumb', _fake_thumb)
    catalog = [
        {'sku': 'IMG-1', 'name': 'Товар', 'available': 5, 'image': 'https://x/a.png'},
        {'sku': 'IMG-2', 'name': 'Без фото', 'available': 3, 'image': None},
    ]
    bio = xlsx_io.export_materials_template_xlsx(catalog)
    from openpyxl import load_workbook as _lw
    tmp = Path(tempfile.gettempdir()) / 'iprm_materials_img.xlsx'
    tmp.write_bytes(bio.getvalue())
    try:
        ws = _lw(tmp).active
        assert len(ws._images) == 1  # only the row with an image URL got a thumbnail
        headers = [c.value for c in ws[1]]
        assert headers[0] == 'Зображення'  # image is the first column
    finally:
        tmp.unlink(missing_ok=True)


def test_materials_filename_includes_date_and_course():
    from types import SimpleNamespace
    from datetime import datetime as _dt
    from app.admin.routes_materials import _materials_filename
    inst = SimpleNamespace(
        start_date=_dt(2026, 7, 15),
        course=SimpleNamespace(title='Плазмотерапія / базовий'),
    )
    # forbidden filename chars removed, whitespace collapsed
    assert _materials_filename(inst) == 'Витр. мат-и на 15.07.2026 Плазмотерапія базовий.xlsx'


# ----------------------------- client signing -----------------------------

def test_sign_matches_hmac_of_timestamp_dot_body():
    ts, body, secret = '1700000000', b'{"a":1}', 'shared'
    expected = hmac.new(secret.encode(), ts.encode() + b'.' + body, hashlib.sha256).hexdigest()
    assert _sign(ts, body, secret) == expected


def test_sign_empty_body_is_defined_for_get():
    ts, secret = '1700000000', 'shared'
    expected = hmac.new(secret.encode(), ts.encode() + b'.', hashlib.sha256).hexdigest()
    assert _sign(ts, b'', secret) == expected


def test_from_settings_requires_enabled_flag():
    class S:
        mm_medic_integration_enabled = False
        mm_medic_api_base_url = 'https://mm-medic.com'
        partner_webhook_secret = 'x'
    with pytest.raises(MMConfigError):
        MMMedicClient.from_settings(S())


def test_from_settings_requires_url_and_secret():
    class S:
        mm_medic_integration_enabled = True
        mm_medic_api_base_url = ''
        partner_webhook_secret = ''
    with pytest.raises(MMConfigError):
        MMMedicClient.from_settings(S())


# ----------------------------- client transport (retry / query) -----------------------------

import app.services.mm_medic_client as mmc


class _FakeResp:
    def __init__(self, status, data):
        self.status_code = status
        self._data = data
        self.ok = 200 <= status < 300
        self.text = ''

    def json(self):
        if self._data is None:
            raise ValueError('no json')
        return self._data


def test_client_retries_on_connection_error(monkeypatch):
    calls = {'n': 0}

    def fake_request(method, url, **kw):
        calls['n'] += 1
        if calls['n'] < 3:
            raise mmc.requests.ConnectionError('boom')
        return _FakeResp(200, {'status': 'ok', 'items': []})

    monkeypatch.setattr(mmc.requests, 'request', fake_request)
    monkeypatch.setattr(mmc.time, 'sleep', lambda *_a, **_k: None)
    res = mmc.MMMedicClient('https://x', 'secret').fetch_catalog()
    assert res.ok is True
    assert calls['n'] == 3  # 2 retries then success


def test_client_stable_request_id_across_retries(monkeypatch):
    seen = []

    def fake_request(method, url, headers=None, **kw):
        seen.append(headers.get('X-IPRM-Request-Id'))
        if len(seen) < 2:
            raise mmc.requests.Timeout('t')
        return _FakeResp(200, {'status': 'ok'})

    monkeypatch.setattr(mmc.requests, 'request', fake_request)
    monkeypatch.setattr(mmc.time, 'sleep', lambda *_a, **_k: None)
    mmc.MMMedicClient('https://x', 's').create_reservation('ref', {}, [{'sku': 'A', 'quantity': 1}])
    assert len(seen) == 2 and seen[0] == seen[1] and seen[0]  # idempotency key stable


def test_fetch_catalog_builds_query(monkeypatch):
    seen = {}

    def fake_request(method, url, **kw):
        seen['url'] = url
        return _FakeResp(200, {'status': 'ok', 'items': []})

    monkeypatch.setattr(mmc.requests, 'request', fake_request)
    mmc.MMMedicClient('https://x', 's').fetch_catalog(consumable=True, search='гол')
    assert '/catalog?' in seen['url']
    assert 'consumable=1' in seen['url'] and 'search=' in seen['url']


def test_client_4xx_not_retried(monkeypatch):
    calls = {'n': 0}

    def fake_request(method, url, **kw):
        calls['n'] += 1
        return _FakeResp(409, {'status': 'insufficient_stock', 'shortfalls': [{'sku': 'A'}]})

    monkeypatch.setattr(mmc.requests, 'request', fake_request)
    res = mmc.MMMedicClient('https://x', 's').create_reservation('r', {}, [{'sku': 'A', 'quantity': 1}])
    assert res.ok is False and res.http_status == 409
    assert calls['n'] == 1 and res.shortfalls


# ----------------------------- submit / update-items (Task 1 routes) -----------------------------

def test_submit_request_posts_items_and_event_meta_no_replace_partial(monkeypatch):
    seen = {}

    def fake_request(method, url, **kw):
        seen['method'] = method
        seen['url'] = url
        seen['body'] = json.loads(kw['data'])
        return _FakeResp(201, {'status': 'created', 'reservation': {}})

    monkeypatch.setattr(mmc.requests, 'request', fake_request)
    res = mmc.MMMedicClient('https://x', 's').submit_request(
        'ref-1', {'event_title': 'Курс', 'event_starts_at': None},
        [{'sku': 'A', 'quantity': 3}],
    )
    assert res.ok is True
    assert seen['method'] == 'POST'
    assert seen['url'].endswith('/reservations/ref-1/submit')
    assert seen['body']['items'] == [{'sku': 'A', 'quantity': 3}]
    assert seen['body']['event_title'] == 'Курс'
    assert 'event_starts_at' not in seen['body']  # None dropped, like create_reservation
    assert 'replace' not in seen['body']  # endpoint has no such concept
    assert 'partial' not in seen['body']
    assert 'external_ref' not in seen['body']  # ref travels in the URL, not the body


def test_update_request_items_posts_to_items_endpoint(monkeypatch):
    seen = {}

    def fake_request(method, url, **kw):
        seen['method'] = method
        seen['url'] = url
        seen['body'] = json.loads(kw['data'])
        return _FakeResp(200, {'status': 'updated'})

    monkeypatch.setattr(mmc.requests, 'request', fake_request)
    res = mmc.MMMedicClient('https://x', 's').update_request_items(
        'ref-1', [{'sku': 'A', 'quantity': 5}],
    )
    assert res.ok is True
    assert seen['method'] == 'POST'
    assert seen['url'].endswith('/reservations/ref-1/items')
    assert seen['body'] == {'items': [{'sku': 'A', 'quantity': 5}]}


# ----------------------------- kits_for_instance (Task 5) -----------------------------

def _course(title='Курс', is_active=False):
    course = Course(title=title, slug=f'kfi-{uuid4().hex[:8]}', is_active=is_active)
    db.session.add(course)
    db.session.flush()
    return course


def _instance(course):
    inst = CourseInstance(course_id=course.id)
    db.session.add(inst)
    db.session.flush()
    return inst


def _kit(course_id=None, name='Набір', is_active=True):
    kit = MaterialKit(name=name, course_id=course_id, is_active=is_active)
    db.session.add(kit)
    db.session.flush()
    return kit


def test_kits_for_instance_returns_course_and_universal_not_other_course(db_session):
    course_a = _course('Курс A')
    course_b = _course('Курс B')
    instance_a = _instance(course_a)

    kit_a = _kit(course_id=course_a.id, name='Набір A')
    kit_universal = _kit(course_id=None, name='Універсальний')
    _kit(course_id=course_b.id, name='Набір B')  # інший курс -- не має потрапити

    kits = mrs.kits_for_instance(instance_a)
    ids = {k.id for k in kits}

    assert kit_a.id in ids
    assert kit_universal.id in ids
    assert len(ids) == 2  # набір курсу B відсутній


def test_kits_for_instance_excludes_inactive(db_session):
    course_a = _course('Курс C')
    instance_a = _instance(course_a)

    active_kit = _kit(course_id=course_a.id, name='Активний')
    _kit(course_id=course_a.id, name='Неактивний', is_active=False)
    _kit(course_id=None, name='Універсальний, неактивний', is_active=False)

    kits = mrs.kits_for_instance(instance_a)
    ids = {k.id for k in kits}

    assert ids == {active_kit.id}


# --------------------- apply-template route (Task 8 debt item) ---------------------
#
# `kits_for_instance` вище перевірено напряму (виклик функції). Тут -- сам
# роут `instance_materials_apply_template`, що нею користується: гарантія
# роута полягає в тому, що `kit_id`, надісланий руками в POST-запиті, не
# може застосувати комплект ЧУЖОГО курсу чи НЕАКТИВНИЙ, бо придатність
# комплекту перераховується ТІЄЮ Ж функцією, що наповнює випадаючий список
# у формі. Досі це підтверджувалось лише читанням коду.

def _admin():
    u = User.create_with_password(
        f'kit-route-{uuid4().hex[:8]}@test.com', 'password123',
        first_name='K', last_name='R', is_admin=True, email_confirmed=True,
    )
    db.session.commit()
    return u


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def _flashes(resp):
    """Тексти flash-повідомлень зі сторінки (JSON-блок, |tojson екранує
    кирилицю у \\uXXXX -- шукати підрядок у сирому HTML не можна)."""
    match = re.search(
        r'<script type="application/json" id="iprm-flash-data">(.*?)</script>',
        resp.get_data(as_text=True), re.S,
    )
    if not match:
        return []
    return [item['message'] for item in json.loads(match.group(1))]


def test_apply_template_rejects_foreign_course_kit_id(client):
    admin = _admin()
    course_own = _course('Курс власний для комплекту')
    course_foreign = _course('Курс чужий для комплекту')
    instance = _instance(course_own)
    kit_foreign = _kit(course_id=course_foreign.id, name='Комплект чужого курсу')
    db.session.add(MaterialKitItem(kit_id=kit_foreign.id, sku='SKU-FOREIGN', quantity=5))
    db.session.commit()

    _login(client, admin)
    resp = client.post(
        f'/admin/instances/{instance.id}/materials/apply-template',
        data={'kit_id': kit_foreign.id, 'multiplier': 1},
    )

    # Роут не приймає rows="чужого" комплекту: редірект без токена
    # передзаповнення -- нічого прикладати.
    assert resp.status_code == 302
    assert 'prefill=' not in resp.headers['Location']

    follow = client.get(resp.headers['Location'])
    assert any('не знайдено' in m for m in _flashes(follow))


def test_apply_template_rejects_inactive_kit_id(client):
    admin = _admin()
    course = _course('Курс з неактивним комплектом')
    instance = _instance(course)
    kit_inactive = _kit(course_id=course.id, name='Неактивний комплект', is_active=False)
    db.session.add(MaterialKitItem(kit_id=kit_inactive.id, sku='SKU-INACTIVE', quantity=4))
    db.session.commit()

    _login(client, admin)
    resp = client.post(
        f'/admin/instances/{instance.id}/materials/apply-template',
        data={'kit_id': kit_inactive.id, 'multiplier': 1},
    )

    assert resp.status_code == 302
    assert 'prefill=' not in resp.headers['Location']

    follow = client.get(resp.headers['Location'])
    assert any('не знайдено' in m for m in _flashes(follow))


def test_apply_template_valid_kit_prefills_selected_items(client, monkeypatch):
    admin = _admin()
    course = _course('Курс з власним комплектом')
    instance = _instance(course)
    kit = _kit(course_id=course.id, name='Комплект курсу')
    db.session.add(MaterialKitItem(kit_id=kit.id, sku='SKU-OWN', quantity=3))
    db.session.commit()

    # Каталог MM Medic не сконфігуровано в тестах -- підмінюємо, щоб рядок
    # SKU-OWN узагалі опинився на сторінці (див. `_build_rows`).
    monkeypatch.setattr(routes.mrs, 'get_catalog', lambda **kw: (
        [{'sku': 'SKU-OWN', 'name': 'Матеріал курсу', 'available': 50, 'price': 1.0}],
        None, False,
    ))

    _login(client, admin)
    resp = client.post(
        f'/admin/instances/{instance.id}/materials/apply-template',
        data={'kit_id': kit.id, 'multiplier': 2},
    )

    assert resp.status_code == 302
    assert 'prefill=' in resp.headers['Location']

    page = client.get(resp.headers['Location'])
    html = page.get_data(as_text=True)
    match = re.search(r'name="sku" value="SKU-OWN">.*?value="(\d+)"', html, re.S)
    assert match is not None, 'рядок SKU-OWN не знайдено на сторінці матеріалів'
    assert match.group(1) == '6'  # 3 (у комплекті) * 2 (множник)
