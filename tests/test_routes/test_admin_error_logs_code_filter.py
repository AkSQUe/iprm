"""`error_code` у журналі помилок приймає будь-який HTTP-статус.

`ErrorLog.error_code` -- будь-яке ціле (`getattr(exception, 'code', 500)` у
`app/models/error_log.py`), а не лише вісім кодів із випадної підказки.
Раніше фільтр звіряв значення проти фіксованого кортежу (`choice_arg`), тож
код на кшталт 502 мовчки не звужував нічого -- адмін бачив невідфільтрований
журнал, вважаючи його звуженим.
"""
import re
from uuid import uuid4

import pytest

from app.extensions import db
from app.models.error_log import ErrorLog
from app.models.user import User


def _uid():
    return uuid4().hex[:8]


@pytest.fixture
def admin(app):
    u = User.create_with_password(
        f'ecf-{_uid()}@test.com', 'password123',
        first_name='E', last_name='C', is_admin=True, email_confirmed=True,
    )
    # commit, а не flush: сторінка журналу починається з захисного
    # db.session.rollback(), який зніс би незакомічені рядки.
    db.session.commit()
    yield u
    db.session.rollback()
    db.session.delete(db.session.merge(u))
    db.session.commit()


@pytest.fixture
def logs(app):
    """502, якого нема у випадній підказці, і 404, який там є."""
    bad_gateway = ErrorLog(
        error_code=502, error_type='BadGateway', error_message=f'lost-{_uid()}',
        url='/upstream/timeout', method='GET',
    )
    not_found = ErrorLog(
        error_code=404, error_type='NotFound', error_message=f'missing-{_uid()}',
        url='/courses/missing', method='GET',
    )
    db.session.add_all([bad_gateway, not_found])
    db.session.commit()
    yield bad_gateway, not_found
    db.session.rollback()
    db.session.delete(db.session.merge(bad_gateway))
    db.session.delete(db.session.merge(not_found))
    db.session.commit()


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def test_code_outside_dropdown_still_narrows(client, admin, logs):
    bad_gateway, not_found = logs
    _login(client, admin)
    html = client.get('/admin/error-logs?error_code=502&days=0').get_data(as_text=True)
    assert bad_gateway.error_message in html
    assert not_found.error_message not in html


@pytest.mark.parametrize('junk', ['abc', '99', '700', ''])
def test_junk_code_behaves_as_no_filter(client, admin, logs, junk):
    _login(client, admin)
    resp = client.get(f'/admin/error-logs?error_code={junk}&days=0')
    assert resp.status_code == 200
    bad_gateway, not_found = logs
    html = resp.get_data(as_text=True)
    assert bad_gateway.error_message in html
    assert not_found.error_message in html


def test_code_outside_dropdown_narrows_export_too(client, admin, logs):
    """Фільтр спільний зі сторінкою -- код 502 звужує і xlsx-експорт."""
    import io

    from openpyxl import load_workbook

    bad_gateway, not_found = logs
    _login(client, admin)
    r = client.get('/admin/error-logs/export?error_code=502&days=0')
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.data))
    ws = wb['Помилки']
    messages = [
        ws.cell(row=row, column=col).value
        for row in range(2, ws.max_row + 1)
        for col in range(1, ws.max_column + 1)
    ]
    assert bad_gateway.error_message in messages
    assert not_found.error_message not in messages


def test_code_outside_dropdown_shows_in_chip(client, admin, logs):
    """Чіпс активного фільтра лишається з підписом, а не порожній «Код: »."""
    _login(client, admin)
    html = client.get('/admin/error-logs?error_code=502&days=0').get_data(as_text=True)
    assert re.search(r'<span class="admin-chip__key">Код:</span>\s*502', html)


def test_junk_per_page_falls_back_to_default(client, admin):
    """`?per_page=<сміття>` не має обходити стелю: журнал тепер читає
    розмір сторінки через `_listing.per_page_arg()`, як і решта реєстрів,
    а не сирим `request.args.get('per_page', 50, type=int)` -- те приймало
    будь-яке ціле напряму в `paginate()`.
    """
    from datetime import datetime, timedelta, timezone

    from app.admin import _listing

    tag = f'perpage-{_uid()}'
    now = datetime.now(timezone.utc)
    for i in range(_listing.LIST_PER_PAGE + 3):
        db.session.add(ErrorLog(
            error_code=500, error_type='ServerError',
            error_message=f'{tag}-{i:03d}', created_at=now - timedelta(seconds=i),
        ))
    # commit, а не flush: та сама причина, що й для admin/logs вище --
    # захисний db.session.rollback() на початку роуту.
    db.session.commit()

    try:
        _login(client, admin)
        html = client.get(
            f'/admin/error-logs?q={tag}&per_page=999999999&days=0',
        ).get_data(as_text=True)
        shown = len(re.findall(
            rf'data-error-message="{re.escape(tag)}-\d{{3}}"', html,
        ))
        assert shown == _listing.LIST_PER_PAGE, (
            f'сміттєвий per_page мав відкотитись на дефолт '
            f'({_listing.LIST_PER_PAGE}), показано {shown}'
        )
    finally:
        db.session.rollback()
        ErrorLog.query.filter(ErrorLog.error_message.like(f'{tag}-%')).delete(
            synchronize_session=False,
        )
        db.session.commit()
