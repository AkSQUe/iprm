"""Пагінація чотирьох вхідних реєстрів адмінки: B2B, запити на курси,
заявки на повернення, відгуки.

Усі чотири раніше тягли ВЕСЬ зріз на кожен рендер списку (`.all()` без
`paginate()`). Три з чотирьох ділять query-helper з роутом /export
(`_b2b_query`, `_course_requests_query`, `_query`) -- головний ризик цієї
правки в тому, щоб пагінація лишилась ЛИШЕ в роуті сторінки. Тому головний
тест тут -- xlsx несе весь зріз, а не одну сторінку.
"""
from tests.support.rbac import grant_role
import io
import json
import re
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.admin._listing import LIST_PER_PAGE
from app.extensions import db
from app.models.auth_identity import AuthIdentity
from app.models.b2b_request import B2BRequest
from app.models.course import Course
from app.models.course_request import CourseRequest, CourseRequestAudit
from app.models.email_log import EmailLog
from app.models.medical_profile import MedicalProfile
from app.models.online_course import OnlineCourse
from app.models.online_enrollment import OnlineEnrollment
from app.models.refund_request import RefundRequest, STATUS_APPROVED, STATUS_NEW
from app.models.review import Review
from app.models.user import User

PAGE = LIST_PER_PAGE  # 50 -- типовий розмір сторінки (без ?per_page=)


def _uid():
    return uuid4().hex[:8]


def _times(n, base=None):
    """n зростаючих UTC-міток -- крок у секундах гарантує детермінований
    порядок незалежно від того, як швидко виконається цикл вставки."""
    base = base or datetime(2024, 1, 1, tzinfo=timezone.utc)
    return [base + timedelta(seconds=i) for i in range(n)]


@pytest.fixture
def admin(app):
    u = User.create_with_password(
        f'pgreq-{_uid()}@test.com', 'password123',
        first_name='П', last_name='А', email_confirmed=True,
    )
    grant_role(u, 'super_admin')
    db.session.commit()
    return u


@pytest.fixture(autouse=True)
def clean(app):
    """Прибрати за собою: заявки, (онлайн-)курси й власних користувачів.

    Явний autouse-teardown поверх відкату транзакції з conftest.db_session:
    для частини шляхів комміт іде через окрему сесію (after_commit-хуки), яка
    зносить і зовнішню транзакцію тесту -- покладатись саме на rollback
    ризиковано (див. tests/test_routes/test_admin_online_listings.py:30-58).
    Діти видаляються перед батьками: RefundRequest -> OnlineEnrollment ->
    OnlineCourse; CourseRequestAudit -> CourseRequest -> Course; User --
    лише після заявок, останнім, разом із AuthIdentity/MedicalProfile.
    """
    def _wipe():
        RefundRequest.query.filter(
            RefundRequest.reason.like('pgrf-%')).delete(synchronize_session=False)
        stale_courses = [
            row.id for row in OnlineCourse.query.filter(
                OnlineCourse.slug.like('pgrf-%')).all()
        ]
        if stale_courses:
            OnlineEnrollment.query.filter(
                OnlineEnrollment.online_course_id.in_(stale_courses),
            ).delete(synchronize_session=False)
            OnlineCourse.query.filter(
                OnlineCourse.id.in_(stale_courses)).delete(synchronize_session=False)

        B2BRequest.query.filter(
            B2BRequest.email.like('pgb2b-%@test.com')).delete(synchronize_session=False)

        # CourseRequestAudit -- дитина CourseRequest (request_id). Масовий
        # .delete() на CourseRequest обходить ORM-каскад
        # (cascade='all, delete-orphan' у моделі), а conftest не вмикає
        # PRAGMA foreign_keys, тож і БД-рівневий ON DELETE CASCADE не
        # спрацьовує -- без явного видалення рядки аудиту лишаються
        # сиротами з changed_by_id на вже видаленого користувача.
        stale_requests = [
            row.id for row in CourseRequest.query.filter(
                CourseRequest.email.like('pgcr-%@test.com')).all()
        ]
        if stale_requests:
            CourseRequestAudit.query.filter(
                CourseRequestAudit.request_id.in_(stale_requests),
            ).delete(synchronize_session=False)
        CourseRequest.query.filter(
            CourseRequest.email.like('pgcr-%@test.com')).delete(synchronize_session=False)
        Course.query.filter(Course.slug.like('pgcr-%')).delete(synchronize_session=False)

        Review.query.filter(Review.text.like('pgrv-%')).delete(synchronize_session=False)

        # Один рядок на прогін test_refund_row_action_redirects_back_to_same_page:
        # лист-відмова (refund_requests.reject -> EmailService) пише
        # EmailLog(to_email=наш admin). Прямого FK на User тут немає, тож
        # нічого сьогодні не падає, але залишене сміття в спільній
        # сесійній БД -- саме той клас витоку, від якого рятує ця фікстура.
        EmailLog.query.filter(
            EmailLog.to_email.like('pgreq-%@test.com')).delete(synchronize_session=False)

        stale_users = [
            row.id for row in User.query.filter(
                User.email.like('pgreq-%@test.com')).all()
        ]
        if stale_users:
            for model in (AuthIdentity, MedicalProfile):
                model.query.filter(model.user_id.in_(stale_users)).delete(
                    synchronize_session=False)
            User.query.filter(User.id.in_(stale_users)).delete(
                synchronize_session=False)
        db.session.commit()

    _wipe()
    yield
    _wipe()


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def _indices(html, pattern):
    return {int(x) for x in pattern.findall(html)}


_PAGER_RE = re.compile(r'<nav class="admin-pagination">.*?</nav>', re.S)


def _pager_html(html):
    """Витягти САМЕ <nav class="admin-pagination">.

    Просто `f'q={tag}' in html` нічого не доводить: той самий рядок пише і
    чіп активного фільтра, і посилання експорту -- обидва є на сторінці
    незалежно від пейджера. Якби хтось написав
    `pager(endpoint, pagination, {})`, загубивши фільтр саме в посиланнях
    next/prev, повна перевірка все одно пройшла б.
    """
    m = _PAGER_RE.search(html)
    assert m, 'пейджер не відрендерився'
    return m.group(0)


_UNDO_RE = re.compile(
    r'<script type="application/json" id="iprm-undo-data">(.*?)</script>', re.S,
)


def _undo_url(html):
    """URL кнопки "Повернути" з тосту відкату (app/undo.py -> base.html)."""
    m = _UNDO_RE.search(html)
    assert m, 'тост відкату не відрендерився'
    return json.loads(m.group(1))['url']


# --------------------------------- B2B ---------------------------------

def _b2b_env(tag, n=PAGE + 1):
    for i, t in enumerate(_times(n)):
        db.session.add(B2BRequest(
            first_name='Тест', last_name=f'{tag}{i:03d}', phone='+380670000000',
            email=f'pgb2b-{tag}{i:03d}@test.com', team_size='3-5',
            status='new', created_at=t,
        ))
    db.session.commit()


def test_b2b_pagination_filter_perpage_export_and_new_count(client, admin):
    tag = _uid()
    _b2b_env(tag)  # PAGE + 1 -- рівно дві сторінки під тегом, усі status=new
    # Ще дві заявки status=new, які НЕ підпадають під q=tag: new_count --
    # ГЛОБАЛЬНИЙ підрахунок, а не розмір поточного (відфільтрованого)
    # зрізу. Без цих двох мутація new_count=pagination.total лишалась би
    # непоміченою -- під самим лише тегом обидва числа збігаються (51).
    _b2b_env(_uid(), n=2)
    expected_new_count = B2BRequest.query.filter_by(status='new').count()
    _login(client, admin)
    pattern = re.compile(rf'pgb2b-{tag}(\d{{3}})@test\.com')

    page1 = client.get(f'/admin/b2b-requests?q={tag}').get_data(as_text=True)
    page2 = client.get(f'/admin/b2b-requests?q={tag}&page=2').get_data(as_text=True)
    idx1, idx2 = _indices(page1, pattern), _indices(page2, pattern)
    # Перша сторінка -- рівно per_page рядків, друга -- решта, без повторів.
    assert len(idx1) == PAGE
    assert len(idx2) == (PAGE + 1) - PAGE
    assert idx1.isdisjoint(idx2)
    assert idx1 | idx2 == set(range(PAGE + 1))
    # page=2 не губить активний фільтр САМЕ у посиланнях пейджера
    # (next/prev), а не десь на сторінці взагалі (чіп/експорт теж
    # містять q={tag}, але це не той факт, що перевіряється тут).
    assert f'q={tag}' in _pager_html(page1)
    assert f'q={tag}' in _pager_html(page2)

    # ?per_page=25 справді дає 25 рядків.
    p25 = client.get(f'/admin/b2b-requests?q={tag}&per_page=25').get_data(as_text=True)
    assert len(_indices(p25, pattern)) == 25
    # Сміттєве значення відкочується на дефолт (50).
    garbage = client.get(f'/admin/b2b-requests?q={tag}&per_page=99999').get_data(as_text=True)
    assert len(_indices(garbage, pattern)) == PAGE

    # new_count -- незалежний ГЛОБАЛЬНИЙ запит: не довжина сторінки і не
    # pagination.total (розмір зрізу під q=tag, 51, а не 53).
    n1 = re.search(r'Нових:\s*(\d+)', page1).group(1)
    n2 = re.search(r'Нових:\s*(\d+)', page2).group(1)
    assert n1 == n2 == str(expected_new_count)

    # Експорт лишається по ВСЬОМУ зрізу (головний тест плану), і per_page не
    # потрапляє в аркуш «Фільтри», як нібито він теж звужує вибірку.
    r = client.get(f'/admin/b2b-requests/export?q={tag}')
    wb = load_workbook(io.BytesIO(r.data))
    assert wb['B2B-заявки'].max_row - 1 == PAGE + 1
    filter_labels = {c.value for c in wb['Фільтри']['A'] if c.value}
    assert 'Рядків на сторінці' not in filter_labels
    assert 'per_page' not in filter_labels


def test_b2b_out_of_range_page_offers_way_back(client, admin):
    """Бокмарк/старий лінк на сторінку за межею останньої не має брехати
    "Заявок поки немає" (записи є, просто не тут) і не має лишати без
    жодного шляху назад.

    БЕЗ q=: активний фільтр сам по собі вже робить `filter_args` непорожнім
    і викликає гілку "Нічого не знайдено" незалежно від сторінки -- це саме
    те, як `empty_state` поводився і ДО фіксу. Єдиний спосіб перевірити, що
    `narrow_args={'page': ...}` справді щось важить -- узяти незвужений
    зріз, де без нього narrowed лишався б порожнім.
    """
    tag = _uid()
    _b2b_env(tag, n=3)
    _login(client, admin)
    html = client.get('/admin/b2b-requests?page=999').get_data(as_text=True)
    assert html.count(f'pgb2b-{tag}') == 0
    assert 'скинути фільтри' in html


def test_b2b_row_action_redirects_back_to_same_page(client, admin):
    """Дія в рядку (збереження статусу) на сторінці 3 черги не має скидати
    менеджера на першу сторінку -- форма несе зріз (фільтр + сторінка) у
    своєму action-URL, а роут повертає туди ж через `_back()`."""
    tag = _uid()
    _b2b_env(tag)  # PAGE + 1 -- рівно дві сторінки під тегом
    _login(client, admin)
    page2 = client.get(f'/admin/b2b-requests?q={tag}&page=2').get_data(as_text=True)
    m = re.search(r'action="(/admin/b2b-requests/\d+/update\?[^"]*page=2[^"]*)"', page2)
    assert m, 'дію рядка з page=2 у action-URL не знайдено на другій сторінці'
    action_url = m.group(1).replace('&amp;', '&')
    resp = client.post(action_url, data={'status': 'contacted', 'admin_notes': ''})
    assert resp.status_code == 302
    assert 'page=2' in resp.headers['Location']
    assert f'q={tag}' in resp.headers['Location']


# ----------------------------- Заявки на курси -----------------------------

def _course_request_env(tag, n=PAGE + 1):
    course = Course(title=f'Курс {tag}', slug=f'pgcr-{tag}', is_active=True)
    db.session.add(course)
    db.session.flush()
    for i, t in enumerate(_times(n)):
        db.session.add(CourseRequest(
            course_id=course.id, email=f'pgcr-{tag}{i:03d}@test.com',
            status='pending', created_at=t,
        ))
    db.session.commit()
    return course


def test_course_requests_pagination_filter_perpage_and_export(client, admin):
    tag = _uid()
    _course_request_env(tag)
    _login(client, admin)
    pattern = re.compile(rf'pgcr-{tag}(\d{{3}})@test\.com')

    page1 = client.get(f'/admin/course-requests?q={tag}').get_data(as_text=True)
    page2 = client.get(f'/admin/course-requests?q={tag}&page=2').get_data(as_text=True)
    idx1, idx2 = _indices(page1, pattern), _indices(page2, pattern)
    assert len(idx1) == PAGE
    assert len(idx2) == 1
    assert idx1.isdisjoint(idx2)
    assert idx1 | idx2 == set(range(PAGE + 1))
    # page=2 не губить активний фільтр САМЕ у посиланнях пейджера
    # (next/prev), а не десь на сторінці взагалі (чіп/експорт теж
    # містять q={tag}, але це не той факт, що перевіряється тут).
    assert f'q={tag}' in _pager_html(page1)
    assert f'q={tag}' in _pager_html(page2)

    p25 = client.get(f'/admin/course-requests?q={tag}&per_page=25').get_data(as_text=True)
    assert len(_indices(p25, pattern)) == 25
    garbage = client.get(f'/admin/course-requests?q={tag}&per_page=99999').get_data(as_text=True)
    assert len(_indices(garbage, pattern)) == PAGE

    r = client.get(f'/admin/course-requests/export?q={tag}')
    wb = load_workbook(io.BytesIO(r.data))
    assert wb['Запити на курси'].max_row - 1 == PAGE + 1
    filter_labels = {c.value for c in wb['Фільтри']['A'] if c.value}
    assert 'Рядків на сторінці' not in filter_labels
    assert 'per_page' not in filter_labels


def test_course_requests_out_of_range_page_offers_way_back(client, admin):
    # БЕЗ q= -- див. докстрінг test_b2b_out_of_range_page_offers_way_back.
    tag = _uid()
    _course_request_env(tag, n=3)
    _login(client, admin)
    html = client.get('/admin/course-requests?page=999').get_data(as_text=True)
    assert html.count(f'pgcr-{tag}') == 0
    assert 'скинути фільтри' in html


def test_course_request_row_action_redirects_back_to_same_page(client, admin):
    tag = _uid()
    _course_request_env(tag)  # PAGE + 1 -- рівно дві сторінки під тегом
    _login(client, admin)
    page2 = client.get(f'/admin/course-requests?q={tag}&page=2').get_data(as_text=True)
    m = re.search(r'action="(/admin/course-requests/\d+/delete\?[^"]*page=2[^"]*)"', page2)
    assert m, 'дію рядка з page=2 у action-URL не знайдено на другій сторінці'
    action_url = m.group(1).replace('&amp;', '&')
    resp = client.post(action_url)
    assert resp.status_code == 302
    assert 'page=2' in resp.headers['Location']
    assert f'q={tag}' in resp.headers['Location']


def test_course_request_edit_redirects_back_to_same_page(client, admin):
    """`course_request_edit` -- окрема сторінка (не інлайн-форма рядка), але
    ту саму пастку відтворює точно так само: посилання «Редагувати» на
    другій сторінці не несло зрізу, форма редагування (без action=, тож
    шле POST на свій же URL) успадковувала цю відсутність, і збереження
    скидало на першу сторінку."""
    tag = _uid()
    _course_request_env(tag)  # PAGE + 1 -- рівно дві сторінки під тегом
    _login(client, admin)
    page2 = client.get(f'/admin/course-requests?q={tag}&page=2').get_data(as_text=True)
    m = re.search(r'href="(/admin/course-requests/\d+/edit\?[^"]*page=2[^"]*)"', page2)
    assert m, 'посилання редагування з page=2 не знайдено на другій сторінці'
    edit_url = m.group(1).replace('&amp;', '&')
    resp = client.post(edit_url, data={'status': 'responded', 'admin_notes': ''})
    assert resp.status_code == 302
    assert 'page=2' in resp.headers['Location']
    assert f'q={tag}' in resp.headers['Location']


def test_course_requests_empty_state_without_has_filters(client, admin):
    """course_requests.html:179 мав саморобний `has_filters`, який нічого не
    керував (empty_state сам рахує зріз через filter_args). Прибраний рядок
    не мав нічого зламати: порожній зріз під фільтром і далі каже "Нічого не
    знайдено", а не 500-ку і не "Запитів немає"."""
    tag = _uid()
    _course_request_env(tag, n=1)
    _login(client, admin)
    r = client.get(f'/admin/course-requests?q={tag}-немає-такого')
    assert r.status_code == 200
    assert 'Нічого не знайдено'.encode() in r.data


# --------------------------- Заявки на повернення ---------------------------

def _refund_row(admin, i, status, created_at, tag):
    # (user_id, online_course_id) унікальна пара -- на кожну заявку свій
    # курс, інакше другий запис того самого власника впав би на constraint.
    course = OnlineCourse(
        sintegrum_id=int(uuid4().int % 10_000_000), slug=f'pgrf-{tag}-{i:03d}',
        remote_name=f'Курс {tag}', remote_status=1,
    )
    db.session.add(course)
    db.session.flush()
    enrollment = OnlineEnrollment(
        user_id=admin.id, online_course_id=course.id,
        payment_amount=Decimal('1000.00'),
    )
    db.session.add(enrollment)
    db.session.flush()
    req = RefundRequest(
        user_id=admin.id, enrollment_id=enrollment.id,
        reason=f'pgrf-{tag}-{i:03d}', status=status, created_at=created_at,
    )
    db.session.add(req)
    return req


def _refund_env(admin, tag, n=PAGE + 1):
    for i, t in enumerate(_times(n)):
        _refund_row(admin, i, STATUS_NEW, t, tag)
    db.session.commit()


def test_refund_requests_pagination_filter_perpage_export_and_new_count(client, admin):
    tag = _uid()
    _refund_env(admin, tag)  # PAGE + 1 -- рівно дві сторінки під тегом
    # Ще дві заявки status=new поза q=tag -- той самий контраргумент, що й
    # у B2B: new_count -- глобальний підрахунок, не розмір відфільтрованого
    # зрізу (pagination.total під самим тегом дав би той самий збіг).
    extra_tag = _uid()
    for i, t in enumerate(_times(2)):
        _refund_row(admin, i, STATUS_NEW, t, extra_tag)
    db.session.commit()
    expected_new_count = RefundRequest.query.filter_by(status=STATUS_NEW).count()
    _login(client, admin)
    pattern = re.compile(rf'pgrf-{tag}-(\d{{3}})')

    page1 = client.get(f'/admin/refund-requests?q={tag}').get_data(as_text=True)
    page2 = client.get(f'/admin/refund-requests?q={tag}&page=2').get_data(as_text=True)
    idx1, idx2 = _indices(page1, pattern), _indices(page2, pattern)
    assert len(idx1) == PAGE
    assert len(idx2) == 1
    assert idx1.isdisjoint(idx2)
    assert idx1 | idx2 == set(range(PAGE + 1))
    # page=2 не губить активний фільтр САМЕ у посиланнях пейджера
    # (next/prev), а не десь на сторінці взагалі (чіп/експорт теж
    # містять q={tag}, але це не той факт, що перевіряється тут).
    assert f'q={tag}' in _pager_html(page1)
    assert f'q={tag}' in _pager_html(page2)

    p25 = client.get(f'/admin/refund-requests?q={tag}&per_page=25').get_data(as_text=True)
    assert len(_indices(p25, pattern)) == 25
    garbage = client.get(f'/admin/refund-requests?q={tag}&per_page=99999').get_data(as_text=True)
    assert len(_indices(garbage, pattern)) == PAGE

    # new_count -- незалежний ГЛОБАЛЬНИЙ запит, не pagination.total.
    n1 = re.search(r'Нових:\s*(\d+)', page1).group(1)
    n2 = re.search(r'Нових:\s*(\d+)', page2).group(1)
    assert n1 == n2 == str(expected_new_count)

    r = client.get(f'/admin/refund-requests/export?q={tag}')
    wb = load_workbook(io.BytesIO(r.data))
    assert wb['Заявки на повернення'].max_row - 1 == PAGE + 1
    filter_labels = {c.value for c in wb['Фільтри']['A'] if c.value}
    assert 'Рядків на сторінці' not in filter_labels
    assert 'per_page' not in filter_labels


def test_refund_requests_out_of_range_page_offers_way_back(client, admin):
    # БЕЗ q= -- див. докстрінг test_b2b_out_of_range_page_offers_way_back.
    tag = _uid()
    _refund_env(admin, tag, n=3)
    _login(client, admin)
    html = client.get('/admin/refund-requests?page=999').get_data(as_text=True)
    assert html.count(f'pgrf-{tag}-') == 0
    assert 'скинути фільтри' in html


def test_refund_row_action_redirects_back_to_same_page(client, admin):
    """Відхилення заявки на третій сторінці черги не має скидати менеджера
    на першу -- та сама пастка, що й з B2B, лише тут ціна помилки вища
    (черга з дедлайном п. 6.3)."""
    tag = _uid()
    _refund_env(admin, tag)  # PAGE + 1 -- рівно дві сторінки під тегом
    _login(client, admin)
    page2 = client.get(f'/admin/refund-requests?q={tag}&page=2').get_data(as_text=True)
    m = re.search(r'action="(/admin/refund-requests/\d+/reject\?[^"]*page=2[^"]*)"', page2)
    assert m, 'дію рядка з page=2 у action-URL не знайдено на другій сторінці'
    action_url = m.group(1).replace('&amp;', '&')
    resp = client.post(action_url, data={'decision_note': 'тестова причина'})
    assert resp.status_code == 302
    assert 'page=2' in resp.headers['Location']
    assert f'q={tag}' in resp.headers['Location']


def test_refund_queue_order_survives_pagination(client, admin):
    """Черга: нові зверху, потім найдавніші -- НЕ "найновіші першими". Новий
    запис (status=new) мусить лишитись першим, навіть якщо він найстаріший
    за датою подання серед усіх рядків зрізу."""
    tag = _uid()
    _refund_row(
        admin, 0, STATUS_NEW,
        datetime(2000, 1, 1, tzinfo=timezone.utc), tag,
    )
    for i in range(1, 4):
        _refund_row(
            admin, i, STATUS_APPROVED,
            datetime(2026, 1, 1, tzinfo=timezone.utc), tag,
        )
    db.session.commit()
    _login(client, admin)

    html = client.get(f'/admin/refund-requests?q={tag}').get_data(as_text=True)
    new_pos = html.index(f'pgrf-{tag}-000')
    later_positions = [html.index(f'pgrf-{tag}-{i:03d}') for i in range(1, 4)]
    assert all(new_pos < p for p in later_positions)


# --------------------------------- Відгуки ---------------------------------

def _review_env(tag, n=PAGE + 1):
    for i, t in enumerate(_times(n)):
        db.session.add(Review(
            author_name='Тест', text=f'pgrv-{tag}-{i:03d} чудовий курс',
            rating=5, is_published=True, sort_order=0, created_at=t,
        ))
    db.session.commit()


def test_reviews_pagination_filter_and_perpage(client, admin):
    tag = _uid()
    _review_env(tag)
    _login(client, admin)
    pattern = re.compile(rf'pgrv-{tag}-(\d{{3}})')

    page1 = client.get(f'/admin/reviews?q={tag}').get_data(as_text=True)
    page2 = client.get(f'/admin/reviews?q={tag}&page=2').get_data(as_text=True)
    idx1, idx2 = _indices(page1, pattern), _indices(page2, pattern)
    assert len(idx1) == PAGE
    assert len(idx2) == 1
    assert idx1.isdisjoint(idx2)
    assert idx1 | idx2 == set(range(PAGE + 1))
    # page=2 не губить активний фільтр САМЕ у посиланнях пейджера
    # (next/prev), а не десь на сторінці взагалі (чіп/експорт теж
    # містять q={tag}, але це не той факт, що перевіряється тут).
    assert f'q={tag}' in _pager_html(page1)
    assert f'q={tag}' in _pager_html(page2)

    p25 = client.get(f'/admin/reviews?q={tag}&per_page=25').get_data(as_text=True)
    assert len(_indices(p25, pattern)) == 25
    garbage = client.get(f'/admin/reviews?q={tag}&per_page=99999').get_data(as_text=True)
    assert len(_indices(garbage, pattern)) == PAGE


def test_reviews_sort_order_survives_pagination(client, admin):
    """sort_order керує порядком публічного блоку на Головній -- і set-based
    перевірка вище ("рівно per_page рядків, без повторів") цього не бачить.

    Два рядки з протилежними sort_order/датою цього не досить: "нижчий
    sort_order = новіша дата" ловить лише мутацію на `created_at.asc()`,
    а `created_at.desc()` (сам по собі, без sort_order) випадково дає той
    самий порядок, що й справжній -- і тест мовчки минає мутанта.

    Тому тут ТРИ рядки з різними sort_order і навмисно переплутаними
    датами: X (sort_order=0, дата СЕРЕДНЯ), Y (sort_order=1, дата
    НАЙДАВНІША), Z (sort_order=2, дата НАЙНОВІША). Справжній порядок --
    X, Y, Z (sort_order рахується першим і сам по собі однозначний, дата
    для tie-break тут узагалі не потрібна). Гола дата ASC дає Y, X, Z;
    гола дата DESC дає Z, X, Y -- жодна не збігається зі справжньою."""
    tag = _uid()
    oldest, middle, newest = _times(3)
    db.session.add(Review(  # X: sort_order=0, дата -- середня
        author_name='Тест', text=f'pgrv-{tag}-x', rating=5,
        is_published=True, sort_order=0, created_at=middle,
    ))
    db.session.add(Review(  # Y: sort_order=1, дата -- найдавніша
        author_name='Тест', text=f'pgrv-{tag}-y', rating=5,
        is_published=True, sort_order=1, created_at=oldest,
    ))
    db.session.add(Review(  # Z: sort_order=2, дата -- найновіша
        author_name='Тест', text=f'pgrv-{tag}-z', rating=5,
        is_published=True, sort_order=2, created_at=newest,
    ))
    db.session.commit()
    _login(client, admin)

    html = client.get(f'/admin/reviews?q={tag}').get_data(as_text=True)
    x_pos = html.index(f'pgrv-{tag}-x')
    y_pos = html.index(f'pgrv-{tag}-y')
    z_pos = html.index(f'pgrv-{tag}-z')
    assert x_pos < y_pos < z_pos


def test_reviews_out_of_range_page_offers_way_back(client, admin):
    # БЕЗ q= -- див. докстрінг test_b2b_out_of_range_page_offers_way_back.
    tag = _uid()
    _review_env(tag, n=3)
    _login(client, admin)
    html = client.get('/admin/reviews?page=999').get_data(as_text=True)
    assert html.count(f'pgrv-{tag}-') == 0
    assert 'скинути фільтри' in html


def test_review_row_action_redirects_back_to_same_page(client, admin):
    tag = _uid()
    _review_env(tag)  # PAGE + 1 -- рівно дві сторінки під тегом
    _login(client, admin)
    page2 = client.get(f'/admin/reviews?q={tag}&page=2').get_data(as_text=True)
    m = re.search(r'action="(/admin/reviews/\d+/toggle\?[^"]*page=2[^"]*)"', page2)
    assert m, 'дію рядка з page=2 у action-URL не знайдено на другій сторінці'
    action_url = m.group(1).replace('&amp;', '&')
    resp = client.post(action_url)
    assert resp.status_code == 302
    assert 'page=2' in resp.headers['Location']
    assert f'q={tag}' in resp.headers['Location']


def test_review_edit_redirects_back_to_same_page(client, admin):
    """`review_edit` -- окрема сторінка (не інлайн-форма рядка), як і
    `course_request_edit`: посилання «Редагувати» на другій сторінці не
    несло зрізу, тоді як «Опублікувати»/«Видалити» в тому самому рядку --
    несли. Форма редагування (без action=) успадковує query-рядок
    власного запиту, тож зріз, якщо він є в href, доїжджає й туди."""
    tag = _uid()
    _review_env(tag)  # PAGE + 1 -- рівно дві сторінки під тегом
    _login(client, admin)
    page2 = client.get(f'/admin/reviews?q={tag}&page=2').get_data(as_text=True)
    m = re.search(r'href="(/admin/reviews/\d+/edit\?[^"]*page=2[^"]*)"', page2)
    assert m, 'посилання редагування з page=2 не знайдено на другій сторінці'
    edit_url = m.group(1).replace('&amp;', '&')
    resp = client.post(edit_url, data={
        'author_name': 'Тест Оновлений', 'text': 'оновлений текст відгуку',
        'rating': '5', 'sort_order': '0', 'course_id': '',
    })
    assert resp.status_code == 302
    assert 'page=2' in resp.headers['Location']
    assert f'q={tag}' in resp.headers['Location']


def test_review_restore_redirects_back_to_same_page(client, admin):
    """Видалення на сторінці 2 правильно повертає на сторінку 2 (уже
    покрито test_review_row_action_redirects_back_to_same_page), але тост
    "Повернути" з видалення досі вів на голий /admin/reviews -- клік по
    ньому саме в тому потоці, яким адмін виправляє власну помилку, скидав
    на першу сторінку. restore_url тепер несе той самий back_args, що й
    редірект після видалення."""
    tag = _uid()
    _review_env(tag)  # PAGE + 1 -- рівно дві сторінки під тегом
    _login(client, admin)
    page2 = client.get(f'/admin/reviews?q={tag}&page=2').get_data(as_text=True)
    m = re.search(r'action="(/admin/reviews/\d+/delete\?[^"]*page=2[^"]*)"', page2)
    assert m, 'дію видалення з page=2 у action-URL не знайдено на другій сторінці'
    delete_url = m.group(1).replace('&amp;', '&')

    after_delete = client.post(delete_url, follow_redirects=True)
    assert after_delete.status_code == 200
    restore_url = _undo_url(after_delete.get_data(as_text=True))
    assert 'page=2' in restore_url
    assert f'q={tag}' in restore_url

    resp = client.post(restore_url)
    assert resp.status_code == 302
    assert 'page=2' in resp.headers['Location']
    assert f'q={tag}' in resp.headers['Location']
