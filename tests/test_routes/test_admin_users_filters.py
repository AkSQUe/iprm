"""Фільтри, пошук і xlsx-експорт списку користувачів (/admin/users).

Сторінка й експорт читають ОДИН набір фільтрів (`_user_filters`), тож тести
перевіряють обидва входи: те, що видно на екрані, і те, що лягає у файл.
"""
from tests.support.rbac import grant_role
import io
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.registration import EventRegistration
from app.models.user import User


def _uid():
    return uuid4().hex[:8]


@pytest.fixture
def admin(app):
    u = User.create_with_password(
        f'usr-admin-{_uid()}@test.com', 'password123',
        first_name='Адмін', last_name='Головний', email_confirmed=True,
    )
    grant_role(u, 'super_admin')
    # Роль пускає в адмінку; прапорець нижче -- для фільтра "role=admin"
    # у списку користувачів, який досі читає стару колонку (Task 8).
    u.is_admin = True
    db.session.flush()
    return u


@pytest.fixture
def people(app, admin):
    """Двоє звичайних: один із реєстрацією і телефоном, другий -- без."""
    with_reg = User.create_with_password(
        f'usr-a-{_uid()}@test.com', 'password123',
        first_name='Оксана', last_name='Шевченко', email_confirmed=True,
    )
    without = User.create_with_password(
        f'usr-b-{_uid()}@test.com', 'password123',
        first_name='Богдан', last_name='Мельник', email_confirmed=False,
    )
    db.session.flush()
    with_reg.medical_profile.phone = '+380671234567'

    course = Course(title='Users Course', slug=f'usr-{_uid()}', is_active=True)
    db.session.add(course)
    db.session.flush()
    inst = CourseInstance(course_id=course.id, status='active',
                          event_format='offline')
    db.session.add(inst)
    db.session.flush()
    db.session.add(EventRegistration(
        user_id=with_reg.id, instance_id=inst.id, phone='+380671234567',
        specialty='T', workplace='Клініка', status='confirmed',
        payment_status='paid',
    ))
    db.session.flush()
    return with_reg, without


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def _export_ids(response):
    wb = load_workbook(io.BytesIO(response.data))
    ws = wb['Користувачі']
    return wb, [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]


def test_search_matches_name_and_filters_out_others(client, admin, people):
    with_reg, without = people
    _login(client, admin)
    r = client.get('/admin/users?q=Шевченко')
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert with_reg.email in html
    assert without.email not in html


def test_search_matches_profile_phone(client, admin, people):
    with_reg, without = people
    _login(client, admin)
    r = client.get('/admin/users?q=380671234567')
    html = r.get_data(as_text=True)
    assert with_reg.email in html
    assert without.email not in html


def test_search_escapes_like_wildcards(client, admin, people):
    """'%' у запиті -- це символ пошуку, а не 'показати всіх'."""
    with_reg, _without = people
    _login(client, admin)
    r = client.get('/admin/users?q=%25')
    assert with_reg.email not in r.get_data(as_text=True)


def test_role_and_registration_filters(client, admin, people):
    with_reg, without = people
    _login(client, admin)

    only_admins = client.get('/admin/users?role=admin').get_data(as_text=True)
    assert admin.email in only_admins
    assert with_reg.email not in only_admins

    with_regs = client.get('/admin/users?regs=with').get_data(as_text=True)
    assert with_reg.email in with_regs
    assert without.email not in with_regs

    unconfirmed = client.get('/admin/users?confirmed=no').get_data(as_text=True)
    assert without.email in unconfirmed
    assert with_reg.email not in unconfirmed


def test_export_honours_filters(client, admin, people):
    with_reg, without = people
    _login(client, admin)
    r = client.get('/admin/users/export?regs=with')
    assert r.status_code == 200
    assert 'spreadsheetml' in r.headers['Content-Type']
    assert 'users-' in r.headers['Content-Disposition']
    wb, ids = _export_ids(r)
    assert with_reg.id in ids
    assert without.id not in ids
    values = {
        row[0].value: row[1].value
        for row in wb['Фільтри'].iter_rows(min_row=2, max_col=2)
    }
    assert values['Реєстрації'] == 'З реєстраціями'
    assert values['Рядків у файлі'] == len(ids)


def test_page_renders_filter_bar_with_export_link(client, admin, people):
    _login(client, admin)
    html = client.get('/admin/users?role=admin').get_data(as_text=True)
    assert 'name="q"' in html
    assert '/admin/users/export?' in html
    assert 'role=admin' in html


def test_unknown_filter_value_is_ignored(client, admin, people):
    """Сміття в query-string не має ані падати, ані ховати список."""
    with_reg, without = people
    _login(client, admin)
    html = client.get('/admin/users?role=%3Cdrop%3E').get_data(as_text=True)
    assert with_reg.email in html
    assert without.email in html
