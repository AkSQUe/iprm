"""Фільтри, пошук і xlsx-експорт на списках адмінки.

Усі ці сторінки ділять один інструментарій (`app/admin/_listing.py` +
макрос `admin/partials/_filter_bar.html`), тож тест ганяє їх однаково:
сторінка звужується так само, як експорт, а сміття в query-string нічого
не ламає.
"""
import io
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.models.b2b_request import B2BRequest
from app.models.certificate import Certificate
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.course_request import CourseRequest
from app.models.email_log import EmailLog
from app.models.error_log import ErrorLog
from app.models.promo_code import PromoCode
from app.models.registration import EventRegistration
from app.models.review import Review
from app.models.user import User


def _uid():
    return uuid4().hex[:8]


@pytest.fixture
def admin(app):
    u = User.create_with_password(
        f'lst-{_uid()}@test.com', 'password123',
        first_name='L', last_name='A', is_admin=True, email_confirmed=True,
    )
    # Саме commit, а не flush: сторінка журналу помилок починається із
    # захисного db.session.rollback(), який зніс би незакомічені рядки.
    db.session.commit()
    yield u
    db.session.rollback()
    db.session.delete(db.session.merge(u))
    db.session.commit()


@pytest.fixture
def data(app, admin):
    """Мінімальний набір: курс, проведення, реєстрація, сертифікат тощо.

    Дані комітяться (див. фікстуру admin) і тому прибираються вручну після
    тесту -- інакше опубліковане проведення потрапляло б у чужі вибірки.
    """
    created = []
    course = Course(title='Плазмотерапія в дерматології', slug=f'lst-{_uid()}',
                    is_active=True)
    db.session.add(course)
    db.session.flush()
    # Місто унікальне: сесія тестів ділить одну БД (conftest не відкочує
    # коміти), тож пошук за "Львів" бачив би і заходи з інших тестів.
    inst = CourseInstance(course_id=course.id, status='published',
                          event_format='offline', location=f'Львів-{_uid()}')
    db.session.add(inst)
    db.session.flush()
    reg = EventRegistration(
        user_id=admin.id, instance_id=inst.id, phone='+380671110001',
        specialty='T', workplace='Клініка', status='confirmed',
        payment_status='paid',
    )
    db.session.add(reg)
    db.session.flush()
    cert = Certificate(
        registration_id=reg.id, user_id=admin.id, number=f'2026-0001-{_uid()}',
        recipient_name='Шевченко Оксана', event_title='Плазмотерапія',
        pdf_path='x.pdf', cpd_points=12,
    )
    # Контрольний код без "%" в описі -- для перевірки екранування LIKE.
    plain_code = f'PLAIN{_uid()[:4]}'.upper()
    extras = [
        cert,
        B2BRequest(
            first_name='Іван', last_name='Мороз', phone='+380671110002',
            email=f'b2b-{_uid()}@test.com', team_size='6-10', status='new',
        ),
        CourseRequest(
            course_id=course.id, email=f'creq-{_uid()}@test.com',
            phone='+380671110003', status='pending', message='Коли набір?',
        ),
        Review(author_name='Оксана Коваль', text='Чудово',
               rating=5, is_published=True, course_id=course.id),
        PromoCode(
            code=f'SUMMER{_uid()[:4]}'.upper(), code_norm=f'summer{_uid()[:4]}',
            description='Літня знижка 50%', discount_type='percent',
            discount_value=10, is_active=True,
        ),
        PromoCode(
            code=plain_code, code_norm=plain_code.lower(),
            description='Осіння знижка', discount_type='amount',
            discount_value=200, is_active=True,
        ),
        EmailLog(
            to_email='fail@test.com', subject='Реєстрація підтверджена',
            template_name='registration', status='failed',
            trigger='registration', error_message='SMTP 535 auth failed',
        ),
        ErrorLog(
            error_code=404, error_type='NotFound', error_message='page missing',
            url='/courses/missing', method='GET',
        ),
    ]
    db.session.add_all(extras)
    db.session.commit()
    created.extend(extras + [reg, inst, course])

    yield {'course': course, 'instance': inst, 'reg': reg, 'cert': cert,
           'plain_promo': plain_code}

    db.session.rollback()
    for obj in created:
        db.session.delete(db.session.merge(obj))
    db.session.commit()


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def _sheet_rows(response, sheet):
    wb = load_workbook(io.BytesIO(response.data))
    ws = wb[sheet]
    return wb, ws.max_row - 1  # без рядка заголовків


# Сторінка -> (URL списку, URL експорту, назва аркуша, запит, що щось знайде,
#              запит, що не знайде нічого).
LISTINGS = [
    ('certificates', '/admin/certificates', '/admin/certificates/export',
     'Сертифікати', 'Шевченко', 'зовсім-інше'),
    ('b2b', '/admin/b2b-requests', '/admin/b2b-requests/export',
     'B2B-заявки', 'Мороз', 'зовсім-інше'),
    ('course_requests', '/admin/course-requests', '/admin/course-requests/export',
     'Запити на курси', 'набір', 'зовсім-інше'),
    ('promo', '/admin/promo-codes', '/admin/promo-codes/export',
     'Промокоди', 'Літня', 'зовсім-інше'),
    ('email_log', '/admin/notifications/log', '/admin/notifications/log/export',
     'Журнал листів', 'fail@test.com', 'зовсім-інше'),
    ('error_log', '/admin/error-logs', '/admin/error-logs/export',
     'Помилки', '/courses/missing', 'зовсім-інше'),
    ('instances', '/admin/instances', '/admin/instances/report.xlsx',
     'Проведення', 'Львів', 'зовсім-інше'),
]


@pytest.mark.parametrize('name,page_url,export_url,sheet,hit,miss', LISTINGS)
def test_search_narrows_page_and_export(client, admin, data, name, page_url,
                                        export_url, sheet, hit, miss):
    _login(client, admin)
    assert client.get(page_url).status_code == 200

    found = client.get(f'{export_url}?q={hit}')
    assert found.status_code == 200
    assert 'spreadsheetml' in found.headers['Content-Type']
    _wb, hit_rows = _sheet_rows(found, sheet)
    assert hit_rows >= 1

    empty = client.get(f'{export_url}?q={miss}')
    _wb, miss_rows = _sheet_rows(empty, sheet)
    assert miss_rows == 0


@pytest.mark.parametrize('name,page_url,export_url,sheet,hit,miss', LISTINGS)
def test_export_documents_filters(client, admin, data, name, page_url,
                                  export_url, sheet, hit, miss):
    """Кожен експорт несе аркуш «Фільтри» з лічильником рядків."""
    _login(client, admin)
    r = client.get(f'{export_url}?q={hit}')
    wb, rows = _sheet_rows(r, sheet)
    assert 'Фільтри' in wb.sheetnames
    values = {
        row[0].value: row[1].value
        for row in wb['Фільтри'].iter_rows(min_row=2, max_col=2)
    }
    assert values['Пошук'] == hit
    assert values['Рядків у файлі'] == rows


@pytest.mark.parametrize('name,page_url,export_url,sheet,hit,miss', LISTINGS)
def test_garbage_filter_value_is_ignored(client, admin, data, name, page_url,
                                         export_url, sheet, hit, miss):
    """Невідомий статус зі старого посилання не має давати 500."""
    _login(client, admin)
    assert client.get(f'{page_url}?status=%3Cdrop%3E').status_code == 200


def test_search_escapes_like_wildcards(client, admin, data):
    """'%' у запиті шукається дослівно, а не як шаблон LIKE."""
    _login(client, admin)
    r = client.get('/admin/promo-codes/export?q=%25')
    wb, rows = _sheet_rows(r, 'Промокоди')
    assert rows >= 1
    # Код з описом "50%" знайдено, а контрольний без "%" -- ні: якби '%'
    # лишився шаблоном LIKE, у вибірку потрапили б геть усі коди.
    dump = '\n'.join(
        str(c.value) for row in wb['Промокоди'].iter_rows() for c in row
    )
    assert 'Літня знижка 50%' in dump
    assert data['plain_promo'] not in dump


def test_certificates_state_and_year_filters(client, admin, data):
    _login(client, admin)
    number = data['cert'].number
    assert number in client.get('/admin/certificates?state=valid').get_data(as_text=True)
    assert number not in client.get('/admin/certificates?state=revoked').get_data(as_text=True)
    assert number not in client.get('/admin/certificates?year=1999').get_data(as_text=True)


def test_reviews_filters(client, admin, data):
    _login(client, admin)
    published = client.get('/admin/reviews?state=published').get_data(as_text=True)
    assert 'Оксана Коваль' in published
    drafts = client.get('/admin/reviews?state=draft').get_data(as_text=True)
    assert 'Оксана Коваль' not in drafts
    assert 'Оксана Коваль' not in client.get('/admin/reviews?rating=1').get_data(as_text=True)


def test_instance_registrations_search(client, admin, data):
    """Пошук у списку учасників заходу; лічильники лишаються по всьому заходу."""
    _login(client, admin)
    url = f"/admin/instances/{data['instance'].id}/registrations"
    html = client.get(f'{url}?q=380671110001').get_data(as_text=True)
    assert admin.email in html
    missing = client.get(f'{url}?q=нікого-такого').get_data(as_text=True)
    assert admin.email not in missing
    assert 'Нічого не знайдено' in missing


def test_instances_report_counts_registrations(client, admin, data):
    _login(client, admin)
    r = client.get(f"/admin/instances/report.xlsx?q={data['instance'].location}")
    wb, rows = _sheet_rows(r, 'Проведення')
    assert rows == 1
    ws = wb['Проведення']
    headers = [c.value for c in ws[1]]
    assert ws.cell(row=2, column=headers.index('Реєстрацій') + 1).value == 1
