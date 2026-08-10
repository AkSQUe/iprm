"""Захисні гарантії модуля списків: санітація файлу, ліміти, пресети.

Ці перевірки описують саме те, чим експорт може нашкодити: чужий текст, що
керує вмістом файлу; зріз, який не влазить у синхронний запит; параметри з
чужого посилання.
"""
import io
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.admin import _listing
from app.extensions import db
from app.models.b2b_request import B2BRequest
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.registration import EventRegistration
from app.models.user import User


def _uid():
    return uuid4().hex[:8]


@pytest.fixture
def admin(app):
    u = User.create_with_password(
        f'hd-{_uid()}@test.com', 'password123',
        first_name='H', last_name='D', is_admin=True, email_confirmed=True,
    )
    db.session.flush()
    return u


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


# --- xlsx: чужий текст не має керувати файлом -------------------------------

def test_control_character_does_not_break_export(app):
    """Один \\x07 у нотатці валив увесь експорт (IllegalCharacterError)."""
    from app.services import xlsx_reports

    req = _fake_b2b(admin_notes='нотатка\x07з керуючим символом')
    wb = load_workbook(io.BytesIO(
        xlsx_reports.export_b2b_requests_xlsx([req]).getvalue()))
    assert wb['B2B-заявки'].cell(row=2, column=8).value == 'нотатказ керуючим символом'


def test_formula_text_is_written_as_text(app):
    """Текст із '=' не має ставати живою формулою у файлі менеджера."""
    from app.services import xlsx_reports

    evil = '=HYPERLINK("http://evil","клік")'
    req = _fake_b2b(last_name=evil, first_name='#N/A')
    ws = load_workbook(io.BytesIO(
        xlsx_reports.export_b2b_requests_xlsx([req]).getvalue()))['B2B-заявки']
    assert ws.cell(row=2, column=2).data_type == 's'
    assert ws.cell(row=2, column=2).value == evil
    assert ws.cell(row=2, column=3).data_type == 's'


def test_overlong_text_is_truncated_to_cell_limit(app):
    from app.services import xlsx_reports
    from app.services.xlsx_io import MAX_CELL_LENGTH

    req = _fake_b2b(admin_notes='x' * (MAX_CELL_LENGTH + 5000))
    ws = load_workbook(io.BytesIO(
        xlsx_reports.export_b2b_requests_xlsx([req]).getvalue()))['B2B-заявки']
    assert len(ws.cell(row=2, column=8).value) == MAX_CELL_LENGTH


class _FakeB2B:
    """Заглушка замість рядка БД: перевіряємо саме побудову файлу."""

    def __init__(self, **kw):
        self.created_at = None
        self.last_name = 'Мороз'
        self.first_name = 'Іван'
        self.phone = '+380670000000'
        self.email = 'b2b@test.com'
        self.team_size_label = '6-10'
        self.status_label = 'Новий'
        self.admin_notes = ''
        self.__dict__.update(kw)


def _fake_b2b(**kw):
    return _FakeB2B(**kw)


# --- ліміти й параметри -----------------------------------------------------

def test_export_refuses_oversized_slice(client, admin, monkeypatch):
    """Понад стелю віддаємо відмову з поясненням, а не мовчазний обрізок."""
    monkeypatch.setattr(_listing, 'MAX_EXPORT_ROWS', 0)
    _login(client, admin)
    db.session.add(B2BRequest(
        first_name='Іван', last_name='Мороз', phone='+380670000001',
        email=f'b2b-{_uid()}@test.com', team_size='6-10', status='new',
    ))
    db.session.flush()

    r = client.get('/admin/b2b-requests/export', follow_redirects=True)
    assert r.status_code == 200
    assert 'більше за ліміт' in r.get_data(as_text=True)


def test_export_failure_does_not_return_500(client, admin, monkeypatch):
    from app.services import xlsx_reports

    def boom(*args, **kwargs):
        raise RuntimeError('зламався openpyxl')

    monkeypatch.setattr(xlsx_reports, 'export_b2b_requests_xlsx', boom)
    _login(client, admin)
    r = client.get('/admin/b2b-requests/export', follow_redirects=True)
    assert r.status_code == 200
    assert 'Не вдалося сформувати файл' in r.get_data(as_text=True)


def test_zero_id_is_not_treated_as_active_filter(client, admin):
    """?course_id=0 не має малювати чіпс фільтра, якого запит не застосовує."""
    _login(client, admin)
    html = client.get('/admin/users?course_id=0').get_data(as_text=True)
    assert 'admin-filters__count' not in html


def test_search_term_is_capped(app):
    with app.test_request_context('/admin/users?q=' + 'я' * 500):
        assert len(_listing.text_arg('q')) == 500
        clause = _listing.search_clause(_listing.text_arg('q'), [User.email])
        # У шаблон LIKE потрапляє обрізаний рядок + два '%'.
        rendered = str(clause.compile(compile_kwargs={'literal_binds': True}))
        assert 'я' * (_listing.MAX_SEARCH_LENGTH + 1) not in rendered


def test_inverted_date_range_is_swapped(client, admin):
    """Переставлені межі дають той самий зріз, а не порожній список."""
    course = Course(title='Дати', slug=f'dt-{_uid()}', is_active=True)
    db.session.add(course)
    db.session.flush()
    inst = CourseInstance(course_id=course.id, status='published',
                          event_format='offline')
    db.session.add(inst)
    db.session.flush()
    db.session.add(EventRegistration(
        user_id=admin.id, instance_id=inst.id, phone='+380670000002',
        specialty='T', workplace='T', status='confirmed', payment_status='paid',
    ))
    db.session.flush()

    _login(client, admin)
    today = _listing.now_kyiv().strftime('%Y-%m-%d')
    straight = client.get(
        f'/admin/registrations?scope=all&date_from={today}&date_to={today}')
    inverted = client.get(
        f'/admin/registrations?scope=all&date_from={today}&date_to=2020-01-01')
    assert admin.email in straight.get_data(as_text=True)
    assert admin.email in inverted.get_data(as_text=True)


# --- пресети ----------------------------------------------------------------

def test_presets_are_plain_filter_links(client, admin):
    """Пресет -- це набір звичайних параметрів, а не окрема гілка запиту."""
    _login(client, admin)
    html = client.get('/admin/registrations?scope=all').get_data(as_text=True)
    assert 'payment=unpaid' in html
    assert 'no_certificate=1' in html

    active = client.get(
        '/admin/registrations?scope=all&payment=unpaid&status=confirmed'
    ).get_data(as_text=True)
    # Активний пресет підсвічено, а його складові видно чіпсами.
    assert 'admin-pill--active' in active
    assert active.count('class="admin-chip"') >= 2


def test_no_certificate_preset_filters_rows(client, admin):
    course = Course(title='Пресет', slug=f'ps-{_uid()}', is_active=True)
    db.session.add(course)
    db.session.flush()
    inst = CourseInstance(course_id=course.id, status='published',
                          event_format='offline')
    db.session.add(inst)
    db.session.flush()
    reg = EventRegistration(
        user_id=admin.id, instance_id=inst.id, phone='+380670000003',
        specialty='T', workplace='T', status='completed', payment_status='paid',
    )
    db.session.add(reg)
    db.session.flush()

    _login(client, admin)
    without = client.get(
        '/admin/registrations?scope=all&status=completed&no_certificate=1'
    ).get_data(as_text=True)
    assert admin.email in without

    from app.models.certificate import Certificate
    db.session.add(Certificate(
        registration_id=reg.id, user_id=admin.id, number=f'PS-{_uid()}',
        recipient_name='Тест', event_title='Пресет', pdf_path='x.pdf',
    ))
    db.session.flush()
    still = client.get(
        '/admin/registrations?scope=all&status=completed&no_certificate=1'
    ).get_data(as_text=True)
    assert admin.email not in still
