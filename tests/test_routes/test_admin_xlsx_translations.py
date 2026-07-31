"""Роути xlsx-каналу перекладів: експорт, прев'ю, застосування, скасування."""
import io
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models.course import Course
from app.models.user import User
from app.services import translation_registry as registry
from app.services import xlsx_translations as xt

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@pytest.fixture
def admin():
    u = User.create_with_password(
        f'a-{uuid4().hex[:6]}@test.com', 'password123',
        first_name='A', last_name='D', is_admin=True, email_confirmed=True,
    )
    db.session.commit()
    return u


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def _course():
    c = Course(title=f'Курс {uuid4().hex[:4]}', slug=f'r-{uuid4().hex[:6]}',
               is_active=True)
    db.session.add(c)
    db.session.commit()
    return c


def _upload_bytes(course, **langs):
    """Мінімальний файл перекладів на одну назву курсу."""
    unit = {u.uid: u for u in registry.units(course)}['title']
    wb = Workbook()
    ws = wb.active
    ws.title = 'Курси'
    cols = xt.translation_cols()
    labels = xt.translation_labels()
    ws.append([labels[c] for c in cols])
    row = {'entity': 'course', 'id': course.id, 'object': course.title,
           'field': unit.label, 'key': 'title', 'uk': unit.source, **langs}
    ws.append([row.get(c, '') for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --- доступ -----------------------------------------------------------------

def test_export_requires_admin(client):
    assert client.get('/admin/translations/export/courses').status_code in (302, 401, 403)


def test_import_requires_admin(client):
    r = client.post('/admin/translations/import/courses')
    assert r.status_code in (302, 401, 403)


def test_unknown_scope_404(client, admin):
    _login(client, admin)
    assert client.get('/admin/translations/export/nope').status_code == 404


# --- експорт ----------------------------------------------------------------

@pytest.mark.parametrize('scope', ['courses', 'trainers', 'instances'])
def test_export_returns_xlsx(client, admin, scope):
    _login(client, admin)
    _course()
    r = client.get(f'/admin/translations/export/{scope}')
    assert r.status_code == 200
    assert r.mimetype == XLSX_MIME
    assert f'translations-{scope}' in r.headers['Content-Disposition']
    wb = load_workbook(io.BytesIO(r.data))
    assert 'Довідка' in wb.sheetnames


def test_export_todo_mode_marks_filename(client, admin):
    _login(client, admin)
    r = client.get('/admin/translations/export/courses?mode=todo')
    assert 'translations-courses-todo' in r.headers['Content-Disposition']


# --- імпорт -----------------------------------------------------------------

def test_import_rejects_non_xlsx(client, admin):
    _login(client, admin)
    r = client.post('/admin/translations/import/courses',
                    data={'xlsx': (io.BytesIO(b'x'), 'notes.txt')},
                    content_type='multipart/form-data')
    assert r.status_code == 302
    assert '/admin/courses' in r.headers['Location']


def test_import_without_file_redirects_back(client, admin):
    _login(client, admin)
    r = client.post('/admin/translations/import/courses',
                    data={}, content_type='multipart/form-data')
    assert r.status_code == 302


def test_full_import_flow(client, admin):
    _login(client, admin)
    course = _course()

    upload = client.post(
        '/admin/translations/import/courses',
        data={'xlsx': (_upload_bytes(course, ru='Курс-РУ'), 'tr.xlsx')},
        content_type='multipart/form-data',
    )
    assert upload.status_code == 302
    preview_url = upload.headers['Location']
    assert '/preview/' in preview_url

    preview = client.get(preview_url)
    assert preview.status_code == 200
    assert 'Курс-РУ' in preview.get_data(as_text=True)
    # Прев'ю нічого не змінює.
    assert db.session.get(Course, course.id).t('title', lang='ru') == course.title

    token = preview_url.rstrip('/').split('/')[-1]
    applied = client.post(f'/admin/translations/import/courses/apply/{token}')
    assert applied.status_code == 302
    assert db.session.get(Course, course.id).t('title', lang='ru') == 'Курс-РУ'


def test_cancel_discards_upload(client, admin):
    _login(client, admin)
    course = _course()
    upload = client.post(
        '/admin/translations/import/courses',
        data={'xlsx': (_upload_bytes(course, ru='Курс-РУ'), 'tr.xlsx')},
        content_type='multipart/form-data',
    )
    token = upload.headers['Location'].rstrip('/').split('/')[-1]

    r = client.post(f'/admin/translations/import/courses/cancel/{token}')
    assert r.status_code == 302
    # Після скасування токен більше не резолвиться.
    assert client.get(
        f'/admin/translations/import/courses/preview/{token}'
    ).status_code == 302
    assert db.session.get(Course, course.id).t('title', lang='ru') == course.title


def test_stale_token_redirects_with_message(client, admin):
    _login(client, admin)
    r = client.get('/admin/translations/import/courses/preview/deadbeef')
    assert r.status_code == 302
    assert '/admin/courses' in r.headers['Location']
