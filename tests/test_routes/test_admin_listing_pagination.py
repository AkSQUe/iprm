"""Пагінація довгих реєстрів + пошук на довідникових списках адмінки.

Сертифікати й користувачі раніше віддавали весь список одним шматком. Тепер
сторінками, і тут перевіряється головне: сторінка ріже вибірку, гортання не
губить фільтр, а експорт лишається по ВСЬОМУ зрізу.
"""
from tests.support.rbac import grant_role
import io
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.models.blog_post import BlogPost
from app.models.certificate import Certificate
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.registration import EventRegistration
from app.models.trainer import Trainer
from app.models.user import User


def _uid():
    return uuid4().hex[:8]


@pytest.fixture
def admin(app):
    u = User.create_with_password(
        f'pag-{_uid()}@test.com', 'password123',
        first_name='P', last_name='A', email_confirmed=True,
    )
    grant_role(u, 'super_admin')
    db.session.flush()
    return u


@pytest.fixture
def many_certificates(app, admin):
    """PER_PAGE + 3 сертифікати одного власника -- рівно щоб було 2 сторінки."""
    from app.admin._listing import LIST_PER_PAGE

    course = Course(title='Пагінація', slug=f'pag-{_uid()}', is_active=True)
    db.session.add(course)
    db.session.flush()

    marker = f'ПАГ{_uid()}'
    total = LIST_PER_PAGE + 3
    instance_ids = []
    for i in range(total):
        # Одне проведення на сертифікат: (user_id, instance_id) у реєстраціях
        # унікальна пара, тож 53 сертифікати одного власника вимагають 53 заходи.
        inst = CourseInstance(course_id=course.id, status='published',
                              event_format='offline')
        db.session.add(inst)
        db.session.flush()
        instance_ids.append(inst.id)
        reg = EventRegistration(
            user_id=admin.id, instance_id=inst.id, phone='+380670000000',
            specialty='T', workplace='T', status='confirmed',
            payment_status='paid',
        )
        db.session.add(reg)
        db.session.flush()
        db.session.add(Certificate(
            registration_id=reg.id, user_id=admin.id,
            number=f'{marker}-{i:04d}', recipient_name=marker,
            event_title='Пагінація', pdf_path='x.pdf',
        ))
    db.session.commit()
    yield {'marker': marker, 'total': total, 'per_page': LIST_PER_PAGE}
    # Прибираємо за собою: conftest коміти не відкочує.
    Certificate.query.filter(Certificate.recipient_name == marker).delete(
        synchronize_session=False)
    EventRegistration.query.filter(
        EventRegistration.instance_id.in_(instance_ids)).delete(
        synchronize_session=False)
    CourseInstance.query.filter(CourseInstance.id.in_(instance_ids)).delete(
        synchronize_session=False)
    db.session.commit()


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def test_certificates_page_is_capped(client, admin, many_certificates):
    marker = many_certificates['marker']
    _login(client, admin)
    first = client.get(f'/admin/certificates?q={marker}').get_data(as_text=True)
    assert first.count(f'{marker}-') <= many_certificates['per_page']
    # Друга сторінка несе решту й зберігає пошук у посиланнях.
    second = client.get(f'/admin/certificates?q={marker}&page=2').get_data(as_text=True)
    assert f'{marker}-' in second
    assert f'q={marker}' in first


def test_certificates_export_covers_whole_slice(client, admin, many_certificates):
    """Пагінація -- властивість екрана: у файл ідуть усі рядки зрізу."""
    marker = many_certificates['marker']
    _login(client, admin)
    r = client.get(f'/admin/certificates/export?q={marker}')
    wb = load_workbook(io.BytesIO(r.data))
    assert wb['Сертифікати'].max_row - 1 == many_certificates['total']


def test_users_pagination_keeps_filters(client, admin):
    _login(client, admin)
    html = client.get('/admin/users?role=admin').get_data(as_text=True)
    assert 'role=admin' in html
    assert client.get('/admin/users?role=admin&page=2').status_code == 200


@pytest.fixture
def catalog(app):
    """Неактивні курс, тренер і чернетка допису.

    Назви з унікальним суфіксом: інші модулі лишають по собі закомічені
    записи з тими самими іменами (conftest коміти не відкочує), і без
    суфікса перевірка "не потрапив у протилежний фільтр" ловила б їх.
    """
    tag = _uid()
    course = Course(title=f'Плазмоліфтинг обличчя {tag}', slug=f'cat-{tag}',
                    is_active=False)
    trainer = Trainer(full_name=f'Гусак Валерія {tag}', slug=f'tr-{tag}',
                      is_active=False)
    post = BlogPost(title=f'Як обрати курс {tag}', slug=f'post-{tag}',
                    content=[], status=BlogPost.STATUS_DRAFT)
    db.session.add_all([course, trainer, post])
    db.session.flush()
    return {'course': course.title, 'trainer': trainer.full_name,
            'post': post.title}


@pytest.mark.parametrize('url,key,hit,miss', [
    ('/admin/courses', 'course', 'q=Плазмоліфтинг', 'q=нічого-такого'),
    ('/admin/trainers', 'trainer', 'q=Гусак', 'q=нічого-такого'),
    ('/admin/blog', 'post', 'q=обрати', 'q=нічого-такого'),
])
def test_catalog_search(client, admin, catalog, url, key, hit, miss):
    _login(client, admin)
    needle = catalog[key]
    assert needle in client.get(f'{url}?{hit}').get_data(as_text=True)
    assert needle not in client.get(f'{url}?{miss}').get_data(as_text=True)


@pytest.mark.parametrize('url,key,state_hit,state_miss', [
    ('/admin/courses', 'course', 'state=inactive', 'state=active'),
    ('/admin/trainers', 'trainer', 'state=inactive', 'state=active'),
    ('/admin/blog', 'post', 'state=draft', 'state=published'),
])
def test_catalog_state_filter(client, admin, catalog, url, key,
                              state_hit, state_miss):
    _login(client, admin)
    needle = catalog[key]
    assert needle in client.get(f'{url}?{state_hit}').get_data(as_text=True)
    assert needle not in client.get(f'{url}?{state_miss}').get_data(as_text=True)
