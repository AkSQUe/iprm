"""XLSX-експорт списку реєстрацій з урахуванням активних фільтрів.

Головне, що тут перевіряється: файл містить рівно той зріз, який менеджер
бачить на екрані (ті самі фільтри), і не обмежений поточною сторінкою.
"""
import io
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.registration import EventRegistration
from app.models.user import User


def _uid():
    return uuid4().hex[:8]


@pytest.fixture
def admin(app):
    u = User.create_with_password(
        f'exp-{_uid()}@test.com', 'password123',
        first_name='E', last_name='A', is_admin=True, email_confirmed=True,
    )
    db.session.flush()
    return u


@pytest.fixture
def regs(app, admin):
    """Дві реєстрації на один захід: підтверджена+оплачена і очікує."""
    course = Course(title='Export Course', slug=f'exp-{_uid()}', is_active=True)
    db.session.add(course)
    db.session.flush()
    inst = CourseInstance(course_id=course.id, status='active',
                          event_format='offline', location='Київ')
    db.session.add(inst)
    db.session.flush()
    paid = EventRegistration(
        user_id=admin.id, instance_id=inst.id, phone='+380000000001',
        specialty='T', workplace='Клініка', status='confirmed',
        payment_status='paid', payment_method='liqpay', payment_amount=7500,
    )
    other = User.create_with_password(
        f'exp2-{_uid()}@test.com', 'password123',
        first_name='P', last_name='B', email_confirmed=True,
    )
    db.session.flush()
    pending = EventRegistration(
        user_id=other.id, instance_id=inst.id, phone='+380000000002',
        specialty='T', workplace='Клініка', status='pending',
        payment_status='unpaid',
    )
    db.session.add_all([paid, pending])
    db.session.flush()
    return paid, pending


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def _rows(response):
    """Значення першої колонки (ID реєстрації) з аркуша «Реєстрації»."""
    wb = load_workbook(io.BytesIO(response.data))
    ws = wb['Реєстрації']
    return wb, [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]


def test_export_returns_xlsx_with_all_rows(client, admin, regs):
    paid, pending = regs
    _login(client, admin)
    r = client.get('/admin/registrations/export?scope=all')
    assert r.status_code == 200
    assert 'spreadsheetml' in r.headers['Content-Type']
    assert 'registrations-' in r.headers['Content-Disposition']
    _wb, ids = _rows(r)
    assert paid.id in ids and pending.id in ids


def test_export_honours_filters(client, admin, regs):
    paid, pending = regs
    _login(client, admin)
    r = client.get(
        '/admin/registrations/export'
        '?scope=all&status=confirmed&payment=paid&payment_method=liqpay'
    )
    assert r.status_code == 200
    _wb, ids = _rows(r)
    # Рядки з попередніх тестів сесії теж можуть бути у вибірці -- перевіряємо
    # саме те, що фільтр відсіює невідповідний статус/оплату.
    assert paid.id in ids
    assert pending.id not in ids


def test_export_documents_applied_filters(client, admin, regs):
    """Аркуш «Фільтри» пояснює, який саме зріз лежить у файлі."""
    _login(client, admin)
    r = client.get('/admin/registrations/export?scope=all&status=confirmed')
    wb, ids = _rows(r)
    assert 'Фільтри' in wb.sheetnames
    values = {
        row[0].value: row[1].value
        for row in wb['Фільтри'].iter_rows(min_row=2, max_col=2)
    }
    assert values['Статус'] == 'Підтверджено'
    assert values['Оплата'] == 'Усі'
    assert values['Рядків у файлі'] == len(ids)


def test_search_filters_page_and_export(client, admin, regs):
    """Пошук по учаснику однаково звужує і сторінку, і файл."""
    paid, pending = regs
    _login(client, admin)
    page = client.get('/admin/registrations?scope=all&q=380000000001')
    html = page.get_data(as_text=True)
    assert paid.user.email in html

    r = client.get('/admin/registrations/export?scope=all&q=380000000001')
    _wb, ids = _rows(r)
    assert ids == [paid.id]


def test_list_page_links_to_filtered_export(client, admin, regs):
    _login(client, admin)
    r = client.get('/admin/registrations?scope=all&status=confirmed')
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert '/admin/registrations/export?' in html
    assert 'status=confirmed' in html
