"""Стеля синхронного експорту рахується COUNT-ом ДО вибірки, а не після.

Дванадцять роутів переведено на `_listing.export_query`, який робить
`COUNT(*)` і лише коли зріз влазить у стелю -- матеріалізує рядки. Тест,
що перевіряє тільки «прийшов редірект», проходив би і на старому коді
(`rows = query.all()`, а стелю міряв уже `xlsx_export` по `len(rows)`) --
відмова та сама, просто пізніша. Доводимо саме ПОРЯДОК ДІЙ: коли зріз
перевищує стелю, у журналі SQL є `COUNT`, і НЕМАЄ жодного іншого SELECT --
бо `query.all()` на старому коді виконався б завжди, незалежно від стелі,
і саме цей запит СТАВ БИ видимим у журналі, якого тут бути не повинно.

Дванадцять роутів зведено в один список `ROUTE_SPECS`, а не продубльовано:
кожен запис несе своє посилання на список, посилання на xlsx-білдер і
фабрику рядків під власну модель.
"""
import io
import re
from datetime import datetime, timedelta, timezone
from urllib.parse import parse_qs, urlparse
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import event

from app.admin import _listing
from app.extensions import db
from app.models.b2b_request import B2BRequest
from app.models.certificate import Certificate
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.course_request import CourseRequest
from app.models.email_log import EmailLog
from app.models.error_log import ErrorLog
from app.models.online_course import OnlineCourse
from app.models.online_enrollment import OnlineEnrollment
from app.models.promo_code import PromoCode
from app.models.referral_reward import ReferralReward
from app.models.refund_request import RefundRequest
from app.models.registration import EventRegistration
from app.models.user import User
from app.services import xlsx_reports


def _uid():
    return uuid4().hex[:8]


@pytest.fixture
def admin(app):
    u = User.create_with_password(
        f'ec-{_uid()}@test.com', 'password123',
        first_name='Е', last_name='С', is_admin=True, email_confirmed=True,
    )
    db.session.flush()
    return u


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


# SQLAlchemy Query.count() завжди огортає весь запит підзапитом:
# "SELECT count(*) AS count_1 FROM (<весь запит>) AS anon_1" -- рядковий
# SELECT так НЕ починається, навіть коли сам несе count() усередині
# корельованого підзапиту (як у users).
_COUNT_SQL_RE = re.compile(r'^select\s+count\(\*\)', re.IGNORECASE)

# `col.ilike(pattern, escape='\\')` на діалекті SQLite нема нативного ILIKE
# -- компілятор розгортає його в "lower(col) LIKE lower(?) ESCAPE '\'".
# Рядка 'ilike' у зібраних SQL НІКОЛИ не буде: ознака справжнього запиту
# зрізу (усі дванадцять специфікацій завжди передають `q=tag`, який
# apply_search перетворює саме на цю форму) -- ця LIKE-конструкція.
_LIKE_LOWER_RE = re.compile(r'like\s+lower\(', re.IGNORECASE)


# --- Спільні фабрики для рядків, що тримають FK-ланцюжок --------------------

def _course(tag):
    c = Course(title=f'Курс {tag}', slug=f'ec-{tag}-{_uid()}', is_active=True)
    db.session.add(c)
    db.session.flush()
    return c


def _instance(course):
    inst = CourseInstance(
        course_id=course.id, status='published', event_format='offline',
        start_date=datetime.now(timezone.utc),
    )
    db.session.add(inst)
    db.session.flush()
    return inst


def _plain_user(tag):
    u = User(email=f'ec-{tag}@test.com', first_name='Т', last_name='У')
    db.session.add(u)
    db.session.flush()
    return u


def _registration(instance, tag):
    user = _plain_user(tag)
    reg = EventRegistration(
        user_id=user.id, instance_id=instance.id, phone='+380670000000',
        specialty='Т', workplace='Клініка', status='confirmed', payment_status='paid',
    )
    db.session.add(reg)
    db.session.flush()
    return reg


# --- Фабрики рядків під кожен роут: seed(n, tag) створює n рядків, у яких --
# --- пошукове поле (те, яке `q` роуту реально фільтрує) містить `tag` -------

def _seed_b2b(n, tag):
    for i in range(n):
        db.session.add(B2BRequest(
            first_name='Іван', last_name='Мороз', phone=f'+38067{i:07d}',
            email=f'ec-{tag}-{i}@test.com', team_size='6-10', status='new',
        ))
    db.session.flush()


def _seed_certificates(n, tag):
    course = _course(tag)
    instance = _instance(course)
    for i in range(n):
        reg = _registration(instance, f'{tag}-{i}')
        db.session.add(Certificate(
            registration_id=reg.id, user_id=reg.user_id, number=f'{tag}-{i}',
            recipient_name='Тест', event_title='Захід', pdf_path='x.pdf',
        ))
    db.session.flush()


def _seed_course_requests(n, tag):
    course = _course(tag)
    for i in range(n):
        db.session.add(CourseRequest(
            course_id=course.id, email=f'ec-{tag}-{i}@test.com', status='pending',
        ))
    db.session.flush()


def _seed_error_logs(n, tag):
    for i in range(n):
        db.session.add(ErrorLog(
            error_code=500, error_type='ServerError', error_message=f'{tag}-{i}',
        ))
    db.session.flush()


def _seed_notifications(n, tag):
    for i in range(n):
        db.session.add(EmailLog(
            to_email=f'ec-{tag}-{i}@test.com', subject='S', template_name='T',
        ))
    db.session.flush()


def _seed_online_orders(n, tag):
    course = OnlineCourse(
        sintegrum_id=uuid4().int % 1_000_000_000,
        remote_name=tag, slug=f'oc-{tag}-{_uid()}',
    )
    db.session.add(course)
    db.session.flush()
    for i in range(n):
        user = _plain_user(f'{tag}-{i}')
        db.session.add(OnlineEnrollment(
            user_id=user.id, online_course_id=course.id,
            status='active', payment_status='paid',
        ))
    db.session.flush()


def _seed_promo_codes(n, tag):
    for i in range(n):
        code = f'{tag}{i}'
        db.session.add(PromoCode(
            code=code, code_norm=code.lower(), discount_type='percent',
            discount_value=10, description=tag,
        ))
    db.session.flush()


def _seed_referrals(n, tag):
    course = _course(tag)
    instance = _instance(course)
    for i in range(n):
        reg = _registration(instance, f'{tag}-{i}')
        db.session.add(ReferralReward(
            registration_id=reg.id, referrer_kind='user', referrer_id=reg.user_id,
            referral_code=tag, points=10, status='granted',
        ))
    db.session.flush()


def _seed_refund_requests(n, tag):
    course = _course(tag)
    instance = _instance(course)
    for i in range(n):
        reg = _registration(instance, f'{tag}-{i}')
        db.session.add(RefundRequest(
            registration_id=reg.id, user_id=reg.user_id, reason=tag, status='new',
        ))
    db.session.flush()


def _seed_registrations(n, tag):
    course = _course(tag)
    instance = _instance(course)
    for i in range(n):
        _registration(instance, f'{tag}-{i}')
    db.session.flush()


def _seed_users(n, tag):
    for i in range(n):
        db.session.add(User(email=f'ec-{tag}-{i}@test.com', first_name='Т', last_name='У'))
    db.session.flush()


def _seed_instances(n, tag):
    course = Course(title=f'Курс {tag}', slug=f'in-{tag}-{_uid()}', is_active=True)
    db.session.add(course)
    db.session.flush()
    for i in range(n):
        db.session.add(CourseInstance(
            course_id=course.id, status='published', event_format='offline',
        ))
    db.session.flush()


# --- Один список замість дванадцяти копій тесту -----------------------------

ROUTE_SPECS = [
    dict(
        name='b2b', list_endpoint='admin.b2b_requests_list',
        export_endpoint='admin.b2b_requests_export',
        builder='export_b2b_requests_xlsx', seed=_seed_b2b,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='certificates', list_endpoint='admin.certificates',
        export_endpoint='admin.certificates_export',
        builder='export_certificates_xlsx', seed=_seed_certificates,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='course_requests', list_endpoint='admin.course_requests_list',
        export_endpoint='admin.course_requests_export',
        builder='export_course_requests_xlsx', seed=_seed_course_requests,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='error_logs', list_endpoint='admin.error_logs',
        export_endpoint='admin.error_logs_export',
        builder='export_error_logs_xlsx', seed=_seed_error_logs,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='notifications', list_endpoint='admin.notifications_log',
        export_endpoint='admin.notifications_log_export',
        builder='export_email_logs_xlsx', seed=_seed_notifications,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='online_orders', list_endpoint='admin.online_orders_list',
        export_endpoint='admin.online_orders_export',
        builder='export_online_orders_xlsx', seed=_seed_online_orders,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='promo_codes', list_endpoint='admin.promo_codes_list',
        export_endpoint='admin.promo_codes_export',
        builder='export_promo_codes_xlsx', seed=_seed_promo_codes,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='referrals', list_endpoint='admin.referrals_overview',
        export_endpoint='admin.referrals_export',
        builder='export_referral_rewards_xlsx', seed=_seed_referrals,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='refund_requests', list_endpoint='admin.refund_requests_list',
        export_endpoint='admin.refund_requests_export',
        builder='export_refund_requests_xlsx', seed=_seed_refund_requests,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='registrations', list_endpoint='admin.registrations_all',
        export_endpoint='admin.registrations_export',
        builder='export_registrations_xlsx', seed=_seed_registrations,
        qs=lambda tag: {'q': tag, 'scope': 'all'},
    ),
    dict(
        name='users', list_endpoint='admin.users',
        export_endpoint='admin.users_export',
        builder='export_users_xlsx', seed=_seed_users,
        qs=lambda tag: {'q': tag},
    ),
    dict(
        name='instances', list_endpoint='admin.instances_list',
        export_endpoint='admin.instances_report_export',
        builder='export_instances_report_xlsx', seed=_seed_instances,
        qs=lambda tag: {'q': tag},
    ),
]

ROUTE_IDS = [spec['name'] for spec in ROUTE_SPECS]


def _record_selects():
    """Слухач before_cursor_execute: збирає SELECT-и цієї транзакції."""
    selects = []

    def _listener(_conn, _cursor, statement, _params, _context, _many):
        s = statement.strip()
        if s.upper().startswith('SELECT'):
            selects.append(s)

    event.listen(db.engine, 'before_cursor_execute', _listener)
    return selects, _listener


# --- Тест 1 (на кожен роут): стеля рахує COUNT ДО вибірки -------------------

@pytest.mark.parametrize('spec', ROUTE_SPECS, ids=ROUTE_IDS)
def test_export_over_ceiling_counts_before_materializing(client, admin, monkeypatch, app, spec):
    """Понад стелею: у журналі SQL є COUNT і НЕМАЄ жодного іншого SELECT.

    Це і є доказ порядку дій -- на старому коді (`rows = query.all()` ДО
    перевірки стелі) цей інший SELECT завжди був би в журналі, незалежно
    від того, що сталось потім. Білдер xlsx теж не має викликатись --
    підміняємо його на такий, що падає.
    """
    ceiling = 3
    monkeypatch.setattr(_listing, 'MAX_EXPORT_ROWS', ceiling)

    # Не дискримінує стару поведінку від нової -- на дореформеному коді
    # xlsx_export теж відмовляв ДО виклику білдера, просто пізніше (по
    # len(rows) замість COUNT). Тримаємо як пасивний запобіжник другого
    # порядку: справжній доказ порядку дій -- нижче, по журналу SQL.
    def _boom(*a, **kw):
        raise AssertionError(f'{spec["builder"]} не мав викликатись понад стелею')

    monkeypatch.setattr(xlsx_reports, spec['builder'], _boom)

    tag = f'{spec["name"]}-{_uid()}'
    spec['seed'](ceiling + 3, tag)

    _login(client, admin)
    with app.test_request_context():
        from flask import url_for
        export_url = url_for(spec['export_endpoint'], **spec['qs'](tag))
        list_path = urlparse(url_for(spec['list_endpoint'])).path

    selects, listener = _record_selects()
    try:
        resp = client.get(export_url)
    finally:
        event.remove(db.engine, 'before_cursor_execute', listener)

    assert resp.status_code == 302, (
        f'{spec["name"]}: понад стелею мав бути редірект, отримано {resp.status_code}'
    )

    # Зріз збережено в редіректі: та сама сторінка списку, ті самі параметри.
    location = urlparse(resp.headers['Location'])
    assert location.path == list_path, (
        f'{spec["name"]}: редірект веде не на свій список ({location.path} != {list_path})'
    )
    q_values = parse_qs(location.query)
    for key, value in spec['qs'](tag).items():
        assert q_values.get(key) == [value], (
            f'{spec["name"]}: параметр {key!r} не зберігся в редіректі ({q_values})'
        )

    # Flash про перевищення -- із сесії ПЕРШОГО запиту (302 вище вже її
    # встановив), а не другим повним запитом до export_url: другий запит
    # заново пройшов би весь той самий шлях (і COUNT, і редірект), а нам
    # тут потрібен лише текст, який flash() уже поклав у сесію.
    with client.session_transaction() as sess:
        raw_flashes = list(sess.get('_flashes', []))
    messages = [msg for _category, msg in raw_flashes]
    assert any('більше за ліміт' in msg for msg in messages), (
        f'{spec["name"]}: немає flash про перевищення стелі ({messages})'
    )

    # Головне твердження: COUNT був, а зріз -- не матеріалізувався.
    #
    # "Іншого SELECT не було" неможливо перевірити буквально: на кожен
    # запит адмінки йдуть побічні SELECT-и, до нашого зрізу не причетні
    # (сесія current_user через flask-login, SiteSettings у контекст-
    # процесорі). Відрізняємо їх не таблицею (для роуту users побічний
    # SELECT б'є в ту саму таблицю users), а формою LIKE-умови: `q=tag` в
    # усіх дванадцяти специфікаціях -- реальний пошуковий фільтр, і РІВНО
    # він завжди потрапляє в WHERE справжнього запиту зрізу (і в COUNT, що
    # той самий запит огортає) у вигляді "lower(col) LIKE lower(?)"
    # (SQLite не має нативного ILIKE -- компілятор розгортає його саме
    # так). Побічні SELECT-и такої умови не несуть.
    #
    # Розпізнавати COUNT ЛИШЕ підрядком 'count(' небезпечно: у users
    # рядковий SELECT сам несе корельований підзапит "(SELECT count(...)
    # AS _registration_count)" -- це підрядок 'count(' усередині запиту, що
    # якраз і мав НЕ виконуватись. Справжній `.count()` SQLAlchemy завжди
    # огортає весь запит підзапитом і починається рівно з "SELECT count(*)"
    # -- цей префікс (а не підрядок будь-де) і є ознакою.
    count_queries = [s for s in selects if _COUNT_SQL_RE.match(s.strip())]
    materializing = [
        s for s in selects
        if not _COUNT_SQL_RE.match(s.strip()) and _LIKE_LOWER_RE.search(s)
    ]
    assert count_queries, (
        f'{spec["name"]}: жодного COUNT-запиту в журналі SQL -- стеля не міряна ДО вибірки'
    )
    # Порожній `materializing` доводить не-матеріалізацію лише тому, що
    # COUNT-запит несе розпізнавальний слід (LIKE lower(...) від `q=tag`).
    # Специфікація без пошукового фільтра (лише status/дата) лишила б цей
    # слід відсутнім і в COUNT-і теж -- і `materializing` був би порожнім
    # ВАКУУМНО, з тієї самої причини, що колись підвела 'ilike': перевірка
    # нічого не побачила б, бо дивилась не туди. Спершу переконуємось, що
    # сигнал узагалі є ДЕ шукати -- інакше висновок нижче не про що.
    assert any(_LIKE_LOWER_RE.search(cq) for cq in count_queries), (
        f'{spec["name"]}: у COUNT-запиті немає розпізнавального LIKE-сліду -- '
        'висновок "не матеріалізувався" з порожнього materializing нічого не '
        'доводить (додайте пошуковий фільтр у qs() цієї специфікації)'
    )
    assert not materializing, (
        f'{spec["name"]}: зріз таки матеріалізувався ДО перевірки стелі:\n'
        + '\n'.join(materializing)
    )


# --- Тест 2 (на кожен роут): у межах стелі -- ВЕСЬ зріз, а не сторінка ------

@pytest.mark.parametrize('spec', ROUTE_SPECS, ids=ROUTE_IDS)
def test_export_under_ceiling_returns_whole_slice(client, admin, app, spec):
    """У межах стелі xlsx містить УСІ рядки зрізу, а не одну сторінку.

    51 -- більше за найбільший спостережений розмір сторінки серед цих
    роутів (LIST_PER_PAGE=50 -- b2b/certificates/course_requests/
    refund_requests/registrations/users; так само 50 -- notifications,
    error_logs; є й менші: _INSTANCES_PER_PAGE=25, PER_PAGE=30 в
    online_orders/promo_codes). Якби експорт випадково почав пагінувати
    (напр., хтось замінив би `.all()` на `.paginate()` при рефакторингу),
    файл мав би не більше 50 рядків замість 51.
    """
    n = 51
    tag = f'{spec["name"]}-{_uid()}'
    spec['seed'](n, tag)

    _login(client, admin)
    with app.test_request_context():
        from flask import url_for
        export_url = url_for(spec['export_endpoint'], **spec['qs'](tag))

    resp = client.get(export_url)
    assert resp.status_code == 200, (
        f'{spec["name"]}: очікували xlsx у межах стелі, отримали {resp.status_code}'
    )
    assert resp.mimetype == _listing.XLSX_MIMETYPE

    wb = load_workbook(io.BytesIO(resp.data))
    data_rows = wb.active.max_row - 1  # мінус рядок заголовка
    assert data_rows == n, (
        f'{spec["name"]}: у файлі {data_rows} рядків замість {n} -- зріз обрізано'
    )


# --- Особливий випадок: routes_registrations.py -- сортування ---------------

def test_registrations_export_keeps_newest_first_order(client, admin, app):
    """`.order_by(EventRegistration.created_at.desc())` лишається на
    матеріалізованому запиті: `export_query` знімає ORDER BY лише для
    свого внутрішнього COUNT (`query.order_by(None).count()`), а не для
    `.all()`, який виконується вже без цього виклику.

    Рядки вставляємо у ЗРОСТАЮЧОМУ хронологічному порядку (найстаріший
    -- перший за id/rowid): якщо колись `.order_by(...)` перед
    `export_query` загубиться, дефолтний порядок видачі SQLite
    (rowid/порядок вставки) буде РІВНО таким -- від найстарішого до
    найновішого, тобто протилежним очікуваному «найновіший перший». Це і
    ловить тест: сторож на ризик, який сам сформулював план -- рядок
    `regs = query.order_by(...).all()` розбили на два, і саме сортування
    можна тихо втратити при майбутній правці.
    """
    tag = f'order-{_uid()}'
    course = _course(tag)
    instance = _instance(course)
    base = datetime.now(timezone.utc)
    reg_ids = []
    for i in range(5):
        user = _plain_user(f'{tag}-{i}')
        reg = EventRegistration(
            user_id=user.id, instance_id=instance.id, phone='+380670000000',
            specialty='Т', workplace='Клініка', status='confirmed',
            payment_status='paid', created_at=base + timedelta(minutes=i),
        )
        db.session.add(reg)
        db.session.flush()
        reg_ids.append(reg.id)

    _login(client, admin)
    with app.test_request_context():
        from flask import url_for
        export_url = url_for('admin.registrations_export', q=tag, scope='all')

    resp = client.get(export_url)
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    exported_ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert exported_ids == list(reversed(reg_ids)), (
        f'реєстрації у файлі не в порядку "найновіша перша": {exported_ids} '
        f'(мало бути {list(reversed(reg_ids))})'
    )


# --- Особливий випадок: routes_instances.py next3 ---------------------------

def test_instances_next3_export_bypasses_ceiling(client, admin, app, monkeypatch):
    """Гілка next3 (не більше 3 рядків) не має відмовляти, навіть коли
    повний (безлімітний) запит перевищив би стелю.

    export_query рахує COUNT БЕЗ limit(3) -- якби next3 уніфікували з
    рештою роутів (застосували стелю до цієї гілки теж), COUNT побачив би
    всі створені тут проведення (їх більше за стелю) і відмовив би
    експорту, який насправді віддав би лише три рядки.
    """
    ceiling = 3
    monkeypatch.setattr(_listing, 'MAX_EXPORT_ROWS', ceiling)

    tag = f'next3-{_uid()}'
    course = Course(title=f'Курс {tag}', slug=f'n3-{tag}', is_active=True)
    db.session.add(course)
    db.session.flush()
    # +1 година: роут рахує "зараз" ПІСЛЯ цього моменту, і рівність тут
    # надто крихка -- start_date мусить лишитись у майбутньому й тоді.
    future = datetime.now(timezone.utc) + timedelta(hours=1)
    # Більше за стелю (ceiling=3): якби стелю накладали і тут, COUNT
    # побачив би усі ці рядки (без limit) і відмовив би.
    for _ in range(ceiling + 3):
        db.session.add(CourseInstance(
            course_id=course.id, status='published', event_format='offline',
            start_date=future,
        ))
    db.session.flush()

    _login(client, admin)
    resp = client.get('/admin/instances/report.xlsx?quick=next3')

    assert resp.status_code == 200, (
        'next3 не мав відмовляти -- завжди віддає не більше трьох рядків'
    )
    wb = load_workbook(io.BytesIO(resp.data))
    assert wb.active.max_row - 1 == 3


def test_instances_non_next3_export_still_respects_ceiling(client, admin, app, monkeypatch):
    """Контроль до попереднього: БЕЗ next3 та сама кількість рядків
    відмовляє як завжди -- виняток стосується лише гілки next3."""
    ceiling = 3
    monkeypatch.setattr(_listing, 'MAX_EXPORT_ROWS', ceiling)

    tag = f'plain-{_uid()}'
    _seed_instances(ceiling + 3, tag)

    _login(client, admin)
    resp = client.get(f'/admin/instances/report.xlsx?q={tag}')
    assert resp.status_code == 302
