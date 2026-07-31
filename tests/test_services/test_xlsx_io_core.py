"""Спільні механізми xlsx_io: читання листа, пошук листа, drop-down-и,
парсер матеріалів.

Регресії, знайдені code review, іменовані за проблемою, яку стережуть.
"""
import io
import re
import zipfile
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from app.services import xlsx_io


def _book(sheet_title, rows, extra_sheets=(), active=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    for row in rows:
        ws.append(row)
    for name in extra_sheets:
        wb.create_sheet(name)['A1'] = 'нотатки'
    if active is not None:
        wb.active = wb.index(wb[active])
    return wb


def _save(wb, tmp_path):
    path = tmp_path / f'x-{uuid4().hex[:6]}.xlsx'
    wb.save(path)
    return path


def _strip_dimension(path):
    """Прибрати запис <dimension> -- так зберігають Google Sheets."""
    with zipfile.ZipFile(path) as src:
        items = [(i, src.read(i.filename)) for i in src.infolist()]
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w') as out:
        for item, data in items:
            if item.filename.endswith('.xml') and b'<dimension' in data:
                data = re.sub(rb'<dimension[^/]*/>', b'', data)
            out.writestr(item, data)
    path.write_bytes(buf.getvalue())
    return path


# --- _read_sheet ------------------------------------------------------------

def test_read_sheet_accepts_internal_and_ukrainian_headers(tmp_path):
    for header in (['sku', 'quantity'], ['Артикул', 'Кількість']):
        wb = _book('Матеріали', [header, ['A-1', 3]])
        ws = load_workbook(_save(wb, tmp_path))['Матеріали']
        rows = xlsx_io._read_sheet(ws, ['sku', 'quantity'], xlsx_io._MATERIALS_LABELS)
        assert rows == [{'sku': 'A-1', 'quantity': 3}]


def test_read_sheet_validates_columns_even_without_data_rows(tmp_path):
    """Раніше перевірка колонок стояла ПІСЛЯ раннього виходу по max_row, тож
    лист із чужими заголовками тихо читався як "порожньо"."""
    wb = _book('Матеріали', [['зовсім', 'інші']])
    ws = load_workbook(_save(wb, tmp_path))['Матеріали']
    with pytest.raises(ValueError, match='бракує колонок'):
        xlsx_io._read_sheet(ws, ['sku', 'quantity'], xlsx_io._MATERIALS_LABELS)


def test_read_sheet_handles_missing_dimension_in_read_only(tmp_path):
    """read_only + файл без <dimension> давав max_row=None і TypeError."""
    path = _strip_dimension(
        _save(_book('Матеріали', [['Артикул', 'Кількість'], ['A-1', 2]]), tmp_path))
    ws = load_workbook(path, read_only=True, data_only=True)['Матеріали']
    rows = xlsx_io._read_sheet(ws, ['sku', 'quantity'], xlsx_io._MATERIALS_LABELS)
    assert rows == [{'sku': 'A-1', 'quantity': 2}]


def test_read_sheet_on_empty_sheet_returns_nothing(tmp_path):
    ws = load_workbook(_save(_book('Матеріали', []), tmp_path))['Матеріали']
    assert xlsx_io._read_sheet(ws, ['sku', 'quantity'], xlsx_io._MATERIALS_LABELS) == []


def test_read_sheet_tolerates_short_rows(tmp_path):
    """Excel обрізає хвостові порожні клітинки -- рядок коротший за заголовок."""
    wb = _book('Матеріали', [['Артикул', 'Кількість'], ['A-1']])
    ws = load_workbook(_save(wb, tmp_path))['Матеріали']
    assert xlsx_io._read_sheet(ws, ['sku', 'quantity'],
                               xlsx_io._MATERIALS_LABELS) == [
        {'sku': 'A-1', 'quantity': None}]


def test_optional_column_may_be_absent(tmp_path):
    wb = _book('Матеріали', [['Артикул'], ['A-1']])
    ws = load_workbook(_save(wb, tmp_path))['Матеріали']
    rows = xlsx_io._read_sheet(ws, ['sku', 'quantity'],
                               xlsx_io._MATERIALS_LABELS, optional=('quantity',))
    assert rows == [{'sku': 'A-1'}]        # ключа немає взагалі, не None


# --- парсер матеріалів ------------------------------------------------------

def _materials_book(active=None, extra_sheets=()):
    return _book('Матеріали',
                 [['Зображення', 'Артикул', 'Назва', 'Наявно', 'Кількість'],
                  [None, 'SKU-1', 'Голка', 10, 3],
                  [None, 'SKU-2', 'Шприц', 5, 0]],
                 extra_sheets=extra_sheets, active=active)


def test_materials_reads_only_positive_quantities(tmp_path):
    assert xlsx_io.parse_materials_xlsx(
        _save(_materials_book(), tmp_path)) == {'SKU-1': 3}


def test_materials_finds_sheet_when_another_is_active(tmp_path):
    """Регресія: _find_sheet шукав ключ 'Матеріали', якого немає в
    SHEET_ALIASES, тож ЗАВЖДИ повертав None і код читав активний лист.
    Варто було адміну лишити активними власні нотатки -- і імпорт мовчки
    повертав порожньо, а повідомлення казало "Завантажено 0 позицій"."""
    path = _save(_materials_book(extra_sheets=('Нотатки',), active='Нотатки'),
                 tmp_path)
    assert xlsx_io.parse_materials_xlsx(path) == {'SKU-1': 3}


def test_materials_single_sheet_file_still_works(tmp_path):
    """Файл, зібраний вручну з іншою назвою листа, лишається робочим."""
    wb = _book('Sheet1', [['Артикул', 'Кількість'], ['SKU-9', 4]])
    assert xlsx_io.parse_materials_xlsx(_save(wb, tmp_path)) == {'SKU-9': 4}


def test_materials_reports_missing_sheet_by_name(tmp_path):
    wb = _book('Нотатки', [['a']], extra_sheets=('Ще один',))
    with pytest.raises(ValueError, match='Матеріали'):
        xlsx_io.parse_materials_xlsx(_save(wb, tmp_path))


def test_materials_ignores_invalid_quantity(tmp_path):
    wb = _book('Матеріали', [['Артикул', 'Кількість'],
                             ['SKU-1', 'багато'], ['SKU-2', 2]])
    assert xlsx_io.parse_materials_xlsx(_save(wb, tmp_path)) == {'SKU-2': 2}


# --- drop-down-и ------------------------------------------------------------

def test_inline_dropdown_skipped_when_over_excel_limit(caplog):
    """Excel не приймає inline-список довший за 255 символів і вважає файл
    пошкодженим. Краще віддати файл без випадайки."""
    ws = Workbook().active
    xlsx_io._add_inline_dropdown(
        ws, 'status', ['status'], options=['x' * 30] * 20, last_data_row=5)
    assert not ws.data_validations.dataValidation
    assert 'ліміт Excel 255' in caplog.text


def test_inline_dropdown_added_when_short_enough():
    ws = Workbook().active
    xlsx_io._add_inline_dropdown(
        ws, 'status', ['status'], options=['Так', 'Ні'], last_data_row=5)
    assert len(ws.data_validations.dataValidation) == 1


def test_reference_dropdown_quotes_sheet_name():
    """Назва листа з пробілом чи дужкою без лапок ламає формулу Excel."""
    ws = Workbook().active
    xlsx_io._add_ref_dropdown(
        ws, 'event', ['event'], last_data_row=5,
        sheet_name='Спеціалізації (довідник)', ref_last_row=10)
    formula = ws.data_validations.dataValidation[0].formula1
    assert formula.startswith("='Спеціалізації (довідник)'!")


# --- _bool ------------------------------------------------------------------

@pytest.mark.parametrize('value, expected', [
    (True, True), (False, False), (None, False), ('', False),
    ('Так', True), ('Ні', False), ('TRUE', True), ('false', False),
    (1, True), (0, False), ('yes', True), ('  так  ', True),
])
def test_bool_round_trip(value, expected):
    assert xlsx_io._bool(value) is expected
