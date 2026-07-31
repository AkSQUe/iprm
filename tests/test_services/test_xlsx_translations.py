"""XLSX-канал перекладів: експорт, розбір, застосування.

Найважливіше тут -- дві властивості каналу:
  1) він НЕ торкається канонічних українських колонок;
  2) якщо оригінал змінився після експорту, переклад не лягає поверх
     іншого тексту, а йде в конфлікти.
"""
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.i18n import source_key
from app.models.city import City
from app.models.course import Course
from app.models.program_block import ProgramBlock
from app.models.trainer import Trainer
from app.services import translation_registry as registry
from app.services import xlsx_translations as xt


# --- хелпери ----------------------------------------------------------------

def _course(**kw):
    kw.setdefault('title', f'Курс {uuid4().hex[:4]}')
    c = Course(slug=f'tr-{uuid4().hex[:6]}', is_active=True, **kw)
    db.session.add(c)
    db.session.commit()
    return c


def _export(scope='courses', only_untranslated=False):
    data = xt.export_translations_xlsx(scope, only_untranslated=only_untranslated)
    return load_workbook(data)


def _rows(ws):
    """Лист -> list[dict] за внутрішніми ключами колонок."""
    cols = xt.translation_cols()
    labels = xt.translation_labels()
    header = [c.value for c in ws[1]]
    index = {}
    for key in cols:
        for i, h in enumerate(header):
            if h in (key, labels[key]):
                index[key] = i
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        out.append({k: row[i] for k, i in index.items()})
    return out


def _write(tmp_path, sheet_rows, sheet_name='Курси'):
    """Створити файл перекладів з переданих рядків."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    cols = xt.translation_cols()
    labels = xt.translation_labels()
    ws.append([labels[c] for c in cols])
    for row in sheet_rows:
        ws.append([row.get(c, '') for c in cols])
    path = tmp_path / f'tr-{uuid4().hex[:6]}.xlsx'
    wb.save(path)
    return path


def _row_for(course, key, **overrides):
    unit = {u.uid: u for u in registry.units(course)}[key]
    base = {'entity': 'course', 'id': course.id, 'object': course.title,
            'field': unit.label, 'key': key, 'uk': unit.source}
    base.update(overrides)
    return base


# --- експорт ----------------------------------------------------------------

def test_export_has_help_sheet_and_entity_sheets(client):
    _course()
    wb = _export()
    assert wb.sheetnames[0] == 'Довідка'
    assert {'Курси', 'Блоки програми', 'Тарифи курсів'} <= set(wb.sheetnames)


def test_export_emits_row_per_unit(client):
    c = _course(subtitle='Підзаголовок',
                faq=[{'question': 'Питання?', 'answer': 'Відповідь.'}])
    rows = [r for r in _rows(_export()['Курси']) if r['id'] == c.id]
    keys = {r['key'] for r in rows}
    assert 'title' in keys
    assert 'subtitle' in keys
    assert f'faq:{source_key("Питання?")}' in keys
    assert f'faq:{source_key("Відповідь.")}' in keys


def test_export_skips_empty_sources(client):
    c = _course()  # subtitle не заповнено
    rows = [r for r in _rows(_export()['Курси']) if r['id'] == c.id]
    assert all((r['uk'] or '').strip() for r in rows)
    assert 'subtitle' not in {r['key'] for r in rows}


def test_export_todo_mode_skips_fully_translated(client):
    c = _course(subtitle='Підзаголовок')
    c.set_translation('ru', 'title', 'РУ')
    c.set_translation('en', 'title', 'EN')
    db.session.commit()

    all_keys = {r['key'] for r in _rows(_export()['Курси']) if r['id'] == c.id}
    todo_keys = {r['key'] for r in _rows(_export(only_untranslated=True)['Курси'])
                 if r['id'] == c.id}
    assert 'title' in all_keys
    assert 'title' not in todo_keys      # перекладено обома мовами
    assert 'subtitle' in todo_keys


def test_export_includes_program_blocks_and_trainers(client):
    c = _course()
    db.session.add(ProgramBlock(course_id=c.id, heading='Теорія',
                                items=['Пункт'], sort_order=0))
    t = Trainer(slug=f'tr-{uuid4().hex[:6]}', full_name='Тренер Тест', role='Лікар')
    db.session.add(t)
    db.session.commit()

    blocks = _rows(_export()['Блоки програми'])
    assert any(r['uk'] == 'Теорія' for r in blocks)

    trainers = _rows(_export('trainers')['Тренери'])
    assert any(r['uk'] == 'Тренер Тест' for r in trainers)


def test_export_instances_scope_has_cities(client):
    name = f'Місто-{uuid4().hex[:4]}'
    db.session.add(City(name=name))
    db.session.commit()
    wb = _export('instances')
    assert {'Тарифи проведень', 'Локації'} <= set(wb.sheetnames)
    assert any(r['uk'] == name for r in _rows(wb['Локації']))


def test_export_locks_source_columns_only(client):
    _course()
    ws = _export()['Курси']
    assert ws.protection.sheet is True
    # Мовні колонки відкриті, ключові -- заблоковані.
    cols = xt.translation_cols()
    uk_col = cols.index('uk') + 1
    ru_col = cols.index('ru') + 1
    assert ws.cell(row=2, column=uk_col).protection.locked is not False
    assert ws.cell(row=2, column=ru_col).protection.locked is False


def test_export_unknown_scope_raises():
    with pytest.raises(ValueError):
        xt.export_translations_xlsx('nope')


# --- розбір й застосування --------------------------------------------------

def test_round_trip_applies_translation(client, tmp_path):
    c = _course(faq=[{'question': 'Питання?', 'answer': 'Відповідь.'}])
    key = f'faq:{source_key("Питання?")}'
    path = _write(tmp_path, [
        _row_for(c, 'title', ru='Курс-РУ', en='Course-EN'),
        _row_for(c, key, ru='Вопрос?'),
    ])

    plan = xt.parse_translations_xlsx(path)
    assert plan.is_valid
    assert plan.counts['add'] == 3

    assert xt.apply_translations_plan(plan)['ok']
    fetched = db.session.get(Course, c.id)
    assert fetched.t('title', lang='ru') == 'Курс-РУ'
    assert fetched.t('title', lang='en') == 'Course-EN'
    assert fetched.t('faq', lang='ru')[0]['question'] == 'Вопрос?'


def test_import_never_touches_ukrainian(client, tmp_path):
    c = _course(title='Оригінальна назва')
    row = _row_for(c, 'title', ru='РУ')
    row['uk'] = 'Оригінальна назва'   # навіть якщо перекладач тут напише інше
    path = _write(tmp_path, [row])
    xt.apply_translations_plan(xt.parse_translations_xlsx(path))

    assert db.session.get(Course, c.id).title == 'Оригінальна назва'


def test_source_changed_becomes_conflict(client, tmp_path):
    c = _course(title='Стара назва')
    path = _write(tmp_path, [_row_for(c, 'title', ru='РУ')])

    # Адмін переписав українську після експорту.
    c.title = 'Нова назва'
    db.session.commit()

    plan = xt.parse_translations_xlsx(path)
    assert plan.counts['conflict'] == 1
    assert plan.counts['add'] == 0
    xt.apply_translations_plan(plan)
    # Переклад не ліг поверх іншого тексту.
    assert db.session.get(Course, c.id).t('title', lang='ru') == 'Нова назва'


def test_empty_cell_keeps_existing_translation(client, tmp_path):
    c = _course()
    c.set_translation('ru', 'title', 'Наявний переклад')
    db.session.commit()

    path = _write(tmp_path, [_row_for(c, 'title', ru='', en='EN')])
    plan = xt.parse_translations_xlsx(path)
    xt.apply_translations_plan(plan)

    fetched = db.session.get(Course, c.id)
    assert fetched.t('title', lang='ru') == 'Наявний переклад'
    assert fetched.t('title', lang='en') == 'EN'


def test_partial_file_keeps_untouched_json_leaves(client, tmp_path):
    """Регресія: apply_units перезбирає мапу поля повністю, тож частковий
    файл міг би стерти фрагменти, яких у ньому немає."""
    c = _course(faq=[{'question': 'Перше?', 'answer': 'Так.'},
                     {'question': 'Друге?', 'answer': 'Ні.'}])
    c.set_translation('ru', 'faq', {
        source_key('Перше?'): 'Первый?',
        source_key('Друге?'): 'Второй?',
    })
    db.session.commit()

    # У файлі лише ОДИН фрагмент поля faq.
    path = _write(tmp_path, [
        _row_for(c, f'faq:{source_key("Перше?")}', ru='Первый (нов)?'),
    ])
    xt.apply_translations_plan(xt.parse_translations_xlsx(path))

    ru = db.session.get(Course, c.id).t('faq', lang='ru')
    assert ru[0]['question'] == 'Первый (нов)?'
    assert ru[1]['question'] == 'Второй?'   # не стерто


def test_unchanged_rows_are_reported_not_applied(client, tmp_path):
    c = _course()
    c.set_translation('ru', 'title', 'РУ')
    db.session.commit()
    path = _write(tmp_path, [_row_for(c, 'title', ru='РУ')])

    plan = xt.parse_translations_xlsx(path)
    assert plan.counts['unchanged'] == 1
    assert plan.applicable == []


def test_bad_rows_do_not_block_the_file(client, tmp_path):
    """Застарілі рядки -- нормальна ситуація; вони не мають відхиляти імпорт."""
    c = _course()
    path = _write(tmp_path, [
        _row_for(c, 'title', ru='Добрий рядок'),
        {'entity': 'course', 'id': 999999, 'object': '?', 'field': 'Назва',
         'key': 'title', 'uk': 'x', 'ru': 'y'},
        {'entity': 'nope', 'id': c.id, 'object': '?', 'field': '?',
         'key': 'title', 'uk': 'x', 'ru': 'y'},
        {'entity': 'course', 'id': c.id, 'object': '?', 'field': '?',
         'key': 'faq:deadbeef0000', 'uk': 'зниклий', 'ru': 'y'},
    ])

    plan = xt.parse_translations_xlsx(path)
    assert plan.is_valid                 # файл придатний
    assert plan.counts['error'] == 3
    assert plan.counts['add'] == 1
    xt.apply_translations_plan(plan)
    assert db.session.get(Course, c.id).t('title', lang='ru') == 'Добрий рядок'


def test_missing_key_column_is_file_level_error(client, tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Курси'
    ws.append(['Тип', 'ID', 'Українська (оригінал)'])   # без колонки "Ключ"
    ws.append(['course', 1, 'x'])
    path = tmp_path / 'broken.xlsx'
    wb.save(path)

    plan = xt.parse_translations_xlsx(path)
    assert not plan.is_valid
    assert any('Ключ' in e for e in plan.errors)
    assert xt.apply_translations_plan(plan)['ok'] is False


def test_help_sheet_is_ignored_on_import(client, tmp_path):
    c = _course()
    path = _write(tmp_path, [_row_for(c, 'title', ru='РУ')])
    wb = load_workbook(path)
    wb.create_sheet('Довідка', 0)['A1'] = 'Інструкція'
    wb.save(path)

    plan = xt.parse_translations_xlsx(path)
    assert plan.is_valid
    assert plan.counts['add'] == 1
