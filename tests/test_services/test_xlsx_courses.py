"""Контентний xlsx курсів: експорт/імпорт і сумісність зі старими файлами.

Цей імпортер пише в канонічні українські колонки і робить REPLACE блоків
програми та FAQ, а тестів не мав узагалі. Приводом стала нова колонка
"Фінальний заклик": додавання колонки не має знецінювати файли, які
менеджери вже тримають на руках.
"""
import io
from uuid import uuid4

from openpyxl import load_workbook

from app.extensions import db
from app.models.course import Course
from app.services import xlsx_io


def _course(**kw):
    kw.setdefault('title', f'Курс {uuid4().hex[:4]}')
    c = Course(slug=f'x-{uuid4().hex[:6]}', event_type='course',
               is_active=True, base_price=0, **kw)
    db.session.add(c)
    db.session.commit()
    return c


def _export_sheet():
    wb = load_workbook(xlsx_io.export_courses_xlsx())
    return wb['Курси']


def _header(ws):
    return [c.value for c in ws[1]]


def _row_for(ws, slug):
    header = _header(ws)
    slug_idx = header.index(xlsx_io.COURSE_LABELS['slug'])
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[slug_idx] == slug:
            return dict(zip(header, row))
    return None


def _write_courses_file(tmp_path, rows, drop_columns=()):
    """Файл у форматі експорту; drop_columns імітує старіший експорт."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Курси'
    cols = [c for c in xlsx_io.COURSE_COLS if c not in drop_columns]
    ws.append([xlsx_io.COURSE_LABELS[c] for c in cols])
    for row in rows:
        ws.append([row.get(c, '') for c in cols])
    path = tmp_path / f'courses-{uuid4().hex[:6]}.xlsx'
    wb.save(path)
    return path


def _base_row(course, **overrides):
    row = {
        'id': course.id,
        'slug': course.slug,
        'title': course.title,
        'event_type': 'Курс',
        'base_price': 0,
        'is_active': True,
        'is_featured': False,
    }
    row.update(overrides)
    return row


# --- узгодженість формату ---------------------------------------------------

def test_export_header_matches_column_list(client):
    _course()
    header = _header(_export_sheet())
    assert header == [xlsx_io.COURSE_LABELS[c] for c in xlsx_io.COURSE_COLS]


def test_export_values_align_with_columns(client):
    """Список значень в експорті позиційний -- зсув на одну колонку зіпсував
    би дані мовчки."""
    c = _course(subtitle='Підзаголовок', agenda='Програма',
                final_cta_text='Готові почати?', tags=['PRP'])
    row = _row_for(_export_sheet(), c.slug)
    assert row[xlsx_io.COURSE_LABELS['subtitle']] == 'Підзаголовок'
    assert row[xlsx_io.COURSE_LABELS['agenda']] == 'Програма'
    assert row[xlsx_io.COURSE_LABELS['final_cta_text']] == 'Готові почати?'
    assert row[xlsx_io.COURSE_LABELS['tags']] == 'PRP'


# --- нова колонка -----------------------------------------------------------

def test_import_sets_final_cta_text(client, tmp_path):
    c = _course()
    path = _write_courses_file(tmp_path, [
        _base_row(c, final_cta_text='Оберіть дату та закріпіть місце.'),
    ])
    plan = xlsx_io.parse_courses_xlsx(path)
    assert plan.is_valid, plan.errors
    assert xlsx_io.apply_courses_plan(plan)['ok']

    assert db.session.get(Course, c.id).final_cta_text == 'Оберіть дату та закріпіть місце.'


def test_import_clears_final_cta_when_column_present_but_empty(client, tmp_path):
    c = _course(final_cta_text='Було')
    path = _write_courses_file(tmp_path, [_base_row(c, final_cta_text='')])
    xlsx_io.apply_courses_plan(xlsx_io.parse_courses_xlsx(path))
    assert db.session.get(Course, c.id).final_cta_text is None


def test_old_file_without_column_still_imports(client, tmp_path):
    """Файл, експортований до появи колонки, має лишатись робочим."""
    c = _course(final_cta_text='Не чіпати')
    path = _write_courses_file(tmp_path, [_base_row(c, title='Нова назва')],
                               drop_columns=('final_cta_text',))

    plan = xlsx_io.parse_courses_xlsx(path)
    assert plan.is_valid, plan.errors
    assert xlsx_io.apply_courses_plan(plan)['ok']

    fetched = db.session.get(Course, c.id)
    assert fetched.title == 'Нова назва'
    # Відсутня колонка = "не чіпати", а не "занулити".
    assert fetched.final_cta_text == 'Не чіпати'


def test_old_file_does_not_report_phantom_change(client, tmp_path):
    c = _course(final_cta_text='Було')
    path = _write_courses_file(tmp_path, [_base_row(c)],
                               drop_columns=('final_cta_text',))
    plan = xlsx_io.parse_courses_xlsx(path)
    change = next(ch for ch in plan.changes if ch.slug == c.slug)
    assert 'final_cta_text' not in change.fields_changed


def test_missing_required_column_is_error(client, tmp_path):
    c = _course()
    path = _write_courses_file(tmp_path, [_base_row(c)], drop_columns=('title',))
    plan = xlsx_io.parse_courses_xlsx(path)
    assert not plan.is_valid
    assert any('Назва' in e for e in plan.errors)


# --- регресія: оновлення існуючого курсу ------------------------------------

def test_existing_course_row_does_not_error(client, tmp_path):
    """Регресія: після переходу на медіа-реєстр _diff_course звертався до
    прибраних hero_image/card_image і кидав AttributeError на КОЖНОМУ
    існуючому курсі, тож файл із оновленнями відхилявся цілком -- імпорт
    працював лише для створення нових курсів."""
    c = _course(subtitle='Старий')
    path = _write_courses_file(tmp_path, [_base_row(c, subtitle='Новий')])

    plan = xlsx_io.parse_courses_xlsx(path)
    assert plan.is_valid, plan.errors
    assert not any(ch.action == 'error' for ch in plan.changes)
    assert xlsx_io.apply_courses_plan(plan)['ok']
    assert db.session.get(Course, c.id).subtitle == 'Новий'


def test_media_compared_by_resolved_id(client, tmp_path):
    """Порожня колонка зображення в експорті курсу без медіа не має
    показуватись як зміна."""
    c = _course()
    path = _write_courses_file(tmp_path, [_base_row(c, hero_image='', card_image='')])
    change = next(ch for ch in xlsx_io.parse_courses_xlsx(path).changes
                  if ch.slug == c.slug)
    assert 'hero_image' not in change.fields_changed
    assert 'card_image' not in change.fields_changed


# --- базові інваріанти імпорту ----------------------------------------------

def test_diff_marks_changed_field(client, tmp_path):
    c = _course(subtitle='Старий')
    path = _write_courses_file(tmp_path, [_base_row(c, subtitle='Новий')])
    plan = xlsx_io.parse_courses_xlsx(path)
    change = next(ch for ch in plan.changes if ch.slug == c.slug)
    assert change.action == 'update'
    assert 'subtitle' in change.fields_changed


def test_unchanged_row_reported_as_unchanged(client, tmp_path):
    c = _course()
    path = _write_courses_file(tmp_path, [_base_row(c)])
    plan = xlsx_io.parse_courses_xlsx(path)
    change = next(ch for ch in plan.changes if ch.slug == c.slug)
    assert change.action == 'unchanged'


def test_duplicate_slug_in_file_is_error(client, tmp_path):
    c = _course()
    path = _write_courses_file(tmp_path, [_base_row(c), _base_row(c)])
    plan = xlsx_io.parse_courses_xlsx(path)
    assert any(ch.action == 'error' for ch in plan.changes)
