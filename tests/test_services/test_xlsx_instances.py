"""XLSX розкладу: експорт, розбір, застосування.

Цей імпортер пише в course_instances (дати, ціни, статуси проведень) і не
мав жодного тесту. Саме такий пробіл минулого разу дозволив зламаному
імпорту курсів прожити місяці непоміченим.
"""
from datetime import datetime, timedelta, timezone
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.trainer import Trainer
from app.services import xlsx_io

KYIV = xlsx_io.KYIV


def _course(**kw):
    kw.setdefault('title', f'Курс {uuid4().hex[:4]}')
    c = Course(slug=f'i-{uuid4().hex[:6]}', event_type='course',
               is_active=True, base_price=0, **kw)
    db.session.add(c)
    db.session.commit()
    return c


def _instance(course, **kw):
    kw.setdefault('start_date', datetime.now(timezone.utc) + timedelta(days=30))
    kw.setdefault('status', 'published')
    kw.setdefault('event_format', 'offline')
    inst = CourseInstance(course_id=course.id, **kw)
    db.session.add(inst)
    db.session.commit()
    return inst


def _write(tmp_path, rows, drop_columns=()):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Розклад'
    cols = [c for c in xlsx_io.INSTANCE_COLS if c not in drop_columns]
    ws.append([xlsx_io.INSTANCE_LABELS[c] for c in cols])
    for row in rows:
        ws.append([row.get(c, '') for c in cols])
    path = tmp_path / f'sch-{uuid4().hex[:6]}.xlsx'
    wb.save(path)
    return path


def _row(course, **overrides):
    row = {
        'course_slug': course.slug,
        'start_date': (datetime.now(timezone.utc) + timedelta(days=30))
        .astimezone(KYIV).replace(tzinfo=None).isoformat(),
        'event_format': 'Офлайн',
        'status': 'Опубліковано',
        'location': 'Харків',
    }
    row.update(overrides)
    return row


# --- узгодженість формату ---------------------------------------------------

def test_export_header_matches_column_list(client):
    _instance(_course())
    ws = load_workbook(xlsx_io.export_instances_xlsx())['Розклад']
    header = [c.value for c in ws[1]]
    assert header == [xlsx_io.INSTANCE_LABELS[c] for c in xlsx_io.INSTANCE_COLS]


def test_export_values_align_with_columns(client):
    course = _course()
    _instance(course, location='Полтава', price=5000, cpd_points=12)
    ws = load_workbook(xlsx_io.export_instances_xlsx())['Розклад']
    header = [c.value for c in ws[1]]
    rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    row = next(r for r in rows if r[xlsx_io.INSTANCE_LABELS['course_slug']] == course.slug)
    assert row[xlsx_io.INSTANCE_LABELS['location']] == 'Полтава'
    assert row[xlsx_io.INSTANCE_LABELS['cpd_points']] == 12


# --- створення й оновлення --------------------------------------------------

def test_import_creates_instance(client, tmp_path):
    course = _course()
    path = _write(tmp_path, [_row(course, location='Львів')])

    plan = xlsx_io.parse_instances_xlsx(path)
    assert plan.is_valid, plan.errors
    assert plan.counts['create'] == 1
    assert xlsx_io.apply_instances_plan(plan)['ok']

    inst = CourseInstance.query.filter_by(course_id=course.id).one()
    assert inst.location == 'Львів'
    assert inst.status == 'published'
    assert inst.event_format == 'offline'


def test_existing_instance_row_does_not_error(client, tmp_path):
    """Той самий клас, що зламав імпорт курсів: _diff_* по неіснуючому полю."""
    course = _course()
    inst = _instance(course, location='Харків')
    path = _write(tmp_path, [_row(course, id=inst.id, location='Одеса')])

    plan = xlsx_io.parse_instances_xlsx(path)
    assert plan.is_valid, plan.errors
    assert not any(ch.action == 'error' for ch in plan.changes)
    assert xlsx_io.apply_instances_plan(plan)['ok']
    assert db.session.get(CourseInstance, inst.id).location == 'Одеса'


def test_unchanged_row_reported_as_unchanged(client, tmp_path):
    """Регресія: naive/aware дати робили кожне проведення "оновити" на SQLite."""
    course = _course()
    start = datetime.now(timezone.utc) + timedelta(days=30)
    inst = _instance(course, start_date=start, location='Харків')

    path = _write(tmp_path, [_row(
        course, id=inst.id, location='Харків',
        start_date=start.astimezone(KYIV).replace(tzinfo=None).isoformat(),
    )])
    change = next(ch for ch in xlsx_io.parse_instances_xlsx(path).changes
                  if ch.course_slug == course.slug)
    assert change.action == 'unchanged', change.fields_changed


def test_diff_reports_changed_date(client, tmp_path):
    course = _course()
    inst = _instance(course)
    later = (inst.start_date + timedelta(days=7)).astimezone(KYIV).replace(tzinfo=None)
    path = _write(tmp_path, [_row(course, id=inst.id, start_date=later.isoformat())])

    change = next(ch for ch in xlsx_io.parse_instances_xlsx(path).changes
                  if ch.course_slug == course.slug)
    assert change.action == 'update'
    assert 'start_date' in change.fields_changed


# --- валідація --------------------------------------------------------------

def test_unknown_course_slug_is_error(client, tmp_path):
    path = _write(tmp_path, [_row(_course(), course_slug='no-such-course')])
    plan = xlsx_io.parse_instances_xlsx(path)
    assert not plan.is_valid
    assert any('не існує' in e for e in plan.errors)


def test_end_before_start_is_error(client, tmp_path):
    course = _course()
    start = datetime.now(timezone.utc) + timedelta(days=30)
    path = _write(tmp_path, [_row(
        course,
        start_date=start.astimezone(KYIV).replace(tzinfo=None).isoformat(),
        end_date=(start - timedelta(days=1)).astimezone(KYIV)
        .replace(tzinfo=None).isoformat(),
    )])
    plan = xlsx_io.parse_instances_xlsx(path)
    assert not plan.is_valid
    assert any('пізніше' in e for e in plan.errors)


def test_unknown_status_is_error(client, tmp_path):
    path = _write(tmp_path, [_row(_course(), status='Невідомо')])
    plan = xlsx_io.parse_instances_xlsx(path)
    assert not plan.is_valid


def test_nonexistent_id_is_error(client, tmp_path):
    path = _write(tmp_path, [_row(_course(), id=999999)])
    plan = xlsx_io.parse_instances_xlsx(path)
    assert not plan.is_valid
    assert any('не існує' in e for e in plan.errors)


def test_missing_column_is_error(client, tmp_path):
    path = _write(tmp_path, [_row(_course())], drop_columns=('status',))
    plan = xlsx_io.parse_instances_xlsx(path)
    assert not plan.is_valid
    assert any('Статус' in e for e in plan.errors)


def test_missing_sheet_is_error(client, tmp_path):
    wb = Workbook()
    wb.active.title = 'Не той лист'
    path = tmp_path / 'wrong.xlsx'
    wb.save(path)
    plan = xlsx_io.parse_instances_xlsx(path)
    assert not plan.is_valid
    assert any('Розклад' in e for e in plan.errors)


def test_invalid_plan_is_not_applied(client, tmp_path):
    course = _course()
    path = _write(tmp_path, [_row(course, status='Невідомо')])
    plan = xlsx_io.parse_instances_xlsx(path)
    assert xlsx_io.apply_instances_plan(plan)['ok'] is False
    assert CourseInstance.query.filter_by(course_id=course.id).count() == 0


def test_trainer_resolved_by_name_and_slug(client, tmp_path):
    course = _course()
    trainer = Trainer(slug=f'tr-{uuid4().hex[:6]}', full_name='Тренер Розкладу')
    db.session.add(trainer)
    db.session.commit()

    for value in (trainer.full_name, trainer.slug):
        path = _write(tmp_path, [_row(course, trainer_slug=value)])
        plan = xlsx_io.parse_instances_xlsx(path)
        assert plan.is_valid, plan.errors
        assert plan.instances[0]['parsed']['trainer_id'] == trainer.id
