"""XLSX учасників: експорт, розбір, застосування.

Найризикованіший імпортер у проєкті -- він створює користувачів і
реєстрації на заходи, а тестів не мав жодного.
"""
from datetime import datetime, timedelta, timezone
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import event

from app.extensions import db
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.registration import EventRegistration
from app.models.user import User
from app.services import xlsx_io


def _instance():
    course = Course(title=f'Курс {uuid4().hex[:4]}', slug=f'p-{uuid4().hex[:6]}',
                    is_active=True, event_type='course')
    db.session.add(course)
    db.session.flush()
    inst = CourseInstance(
        course_id=course.id, status='published', event_format='offline',
        start_date=datetime.now(timezone.utc) + timedelta(days=30),
    )
    db.session.add(inst)
    db.session.commit()
    return inst


def _write(tmp_path, rows, drop_columns=()):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Учасники'
    cols = [c for c in xlsx_io.PARTICIPANT_COLS if c not in drop_columns]
    ws.append([xlsx_io.PARTICIPANT_LABELS[c] for c in cols])
    for row in rows:
        ws.append([row.get(c, '') for c in cols])
    path = tmp_path / f'part-{uuid4().hex[:6]}.xlsx'
    wb.save(path)
    return path


def _row(inst, **overrides):
    row = {
        'event': f'#{inst.id}',
        'last_name': 'Тестовий',
        'first_name': 'Тест',
        'email': f'u-{uuid4().hex[:6]}@test.com',
        'phone': '+380501112233',
        'workplace': 'Клініка',
        'status': 'Підтверджено',
        'payment_status': 'Не оплачено',
        'attended': 'Ні',
    }
    row.update(overrides)
    return row


class _Counter:
    """Лічильник SELECT-ів по конкретній таблиці."""

    def __init__(self, table):
        self.table = table
        self.count = 0

    def __enter__(self):
        self._listener = lambda conn, cur, stmt, *a: (
            setattr(self, 'count', self.count + 1)
            if stmt.lstrip().lower().startswith('select') and self.table in stmt.lower()
            else None
        )
        event.listen(db.engine, 'before_cursor_execute', self._listener)
        return self

    def __exit__(self, *exc):
        event.remove(db.engine, 'before_cursor_execute', self._listener)


# --- експорт ----------------------------------------------------------------

def test_export_header_matches_column_list(client):
    ws = load_workbook(xlsx_io.export_participants_xlsx(blank=True))['Учасники']
    header = [c.value for c in ws[1]]
    assert header == [xlsx_io.PARTICIPANT_LABELS[c] for c in xlsx_io.PARTICIPANT_COLS]


def test_blank_template_has_no_data_rows(client):
    _instance()
    ws = load_workbook(xlsx_io.export_participants_xlsx(blank=True))['Учасники']
    assert ws.max_row == 1


def test_export_includes_reference_sheets(client):
    _instance()
    wb = load_workbook(xlsx_io.export_participants_xlsx(blank=True))
    assert 'Заходи' in wb.sheetnames
    assert any('Спеціалізації' in n for n in wb.sheetnames)


# --- створення --------------------------------------------------------------

def test_import_creates_user_and_registration(client, tmp_path):
    inst = _instance()
    email = f'new-{uuid4().hex[:6]}@test.com'
    path = _write(tmp_path, [_row(inst, email=email, last_name='Новий')])

    plan = xlsx_io.parse_participants_xlsx(path)
    assert plan.is_valid, plan.errors
    assert plan.counts['create'] == 1
    assert xlsx_io.apply_participants_plan(plan)['ok']

    user = User.query.filter_by(email=email).one()
    reg = EventRegistration.query.filter_by(user_id=user.id,
                                            instance_id=inst.id).one()
    assert reg.status == 'confirmed'
    assert user.last_name == 'Новий'


def test_reimport_same_person_updates_not_duplicates(client, tmp_path):
    inst = _instance()
    email = f'dup-{uuid4().hex[:6]}@test.com'
    xlsx_io.apply_participants_plan(xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst, email=email)])))

    plan = xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst, email=email, workplace='Інша клініка')]))
    assert plan.counts['update'] == 1
    assert xlsx_io.apply_participants_plan(plan)['ok']

    user = User.query.filter_by(email=email).one()
    assert EventRegistration.query.filter_by(user_id=user.id,
                                             instance_id=inst.id).count() == 1


def test_unchanged_row_is_skipped_on_apply(client, tmp_path):
    inst = _instance()
    email = f'same-{uuid4().hex[:6]}@test.com'
    row = _row(inst, email=email)
    xlsx_io.apply_participants_plan(
        xlsx_io.parse_participants_xlsx(_write(tmp_path, [row])))

    plan = xlsx_io.parse_participants_xlsx(_write(tmp_path, [row]))
    assert plan.counts['unchanged'] == 1
    result = xlsx_io.apply_participants_plan(plan)
    assert result['skipped'] == 1
    assert result['created'] == 0


# --- валідація --------------------------------------------------------------

@pytest.mark.parametrize('field, message', [
    ('last_name', 'Прізвище'),
    ('first_name', "Ім'я"),
    ('phone', 'Телефон'),
])
def test_required_fields(client, tmp_path, field, message):
    inst = _instance()
    plan = xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst, **{field: ''})]))
    assert not plan.is_valid
    assert any(message in e for e in plan.errors)


def test_unknown_event_is_error(client, tmp_path):
    inst = _instance()
    plan = xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst, event='#999999')]))
    assert not plan.is_valid
    assert any('не знайдено' in e for e in plan.errors)


def test_unknown_specialization_is_error(client, tmp_path):
    inst = _instance()
    plan = xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst, specializations='Астрологія')]))
    assert not plan.is_valid
    assert any('спеціалізац' in e for e in plan.errors)


@pytest.mark.parametrize('field, value', [
    ('payment_amount', -100),
    ('cpd_points_awarded', -5),
    ('experience_years', 120),
    ('birth_date', '1800-01-01'),
])
def test_numeric_ranges_rejected(client, tmp_path, field, value):
    inst = _instance()
    plan = xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst, **{field: value})]))
    assert not plan.is_valid


def test_nonexistent_reg_id_is_error(client, tmp_path):
    inst = _instance()
    plan = xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst, reg_id=999999)]))
    assert not plan.is_valid
    assert any('не знайдено' in e for e in plan.errors)


def test_invalid_plan_is_not_applied(client, tmp_path):
    inst = _instance()
    email = f'bad-{uuid4().hex[:6]}@test.com'
    plan = xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst, email=email, last_name='')]))
    assert xlsx_io.apply_participants_plan(plan)['ok'] is False
    assert User.query.filter_by(email=email).count() == 0


def test_missing_column_is_error(client, tmp_path):
    inst = _instance()
    plan = xlsx_io.parse_participants_xlsx(
        _write(tmp_path, [_row(inst)], drop_columns=('phone',)))
    assert not plan.is_valid
    assert any('Телефон' in e for e in plan.errors)


# --- регресія на N+1 --------------------------------------------------------

def test_reg_ids_without_email_are_prefetched(client, tmp_path):
    """Регресія на N+1 у розборі учасників.

    Префетч за email покривав лише рядки з email. А експорт пише порожній
    email для учасників із placeholder-адресою -- саме для них у циклі
    лишався db.session.get на КОЖЕН рядок. Тепер реєстрації тягнуться
    одним запитом за reg_id.

    Міряємо саме масштабування: кількість запитів не має рости з кількістю
    рядків. Перевіряти "менше за N" безглуздо -- таке проходить і зі
    зламаним кодом, поки якийсь інший запит устиг наповнити identity map.
    """
    def queries_for(n):
        inst = _instance()
        rows = [_row(inst, email='', phone=f'+3805011{uuid4().int % 100000:05d}')
                for _ in range(n)]
        xlsx_io.apply_participants_plan(
            xlsx_io.parse_participants_xlsx(_write(tmp_path, rows)))

        regs = EventRegistration.query.filter_by(instance_id=inst.id).all()
        assert len(regs) == n
        path = _write(tmp_path, [dict(r, reg_id=reg.id)
                                 for r, reg in zip(rows, regs)])

        db.session.expunge_all()  # порожній identity map
        with _Counter('event_registrations') as counter:
            plan = xlsx_io.parse_participants_xlsx(path)
        assert plan.is_valid, plan.errors
        return counter.count

    few, many = queries_for(2), queries_for(10)
    assert few == many, (
        f'запити ростуть із кількістю рядків: {few} на 2 рядки, '
        f'{many} на 10 -- це N+1'
    )
