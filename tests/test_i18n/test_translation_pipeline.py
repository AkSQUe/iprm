"""Наскрізний конвеєр перекладу: експорт xlsx -> заповнення -> імпорт ->
публічна сторінка трьома мовами.

Окремі ланки покриті своїми тестами; тут перевіряється, що вони стикуються.
Найдорожча помилка в такій системі -- та, де кожна частина працює, а разом
переклад не доїжджає до відвідувача.

Перекладач імітується чесно: беруться саме ті рядки, які віддав експорт, і
заповнюються мовні колонки -- жодних припущень про внутрішні ключі.
"""
from datetime import datetime, timedelta, timezone
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models.city import City
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.instance_tariff import InstanceTariff
from app.models.program_block import ProgramBlock
from app.models.trainer import Trainer
from app.services import city_glossary
from app.services import xlsx_translations as xt


# --- імітація роботи перекладача -------------------------------------------

def _export_rows(scope):
    """{лист: [рядок-dict]} -- як їх бачить перекладач у файлі."""
    wb = load_workbook(xt.export_translations_xlsx(scope))
    cols = xt.translation_cols()
    labels = xt.translation_labels()
    sheets = {}
    for name in wb.sheetnames:
        if name == xt.HELP_SHEET:
            continue
        ws = wb[name]
        header = [c.value for c in ws[1]]
        index = {k: header.index(labels[k]) for k in cols if labels[k] in header}
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and str(v).strip() for v in row):
                continue
            rows.append({k: row[i] for k, i in index.items()})
        sheets[name] = rows
    return sheets


def _translate(sheets, translator, tmp_path):
    """Заповнити мовні колонки і зберегти файл, як це зробив би перекладач.

    translator(uk_text) -> (ru, en) або None, щоб лишити рядок порожнім.
    """
    wb = Workbook()
    wb.remove(wb.active)
    cols = xt.translation_cols()
    labels = xt.translation_labels()
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append([labels[c] for c in cols])
        for row in rows:
            filled = dict(row)
            pair = translator(row['uk'])
            if pair:
                filled['ru'], filled['en'] = pair
            ws.append([filled.get(c, '') for c in cols])
    path = tmp_path / f'tr-{uuid4().hex[:6]}.xlsx'
    wb.save(path)
    return path


def _import(path):
    plan = xt.parse_translations_xlsx(path)
    assert plan.is_valid, plan.errors
    result = xt.apply_translations_plan(plan)
    assert result['ok'], result
    return plan, result


# --- курс -------------------------------------------------------------------

def test_course_pipeline_reaches_public_page(get_localized, tmp_path):
    slug = f'pipe-{uuid4().hex[:6]}'
    course = Course(
        title='Практикум з плазмотерапії', slug=slug, is_active=True,
        subtitle='Базовий рівень',
        faq=[{'question': 'Кому підійде?', 'answer': 'Лікарям.'}],
        tags=['Плазмотерапія'],
    )
    db.session.add(course)
    db.session.flush()
    db.session.add(ProgramBlock(course_id=course.id, heading='Теоретична частина',
                                items=['Механізми дії'], sort_order=0))
    db.session.commit()

    mapping = {
        'Практикум з плазмотерапії': ('Практикум по плазмотерапии', 'Plasma Therapy Workshop'),
        'Базовий рівень': ('Базовый уровень', 'Basic level'),
        'Кому підійде?': ('Кому подойдёт?', 'Who is it for?'),
        'Лікарям.': ('Врачам.', 'Doctors.'),
        'Плазмотерапія': ('Плазмотерапия', 'Plasma therapy'),
        'Теоретична частина': ('Теоретическая часть', 'Theory'),
        'Механізми дії': ('Механизмы действия', 'Mechanisms of action'),
    }
    path = _translate(_export_rows('courses'), mapping.get, tmp_path)
    plan, result = _import(path)
    assert result['applied'] >= len(mapping) * 2

    uk = get_localized(f'/courses/{slug}').get_data(as_text=True)
    ru = get_localized(f'/ru/courses/{slug}').get_data(as_text=True)
    en = get_localized(f'/en/courses/{slug}').get_data(as_text=True)

    for source, (ru_text, en_text) in mapping.items():
        assert source in uk, f'укр загубився: {source}'
        assert ru_text in ru, f'ru не доїхав: {ru_text}'
        assert en_text in en, f'en не доїхав: {en_text}'
        assert ru_text not in uk, f'ru протік на укр-сторінку: {ru_text}'


def test_course_pipeline_leaves_ukrainian_untouched(client, tmp_path):
    course = Course(title='Недоторканна назва', slug=f'pipe-{uuid4().hex[:6]}',
                    is_active=True, faq=[{'question': 'Q?', 'answer': 'A.'}])
    db.session.add(course)
    db.session.commit()
    before = (course.title, list(course.faq))

    path = _translate(_export_rows('courses'),
                      lambda uk: (f'ru::{uk}', f'en::{uk}'), tmp_path)
    _import(path)

    fetched = db.session.get(Course, course.id)
    assert (fetched.title, list(fetched.faq)) == before


# --- тренер -----------------------------------------------------------------

def test_trainer_pipeline_reaches_public_page(get_localized, tmp_path):
    slug = f'trn-{uuid4().hex[:6]}'
    trainer = Trainer(
        slug=slug, full_name='Гусак Валерія', role='Лікарка-дерматологиня',
        bio='Практикує з 2015 року.',
        skills=['Ін’єкційні техніки'],
    )
    db.session.add(trainer)
    db.session.commit()

    mapping = {
        'Лікарка-дерматологиня': ('Врач-дерматолог', 'Dermatologist'),
        'Практикує з 2015 року.': ('Практикует с 2015 года.', 'In practice since 2015.'),
        'Ін’єкційні техніки': ('Инъекционные техники', 'Injection techniques'),
    }
    path = _translate(_export_rows('trainers'), mapping.get, tmp_path)
    _import(path)

    ru = get_localized(f'/ru/trainers/{slug}').get_data(as_text=True)
    en = get_localized(f'/en/trainers/{slug}').get_data(as_text=True)
    uk = get_localized(f'/trainers/{slug}').get_data(as_text=True)

    for source, (ru_text, en_text) in mapping.items():
        assert ru_text in ru, f'ru не доїхав: {ru_text}'
        assert en_text in en, f'en не доїхав: {en_text}'
        assert source in uk, f'укр загубився: {source}'


# --- розклад: тарифи проведення + довідник локацій --------------------------

def test_instances_pipeline_reaches_public_page(get_localized, tmp_path):
    slug = f'sch-{uuid4().hex[:6]}'
    course = Course(title='Курс розкладу', slug=slug, is_active=True)
    db.session.add(course)
    db.session.flush()
    instance = CourseInstance(
        course_id=course.id, location='Харків', event_format='offline',
        status='published',
        start_date=datetime.now(timezone.utc) + timedelta(days=30),
    )
    db.session.add(instance)
    db.session.flush()
    db.session.add(InstanceTariff(
        instance_id=instance.id, name='Практикум', price=6000,
        description='Лекція наживо', sort_order=0, is_active=True,
    ))
    db.session.add(City(name='Харків'))
    db.session.commit()

    mapping = {
        'Практикум': ('Практикум-РУ', 'Workshop'),
        'Лекція наживо': ('Лекция вживую', 'Live lecture'),
        'Харків': ('Харьков', 'Kharkiv'),
    }
    path = _translate(_export_rows('instances'), mapping.get, tmp_path)
    _import(path)

    from flask import g
    g.pop(city_glossary._CACHE_ATTR, None)
    ru = get_localized(f'/ru/courses/{slug}').get_data(as_text=True)
    g.pop(city_glossary._CACHE_ATTR, None)
    en = get_localized(f'/en/courses/{slug}').get_data(as_text=True)
    g.pop(city_glossary._CACHE_ATTR, None)
    uk = get_localized(f'/courses/{slug}').get_data(as_text=True)

    assert 'Практикум-РУ' in ru and 'Workshop' in en
    assert 'Харьков' in ru and 'Kharkiv' in en
    assert 'Харків' in uk and 'Харьков' not in uk


# --- повторний прогін -------------------------------------------------------

def test_second_import_is_idempotent(client, tmp_path):
    course = Course(title='Курс повтору', slug=f'pipe-{uuid4().hex[:6]}',
                    is_active=True, faq=[{'question': 'Q?', 'answer': 'A.'}])
    db.session.add(course)
    db.session.commit()

    path = _translate(_export_rows('courses'),
                      lambda uk: (f'ru::{uk}', f'en::{uk}'), tmp_path)
    _import(path)
    snapshot = dict(db.session.get(Course, course.id).translations)

    plan = xt.parse_translations_xlsx(path)
    assert plan.counts['add'] == 0        # усе вже застосовано
    assert plan.counts['unchanged'] > 0
    xt.apply_translations_plan(plan)
    assert db.session.get(Course, course.id).translations == snapshot


def test_reexport_carries_translations_back(client, tmp_path):
    """Другий експорт має віддати вже заповнені мовні колонки -- інакше
    перекладач не бачить, що вже зроблено."""
    course = Course(title='Курс реекспорту', slug=f'pipe-{uuid4().hex[:6]}',
                    is_active=True)
    db.session.add(course)
    db.session.commit()

    path = _translate(_export_rows('courses'),
                      lambda uk: (f'ru::{uk}', f'en::{uk}'), tmp_path)
    _import(path)

    rows = [r for r in _export_rows('courses')['Курси'] if r['id'] == course.id]
    title_row = next(r for r in rows if r['key'] == 'title')
    assert title_row['ru'] == 'ru::Курс реекспорту'
    assert title_row['en'] == 'en::Курс реекспорту'
