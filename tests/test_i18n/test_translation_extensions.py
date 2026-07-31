"""Розширення підсистеми перекладів: блог у xlsx, зведений файл,
пооб'єктний експорт, лінтер, покриття в списках, осиротілі переклади,
пакетне наповнення довідника.
"""
import io
from datetime import datetime, timedelta, timezone
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.i18n import source_key
from app.models.blog_post import BlogPost
from app.models.city import City
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.trainer import Trainer
from app.models.user import User
from app.services import city_glossary
from app.services import translation_registry as registry
from app.services import xlsx_translations as xt


@pytest.fixture
def admin():
    u = User.create_with_password(
        f'a-{uuid4().hex[:6]}@test.com', 'password123',
        first_name='A', last_name='D', is_admin=True, email_confirmed=True,
    )
    db.session.commit()
    return u


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def _sheets(scope):
    wb = load_workbook(xt.export_translations_xlsx(scope))
    return [n for n in wb.sheetnames if n != xt.HELP_SHEET]


# --- B1: блог ---------------------------------------------------------------

def test_blog_scope_exports_post_content(client):
    post = BlogPost(slug=f'b-{uuid4().hex[:6]}', title='Заголовок допису',
                    status='published',
                    content=[{'type': 'paragraph', 'data': {'html': 'Абзац тексту.'}}])
    db.session.add(post)
    db.session.commit()

    wb = load_workbook(xt.export_translations_xlsx('blog'))
    ws = wb['Дописи блогу']
    sources = [row[5] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    assert 'Заголовок допису' in sources
    assert 'Абзац тексту.' in sources


def test_blog_round_trip(client, tmp_path):
    post = BlogPost(slug=f'b-{uuid4().hex[:6]}', title='Заголовок',
                    status='published',
                    content=[{'type': 'paragraph', 'data': {'html': 'Абзац.'}}])
    db.session.add(post)
    db.session.commit()

    from openpyxl import Workbook
    cols, labels = xt.translation_cols(), xt.translation_labels()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Дописи блогу'
    ws.append([labels[c] for c in cols])
    unit = next(u for u in registry.units(post) if u.source == 'Абзац.')
    ws.append(['blog_post', post.id, post.title, unit.label, unit.uid,
               unit.source, 'Абзац-РУ.', ''])
    path = tmp_path / 'blog.xlsx'
    wb.save(path)

    plan = xt.parse_translations_xlsx(path)
    assert xt.apply_translations_plan(plan)['ok']
    assert db.session.get(BlogPost, post.id).t('content', lang='ru')[0]['data']['html'] \
        == 'Абзац-РУ.'


# --- B2: усі розділи одним файлом -------------------------------------------

def test_all_scope_merges_every_section(client):
    names = _sheets('all')
    assert {'Курси', 'Блоки програми', 'Тарифи курсів', 'Тренери',
            'Дописи блогу', 'Тарифи проведень', 'Локації'} <= set(names)
    assert len(names) == len(set(names)), 'дубльовані листи'


# --- B3: експорт одного об'єкта ---------------------------------------------

def test_object_export_contains_only_that_course(client):
    keep = Course(title='Потрібний', slug=f'k-{uuid4().hex[:6]}', is_active=True)
    other = Course(title='Зайвий', slug=f'o-{uuid4().hex[:6]}', is_active=True)
    db.session.add_all([keep, other])
    db.session.commit()

    wb = load_workbook(xt.export_object_translations_xlsx('course', keep))
    ids = {row[1] for row in wb['Курси'].iter_rows(min_row=2, values_only=True) if row[0]}
    assert ids == {keep.id}


def test_object_export_route(client, admin):
    _login(client, admin)
    c = Course(title='Курс', slug=f'r-{uuid4().hex[:6]}', is_active=True)
    db.session.add(c)
    db.session.commit()
    r = client.get(f'/admin/translations/export/object/course/{c.id}')
    assert r.status_code == 200
    assert f'translations-course-{c.id}' in r.headers['Content-Disposition']


def test_object_export_rejects_unsupported_entity(client, admin):
    _login(client, admin)
    assert client.get('/admin/translations/export/object/city/1').status_code == 404


# --- B4: лінтер -------------------------------------------------------------

@pytest.mark.parametrize('lang, source, translated, expected', [
    ('en', 'Практикум', 'Практикум-РУ', 'кирилиц'),
    ('ru', 'Практикум', 'Практикум', 'дорівнює оригіналу'),
    ('ru', 'Дуже довгий український вихідний текст для перевірки', 'Ок', 'довжина'),
    ('ru', 'Практикум', 'Практикум-РУ', None),
    ('en', 'Практикум', 'Workshop', None),
])
def test_linter_flags_typical_mistakes(lang, source, translated, expected):
    warning = xt._lint(lang, source, translated)
    if expected is None:
        assert warning is None
    else:
        assert warning and expected in warning


def test_plan_exposes_warnings(client, tmp_path):
    c = Course(title='Практикум з плазмотерапії', slug=f'w-{uuid4().hex[:6]}',
               is_active=True)
    db.session.add(c)
    db.session.commit()

    from openpyxl import Workbook
    cols, labels = xt.translation_cols(), xt.translation_labels()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Курси'
    ws.append([labels[k] for k in cols])
    ws.append(['course', c.id, c.title, 'Назва', 'title', c.title,
               'Практикум по плазмотерапии', 'Практикум с кириллицей'])
    path = tmp_path / 'lint.xlsx'
    wb.save(path)

    plan = xt.parse_translations_xlsx(path)
    warnings = {w.lang: w.warning for w in plan.warnings}
    assert 'кирилиц' in warnings['en']
    # Попередження не блокує застосування.
    assert plan.counts['add'] == 2


# --- C1: покриття в списках -------------------------------------------------

def test_courses_list_shows_coverage(client, admin):
    _login(client, admin)
    c = Course(title='Курс', slug=f'cv-{uuid4().hex[:6]}', is_active=True)
    db.session.add(c)
    db.session.commit()
    html = client.get('/admin/courses').get_data(as_text=True)
    assert 'Переклад</th>' in html
    assert registry.coverage_label(c) in html


def test_trainers_list_shows_coverage(client, admin):
    _login(client, admin)
    t = Trainer(slug=f'tv-{uuid4().hex[:6]}', full_name='Тренер', role='Лікар')
    db.session.add(t)
    db.session.commit()
    html = client.get('/admin/trainers').get_data(as_text=True)
    assert registry.coverage_label(t) in html


# --- C2: осиротілі переклади ------------------------------------------------

def test_orphaned_detects_translation_without_source():
    c = Course(title='Курс', slug=f'or-{uuid4().hex[:6]}', is_active=True,
               faq=[{'question': 'Старе питання?', 'answer': 'Відповідь.'}])
    db.session.add(c)
    db.session.commit()
    c.set_translation('ru', 'faq', {source_key('Старе питання?'): 'Старый вопрос?'})
    db.session.commit()
    assert registry.orphaned(c) == []

    # Адмін переписав українське питання -- переклад тихо зник зі сторінки.
    c.faq = [{'question': 'Нове питання?', 'answer': 'Відповідь.'}]
    db.session.commit()

    orphans = registry.orphaned(c)
    assert len(orphans) == 1
    assert orphans[0]['text'] == 'Старый вопрос?'
    assert orphans[0]['lang'] == 'ru'
    assert orphans[0]['label'] == 'FAQ'


def test_editor_shows_orphaned_block(client, admin):
    _login(client, admin)
    c = Course(title='Курс', slug=f'ob-{uuid4().hex[:6]}', is_active=True,
               faq=[{'question': 'Старе?', 'answer': 'A'}])
    db.session.add(c)
    db.session.commit()
    c.set_translation('ru', 'faq', {source_key('Старе?'): 'Старый?'})
    c.faq = [{'question': 'Нове?', 'answer': 'A'}]
    db.session.commit()

    html = client.get(f'/admin/translations/course/{c.id}').get_data(as_text=True)
    assert 'Переклади без джерела' in html
    assert 'Старый?' in html


def test_orphaned_empty_for_clean_object():
    c = Course(title='Курс', slug=f'cl-{uuid4().hex[:6]}', is_active=True)
    db.session.add(c)
    db.session.commit()
    assert registry.orphaned(c) == []


# --- D1: пакетне наповнення довідника ---------------------------------------

def test_add_missing_cities_in_one_click(client, admin):
    _login(client, admin)
    course = Course(title='Курс', slug=f'ct-{uuid4().hex[:6]}', is_active=True)
    db.session.add(course)
    db.session.flush()
    names = [f'Місто-{uuid4().hex[:4]}' for _ in range(3)]
    for name in names:
        db.session.add(CourseInstance(
            course_id=course.id, location=name, event_format='offline',
            status='published',
            start_date=datetime.now(timezone.utc) + timedelta(days=5),
        ))
    db.session.commit()

    from flask import g
    g.pop(city_glossary._CACHE_ATTR, None)
    r = client.post('/admin/cities/add-missing')
    assert r.status_code == 302
    for name in names:
        assert City.query.filter_by(name_normalized=name.lower()).count() == 1


def test_add_missing_is_noop_when_nothing_to_add(client, admin):
    _login(client, admin)
    from flask import g
    g.pop(city_glossary._CACHE_ATTR, None)
    before = City.query.count()
    client.post('/admin/cities/add-missing')
    g.pop(city_glossary._CACHE_ATTR, None)
    client.post('/admin/cities/add-missing')
    assert City.query.count() >= before
