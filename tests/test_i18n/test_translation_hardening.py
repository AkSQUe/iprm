"""Регресії підсистеми перекладів, знайдені code review.

Кожен тест named за проблемою, яку стереже, а не за функцією -- інакше
через півроку буде незрозуміло, чому саме так.
"""
from tests.support.rbac import grant_role
import io
import re
import zipfile
from uuid import uuid4

import pytest
from openpyxl import Workbook

from app.extensions import db
from app.i18n import source_key
from app.models.city import City
from app.models.course import Course
from app.models.user import User
from app.services import translation_registry as registry
from app.services import xlsx_translations as xt


@pytest.fixture
def admin():
    u = User.create_with_password(
        f'a-{uuid4().hex[:6]}@test.com', 'password123',
        first_name='A', last_name='D', email_confirmed=True,
    )
    grant_role(u, 'super_admin')
    db.session.commit()
    return u


def _login(client, user):
    with client.session_transaction() as s:
        s['_user_id'] = str(user.id)


def _course(**kw):
    kw.setdefault('title', f'Курс {uuid4().hex[:4]}')
    c = Course(slug=f'h-{uuid4().hex[:6]}', is_active=True, **kw)
    db.session.add(c)
    db.session.commit()
    return c


# --- BUG-1: legacy path-ключі зникали при збереженні -----------------------

def test_editor_shows_legacy_path_translation():
    """Оверрайд зі старим path-ключем має бути видимий редактору.

    Інакше адмін бачив порожнє поле там, де публічна сторінка показувала
    переклад, і перше ж збереження форми стирало його назовсім.
    """
    c = _course(faq=[{'question': 'Питання?', 'answer': 'Відповідь.'}])
    c.translations = {'ru': {'faq': {'0.question': 'Вопрос?'}}}
    db.session.commit()

    unit = next(u for u in registry.units(c) if u.source == 'Питання?')
    assert registry.stored_value(c, 'ru', unit) == 'Вопрос?'


def test_saving_editor_does_not_wipe_legacy_translation(client, admin):
    _login(client, admin)
    c = _course(faq=[{'question': 'Питання?', 'answer': 'Відповідь.'}])
    c.translations = {'ru': {'faq': {'0.question': 'Вопрос?'}}}
    db.session.commit()

    unit = next(u for u in registry.units(c) if u.source == 'Питання?')
    # Адмін відкрив редактор і зберіг, нічого не змінюючи: у формі буде те,
    # що показав stored_value.
    client.post(f'/admin/translations/course/{c.id}',
                data={f'ru__faq__{unit.src_key}': 'Вопрос?'})

    fetched = db.session.get(Course, c.id)
    assert fetched.t('faq', lang='ru')[0]['question'] == 'Вопрос?'
    # Заразом ключ переведено на новий формат.
    assert unit.src_key in fetched.translations['ru']['faq']


def test_hash_key_wins_over_legacy_path_in_stored_value():
    c = _course(faq=[{'question': 'Питання?', 'answer': 'Відповідь.'}])
    unit = next(u for u in registry.units(c) if u.source == 'Питання?')
    c.translations = {'ru': {'faq': {'0.question': 'старе', unit.src_key: 'нове'}}}
    db.session.commit()
    assert registry.stored_value(c, 'ru', unit) == 'нове'


# --- BUG-2: файл без <dimension> клав прев'ю в 500 -------------------------

def _strip_dimension(path):
    """Прибрати запис <dimension> -- так зберігають Google Sheets."""
    with zipfile.ZipFile(path) as src:
        items = [(i, src.read(i.filename)) for i in src.infolist()]
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w') as out:
        for item, data in items:
            if item.filename.endswith('.xml') and b'<dimension' in data:
                data = re.sub(rb'<dimension[^/]*/>', b'', data)
            out.writestr(item, data)
    buf.seek(0)
    path.write_bytes(buf.read())
    return path


def _minimal_file(tmp_path, course, **langs):
    unit = {u.uid: u for u in registry.units(course)}['title']
    wb = Workbook()
    ws = wb.active
    ws.title = 'Курси'
    cols = xt.translation_cols()
    labels = xt.translation_labels()
    ws.append([labels[c] for c in cols])
    row = {'entity': 'course', 'id': course.id, 'object': course.title,
           'field': unit.label, 'key': 'title', 'uk': unit.source, **langs}
    ws.append([row.get(c, '') for c in cols])
    path = tmp_path / f'h-{uuid4().hex[:6]}.xlsx'
    wb.save(path)
    return path


def test_file_without_dimension_parses(client, tmp_path):
    """openpyxl у read-only дає max_row=None, якщо у файлі немає <dimension>;
    порівняння з числом падало TypeError -- прев'ю віддавало 500."""
    c = _course()
    path = _strip_dimension(_minimal_file(tmp_path, c, ru='Курс-РУ'))

    plan = xt.parse_translations_xlsx(path)
    assert plan.is_valid, plan.errors
    assert plan.counts['add'] == 1


def test_empty_sheet_parses_without_crash(client, tmp_path):
    wb = Workbook()
    wb.active.title = 'Курси'
    path = tmp_path / 'empty.xlsx'
    wb.save(path)
    plan = xt.parse_translations_xlsx(path)
    assert plan.changes == []


# --- EDGE: порожнє джерело більше не обходить перевірку --------------------

def test_empty_source_column_is_refused(client, tmp_path):
    """Без цього весь захист каналу вимикався очищенням однієї колонки."""
    c = _course()
    path = _minimal_file(tmp_path, c, ru='Курс-РУ')
    # Перекладач стер колонку "Українська".
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb['Курси']
    uk_col = xt.translation_cols().index('uk') + 1
    # Саме присвоєння: cell(..., value=None) в openpyxl комірку НЕ очищає.
    ws.cell(row=2, column=uk_col).value = None
    wb.save(path)

    plan = xt.parse_translations_xlsx(path)
    assert plan.counts['add'] == 0
    assert plan.counts['error'] == 1
    xt.apply_translations_plan(plan)
    assert db.session.get(Course, c.id).t('title', lang='ru') == c.title


def test_entity_is_case_insensitive(client, tmp_path):
    c = _course()
    path = _minimal_file(tmp_path, c, ru='Курс-РУ')
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb['Курси']
    ws.cell(row=2, column=1, value='Course')
    wb.save(path)

    plan = xt.parse_translations_xlsx(path)
    assert plan.counts['add'] == 1


# --- RELIABILITY: файл звільняється навіть при помилці розбору -------------

def test_upload_file_is_removable_after_parse(client, tmp_path):
    """Незакритий хендл книги на Windows блокував видалення завантаження."""
    c = _course()
    path = _minimal_file(tmp_path, c, ru='Курс-РУ')
    xt.parse_translations_xlsx(path)
    path.unlink()  # має пройти без PermissionError
    assert not path.exists()


# --- RELIABILITY: довідник локацій не падає 500 на помилці БД --------------

def test_duplicate_city_is_flashed_not_500(client, admin, monkeypatch):
    """Гонка двох адмінів: unique-конфлікт має дати повідомлення, не 500."""
    _login(client, admin)
    name = f'Місто-{uuid4().hex[:4]}'
    db.session.add(City(name=name))
    db.session.commit()

    # Прикидаємось, що перевірка на існування не спрацювала (інший адмін
    # додав рядок між SELECT і INSERT).
    monkeypatch.setattr(
        'app.admin.routes_cities.City.query',
        type('Q', (), {'filter_by': staticmethod(
            lambda **kw: type('R', (), {'first': staticmethod(lambda: None)})())})(),
    )
    r = client.post('/admin/cities/add', data={'name': name})
    assert r.status_code == 302

    monkeypatch.undo()
    assert City.query.filter_by(name_normalized=name.lower()).count() == 1


# --- каталоги: fuzzy-записи ловимо на рівні репозиторію ----------------------

def test_catalogs_have_no_fuzzy_entries():
    """`pybabel update` не має лишати після себе fuzzy-записів.

    Для кожного НОВОГО msgid babel шукає найсхожіший наявний (difflib) і копіює
    його переклад, позначаючи `#, fuzzy`. Схожість буває оманливою: при
    впровадженні тестування «Пройти тестування» отримало `Sorting`, «Продовжити
    тестування» -- `Continue with Apple`, «Завантажити сертифікат (PDF)» --
    `Download invoice (PDF)`.

    Небезпека саме в тихості: `pybabel compile` fuzzy пропускає, тож у проді
    з'явилась би не абракадабра, а фолбек на українську. Сторінка виглядає
    робочою, просто неперекладеною -- і так живе довго.

    Якщо цей тест упав: перегляньте позначені записи, впишіть правильний
    переклад і ЗНІМІТЬ прапорець `#, fuzzy` (без цього рядок не потрапить у
    .mo), далі `pybabel compile -d app/translations`.
    """
    from pathlib import Path

    from babel.messages.pofile import read_po

    root = Path(__file__).resolve().parents[2] / 'app' / 'translations'
    offenders = []
    for lang in ('ru', 'en'):
        path = root / lang / 'LC_MESSAGES' / 'messages.po'
        with open(path, encoding='utf-8') as fh:
            catalog = read_po(fh)
        for message in catalog:
            if message.id and message.fuzzy:
                source = ' '.join(str(message.id).split())[:60]
                target = ' '.join(str(message.string).split())[:40]
                offenders.append(f'{lang}: {source!r} -> {target!r}')

    assert not offenders, (
        'fuzzy-записи в каталогах (переклад не потрапить у .mo):\n  '
        + '\n  '.join(offenders)
    )
