"""XLSX export/import для адмінки: курси (з program_blocks + faq) і
проведення курсів (CourseInstance).

Загальний контракт:
  - Експорт повертає `io.BytesIO` з готовим xlsx, який віддається через
    `send_file` в роуті.
  - Імпорт працює у дві стадії:
      1. `parse_*_xlsx(file_path) -> ImportPlan` -- read + validate, БЕЗ
         запису в БД. Якщо є помилки, plan.errors заповнений; apply
         відмовляється виконувати.
      2. `apply_*_plan(plan) -> ApplyResult` -- atomic commit. Усе або
         нічого.
  - Тимчасовий xlsx-файл під час preview зберігається у
    `instance/xlsx_imports/{token}.xlsx`, де token = uuid4. Apply
    видаляє файл після успіху.

Дизайн-рішення (узгоджено з admin-користувачем):
  - Рядки, що є в БД, але відсутні у xlsx, -> залишаємо без змін.
  - program_blocks та faq для course_slug, який присутній у відповідній
    sheet, ПОВНІСТЮ замінюються (REPLACE). Якщо course_slug не зустрі-
    чається у sheet -> блоки/FAQ цього курсу не чіпаємо.
  - Trainer-FK у xlsx подається як trainer_slug (human-readable).
"""
from __future__ import annotations

import io
import logging
import re
import time
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import joinedload

from app.extensions import db
from app.models.course import Course
from app.models.course_instance import CourseInstance
from app.models.medical_profile import MedicalProfile
from app.models.media_file import MediaFile
from app.models.program_block import ProgramBlock
from app.models.registration import EventRegistration
from app.models.specializations import SPECIALIZATIONS
from app.models.trainer import Trainer
from app.models.user import User
from app.utils import ensure_utc

logger = logging.getLogger(__name__)


def _resolve_media_id(url):
    """URL-зображення (/media/...) -> id MediaFile у реєстрі, або None.

    Після Фази 6 курси зберігають зображення лише через media_id. У xlsx
    лишається людиночитний URL; на імпорті резолвимо його назад у реєстр за
    file_path. Невідомі/зовнішні URL -> None (адмін довантажує через UI)."""
    url = (url or '').strip()
    prefix = '/media/'
    if not url.startswith(prefix):
        return None
    file_path = url[len(prefix):]
    media = MediaFile.query.filter_by(file_path=file_path).first()
    return media.id if media else None


# ----------------------------------------------------------------------
# Загальні константи / утиліти
# ----------------------------------------------------------------------

KYIV = timezone(timedelta(hours=3))  # UTC+3

HEADER_FILL = PatternFill('solid', fgColor='4F46E5')
HEADER_FONT = Font(color='FFFFFF', bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')

# ----- Number formats ------------------------------------------------
# Дроблені формати для різних типів даних. Use cell.number_format = ...
FMT_INT = '0'
FMT_CURRENCY_UAH = '#,##0 "₴"'
FMT_DATETIME = 'YYYY-MM-DD HH:MM'
FMT_DATE = 'DD.MM.YYYY'

# Per-key number_format мапа. Якщо ключ відсутній — формат не виставляємо
# (текстовий за замовчуванням).
NUMBER_FORMATS = {
    # Courses
    'id': FMT_INT,
    'base_price': FMT_CURRENCY_UAH,
    'cpd_points': FMT_INT,
    'max_participants': FMT_INT,
    # Instances
    'price': FMT_CURRENCY_UAH,
    'start_date': FMT_DATETIME,
    'end_date': FMT_DATETIME,
    # Program blocks
    'sort_order': FMT_INT,
    # Participants
    'reg_id': FMT_INT,
    'birth_date': FMT_DATE,
    # Списки адмінки (реєстрації, користувачі)
    'event_date': FMT_DATE,
    'created_at': FMT_DATETIME,
    'last_login_at': FMT_DATETIME,
    'place_number': FMT_INT,
    'user_id': FMT_INT,
    'registrations': FMT_INT,
    'issued_at': FMT_DATETIME,
    'updated_at': FMT_DATETIME,
    'sent_at': FMT_DATETIME,
    'retry_count': FMT_INT,
    'resolved_at': FMT_DATETIME,
    'error_code': FMT_INT,
    'seats_left': FMT_INT,
    'used_count': FMT_INT,
    'max_uses': FMT_INT,
    'per_user_limit': FMT_INT,
    'valid_from': FMT_DATE,
    'valid_until': FMT_DATE,
    'payment_amount': FMT_CURRENCY_UAH,
    'discount_amount': FMT_CURRENCY_UAH,
    'cpd_points_awarded': FMT_INT,
    'experience_years': FMT_INT,
}

# ----- Color fills для enum-полів ------------------------------------


def _fill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color)


EVENT_TYPE_FILLS = {
    'course': _fill('DBEAFE'),       # blue-100
    'seminar': _fill('FED7AA'),      # orange-200
    'webinar': _fill('D1FAE5'),      # green-100 (для онлайн-формату)
    'masterclass': _fill('E9D5FF'),  # purple-200
    'conference': _fill('FEF3C7'),   # yellow-100
}

EVENT_FORMAT_FILLS = {
    'online': _fill('DBEAFE'),       # blue
    'offline': _fill('D1FAE5'),      # green
    'hybrid': _fill('E9D5FF'),       # purple
}

STATUS_FILLS = {
    'draft': _fill('F3F4F6'),        # gray
    'published': _fill('DBEAFE'),    # blue
    'active': _fill('D1FAE5'),       # green
    'completed': _fill('A7F3D0'),    # darker green
    'cancelled': _fill('FECACA'),    # red
}

BOOL_TRUE_FILL = _fill('D1FAE5')     # light green
BOOL_FALSE_FILL = _fill('FEE2E2')    # light red

# Ledь-помітна зебра для непарних data-рядків. Робимо її вручну (а не
# через Excel TableStyle), бо вбудовані стилі дають занадто помітне
# банінг -- ledь-помітної опції серед них нема.
ZEBRA_FILL = _fill('FAFAFA')

# ----- Column widths (ширина = "шт. символів"). ----------------------
# Дають xlsx-у форму "зручний для перегляду", не "ALL DEFAULT 14".
COURSE_WIDTHS = {
    'id': 6,
    'slug': 32,
    'title': 55,
    'subtitle': 40,
    'short_description': 50,
    'description': 60,
    'event_type': 16,
    'base_price': 14,
    'cpd_points': 10,
    'max_participants': 12,
    'trainer_slug': 24,
    'hero_image': 50,
    'card_image': 50,
    'speaker_info': 40,
    'agenda': 40,
    'final_cta_text': 50,
    'target_audience': 50,
    'tags': 28,
    'is_active': 12,
    'is_featured': 14,
}

INSTANCE_WIDTHS = {
    'id': 6,
    'course_slug': 32,
    'start_date': 22,
    'end_date': 22,
    'event_format': 14,
    'price': 14,
    'cpd_points': 10,
    'max_participants': 12,
    'trainer_slug': 24,
    'location': 18,
    'online_link': 40,
    'status': 14,
}

PROGRAM_WIDTHS = {
    'course_slug': 32,
    'sort_order': 10,
    'heading': 40,
    'items': 70,
}

FAQ_WIDTHS = {
    'course_slug': 32,
    'question': 50,
    'answer': 70,
}

TRAINER_WIDTHS = {'slug': 28, 'full_name': 36, 'role': 50}

VALID_EVENT_TYPES = {t[0] for t in Course.EVENT_TYPES}
VALID_FORMATS = {t[0] for t in CourseInstance.FORMATS}
VALID_STATUSES = {t[0] for t in CourseInstance.STATUSES}

# key -> Ukrainian label (для відображення в xlsx).
EVENT_TYPE_LABEL = dict(Course.EVENT_TYPES)
EVENT_TYPE_KEY_BY_LABEL = {v: k for k, v in EVENT_TYPE_LABEL.items()}

FORMAT_LABEL = dict(CourseInstance.FORMATS)  # 'online' -> 'Онлайн' тощо
FORMAT_KEY_BY_LABEL = {v: k for k, v in FORMAT_LABEL.items()}

STATUS_LABEL = dict(CourseInstance.STATUSES)  # 'draft' -> 'Чернетка' тощо
STATUS_KEY_BY_LABEL = {v: k for k, v in STATUS_LABEL.items()}


def _import_dir() -> Path:
    """instance/xlsx_imports -- лежить поза static, недоступне з вебу."""
    inst = Path(current_app.instance_path)
    target = inst / 'xlsx_imports'
    target.mkdir(parents=True, exist_ok=True)
    return target


def save_uploaded_xlsx(file_storage) -> str:
    """Зберегти upload з UI у тимчасову директорію, повернути token."""
    token = uuid.uuid4().hex
    path = _import_dir() / f'{token}.xlsx'
    file_storage.save(str(path))
    return token


def get_uploaded_path(token: str) -> Path | None:
    """Знайти збережений файл за token. Захист від path traversal."""
    if not token.isalnum() or len(token) != 32:
        return None
    p = _import_dir() / f'{token}.xlsx'
    return p if p.is_file() else None


def cleanup_upload(token: str) -> None:
    p = get_uploaded_path(token)
    if p is not None:
        try:
            p.unlink()
        except OSError:
            logger.exception('Failed to remove temp xlsx %s', p)


def cleanup_stale_xlsx_uploads(max_age_minutes: int = 30) -> int:
    """Видалити завантажені для preview xlsx-файли, старші за `max_age_minutes`.

    Викликається з APScheduler-job-у. Повертає кількість видалених файлів
    (для логування).
    """
    target = _import_dir()
    cutoff = datetime.now().timestamp() - max_age_minutes * 60
    removed = 0
    for p in target.glob('*.xlsx'):
        try:
            if p.stat().st_mtime < cutoff:
                p.unlink()
                removed += 1
        except OSError:
            logger.exception('Failed to inspect/remove stale xlsx %s', p)
    if removed:
        logger.info('Cleaned up %d stale xlsx upload(s)', removed)
    return removed


def _style_header(ws, columns: list[str], labels: dict[str, str] | None = None) -> None:
    """Записати заголовки (без виставлення ширин — ширини окремо через
    `_set_column_widths`, бо вони залежать від типу контенту).
    """
    for col_idx, key in enumerate(columns, start=1):
        display = labels.get(key, key) if labels else key
        cell = ws.cell(row=1, column=col_idx, value=display)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', horizontal='left')
    ws.freeze_panes = 'A2'
    # Висота header-рядка для зручності
    ws.row_dimensions[1].height = 22


def _set_column_widths(ws, columns: list[str], widths: dict[str, int]) -> None:
    """Виставити ширину кожної колонки за мапою `widths`. Якщо ключа немає,
    використовуємо дефолт 14."""
    for col_idx, key in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, 14)


def _apply_number_formats(ws, columns: list[str], last_row: int) -> None:
    """Призначити number_format на кожну колонку (одразу на всі дата-клітинки
    від рядка 2 до last_row), якщо key є в NUMBER_FORMATS."""
    if last_row < 2:
        return
    for col_idx, key in enumerate(columns, start=1):
        fmt = NUMBER_FORMATS.get(key)
        if not fmt:
            continue
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx).number_format = fmt


def _apply_table_style(ws, columns: list[str], table_name: str, last_data_row: int) -> None:
    """Перетворити діапазон A1:<last_col><last_data_row> на Excel-Table.

    Дає авто-фільтри в заголовку + іменований range. Зебру самі малюємо
    через `_apply_zebra` (бо вбудоване Excel-банінг -- надто помітне).
    """
    if last_data_row < 2:
        return
    last_col = get_column_letter(len(columns))
    ref = f'A1:{last_col}{last_data_row}'
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleLight1',   # майже-білий, без сильних кольорів
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,      # ВЛАСНА зебра нижче
        showColumnStripes=False,
    )
    ws.add_table(table)


def _apply_zebra(ws, n_cols: int, first_data_row: int, last_data_row: int) -> None:
    """Заповнити кожен 2-й data-рядок ledь-помітним сірим. Викликати
    ПЕРЕД призначенням enum-fills, щоб кольорові клітинки (event_type,
    status, is_active, ...) перекривали zebra-fill своїм кольором.
    """
    if last_data_row < first_data_row:
        return
    # Колір на другому, четвертому, шостому data-рядку
    # (visually -- стовпчик `1st row=white, 2nd row=gray, 3rd=white...`).
    for row in range(first_data_row + 1, last_data_row + 1, 2):
        for col in range(1, n_cols + 1):
            ws.cell(row, col).fill = ZEBRA_FILL


# Excel не приймає керуючі символи (крім \t \n \r): один \x07 у нотатці
# піднімав IllegalCharacterError і валив ВЕСЬ експорт, а не рядок.
_CONTROL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
# Ліміт довжини комірки в xlsx; довший рядок Excel вважає файл пошкодженим.
MAX_CELL_LENGTH = 32767


def _safe_text(value: str) -> str:
    """Прибрати керуючі символи й обрізати до ліміту комірки."""
    text = _CONTROL_CHARS_RE.sub('', value)
    if len(text) > MAX_CELL_LENGTH:
        text = text[:MAX_CELL_LENGTH - 1] + '…'
    return text


def write_cell(ws, row, column, value):
    """Записати комірку так, щоб дані лишились ДАНИМИ.

    openpyxl визначає тип за вмістом рядка: значення, що починається з '=',
    стає ЖИВОЮ формулою (data_type 'f'), а '#N/A' та решта ERROR_CODES --
    коміркою-помилкою ('e'). У звіти йде чужий текст (місце роботи, нотатки,
    відповіді SMTP), тобто учасник міг би керувати вмістом файлу, який
    відкриє менеджер. Повертаємо тип у 's': значення зберігається дослівно,
    Excel показує текст.
    """
    if isinstance(value, str):
        value = _safe_text(value)
    cell = ws.cell(row=row, column=column, value=value)
    if cell.data_type in ('f', 'e'):
        cell.data_type = 's'
    return cell


def _to_kyiv_naive(dt):
    """Зняти TZ, попередньо переконвертувавши в Київ. openpyxl попереджує
    про tz-aware datetimes; у клітинці маємо bare datetime, який Excel
    розуміє як «локальний час»."""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        return dt.astimezone(KYIV).replace(tzinfo=None)
    return dt


def _str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_lines(value: list | None) -> str:
    """JSON list -> текст по 1 елементу на рядок."""
    if not value:
        return ''
    return '\n'.join(str(x) for x in value)


def _from_lines(value: str | None) -> list[str]:
    if not value:
        return []
    return [line.strip() for line in str(value).splitlines() if line.strip()]


def _bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {'true', 'yes', 'y', '1', 'так', 'да'}


def _decimal(v) -> Decimal | None:
    if v is None or v == '':
        return None
    try:
        return Decimal(str(v).replace(' ', '').replace(',', '.'))
    except (InvalidOperation, ValueError):
        raise ValueError(f'не число: {v!r}')


def _int(v) -> int | None:
    if v is None or v == '':
        return None
    try:
        return int(float(str(v).replace(' ', '').replace(',', '.')))
    except (TypeError, ValueError):
        raise ValueError(f'не ціле число: {v!r}')


def _dt(v) -> datetime | None:
    """Прийняти або datetime (openpyxl auto-parses), або ISO-рядок."""
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        dt = v
    else:
        try:
            dt = datetime.fromisoformat(str(v))
        except ValueError:
            raise ValueError(f'неможливо розпарсити дату: {v!r}')
    if dt.tzinfo is None:
        # припускаємо Київ (UTC+3), щоб збігалось з seed-розкладом
        dt = dt.replace(tzinfo=KYIV)
    return dt


# ======================================================================
# COURSES
# ======================================================================

COURSE_COLS = [
    'id', 'slug', 'title', 'subtitle', 'short_description', 'description',
    'event_type', 'base_price', 'cpd_points', 'max_participants',
    'trainer_slug', 'hero_image', 'card_image', 'speaker_info', 'agenda',
    'final_cta_text', 'target_audience', 'tags', 'is_active', 'is_featured',
]

# Українські назви колонок для заголовків xlsx. Імпорт приймає обидва
# варіанти (англ. internal key АБО українську підпис) -- це гарантує
# що файли, експортовані раніше зі старими заголовками, ще можна
# завантажувати.
COURSE_LABELS = {
    'id': 'ID',
    'slug': 'Slug (URL)',
    'title': 'Назва',
    'subtitle': 'Підзаголовок',
    'short_description': 'Короткий опис',
    'description': 'Повний опис',
    'event_type': 'Тип',
    'base_price': 'Ціна (грн)',
    'cpd_points': 'Бали БПР',
    'max_participants': 'Макс. учасників',
    'trainer_slug': 'Тренер',
    'hero_image': 'Hero-зображення',
    'card_image': 'Зображення картки',
    'speaker_info': 'Інфо про спікера',
    'agenda': 'Програма (опис)',
    'final_cta_text': 'Фінальний заклик',
    'target_audience': 'Цільова аудиторія',
    'tags': 'Теги',
    'is_active': 'Активний',
    'is_featured': 'Рекомендований',
}

# Колонки, додані після того, як менеджери вже мали на руках експорти.
# Їх відсутність не ламає імпорт старого файлу, а поле лишається як у БД --
# інакше кожне нове поле знецінювало б усі раніше збережені файли.
OPTIONAL_COURSE_COLS = ('final_cta_text',)

PROGRAM_COLS = ['course_slug', 'sort_order', 'heading', 'items']
PROGRAM_LABELS = {
    'course_slug': 'Курс (slug)',
    'sort_order': 'Порядок',
    'heading': 'Заголовок',
    'items': 'Пункти',
}

FAQ_COLS = ['course_slug', 'question', 'answer']
FAQ_LABELS = {
    'course_slug': 'Курс (slug)',
    'question': 'Запитання',
    'answer': 'Відповідь',
}

# Назви sheet-ів. Експортуємо в українській, парсинг приймає обидва.
SHEET_ALIASES = {
    'courses': ['Курси', 'Courses'],
    'program_blocks': ['Блоки програми', 'Program blocks'],
    'faq': ['FAQ'],
    'instances': ['Розклад', 'Instances'],
    'participants': ['Учасники', 'Participants'],
    'materials': ['Матеріали', 'Materials'],
}


def _find_sheet(wb, key: str):
    """Знайти sheet за будь-яким з прийнятних псевдонімів."""
    for name in SHEET_ALIASES.get(key, []):
        if name in wb.sheetnames:
            return wb[name]
    return None


# Кількість рядків, на які поширюється data-validation drop-down у
# колонці trainer_slug. Менеджер може дописувати нові рядки знизу --
# валідація все одно покриватиме. 500 з запасом.
_DROPDOWN_BUFFER_ROWS = 500
_TRAINERS_SHEET_NAME = 'Тренери'


def _add_trainers_sheet(wb) -> int:
    """Додати reference-sheet з активними тренерами. ПІБ йде в колонці A,
    щоб саме воно потрапляло у drop-down тренерів. Slug -- у колонці B
    (для довідки). Повертає номер останнього рядка з даними."""
    ws = wb.create_sheet(_TRAINERS_SHEET_NAME)
    cols = ['full_name', 'slug', 'role']
    _style_header(
        ws,
        cols,
        {'full_name': 'ПІБ (значення)', 'slug': 'Slug', 'role': 'Посада'},
    )

    trainers = (
        Trainer.query.filter_by(is_active=True)
        .order_by(Trainer.full_name)
        .all()
    )
    for row_idx, t in enumerate(trainers, start=2):
        ws.cell(row=row_idx, column=1, value=t.full_name).alignment = WRAP
        ws.cell(row=row_idx, column=2, value=t.slug)
        ws.cell(row=row_idx, column=3, value=t.role or '').alignment = WRAP

    widths = {'full_name': 36, 'slug': 28, 'role': 50}
    _set_column_widths(ws, cols, widths)
    _apply_zebra(ws, len(cols), first_data_row=2, last_data_row=1 + len(trainers))
    _apply_table_style(ws, cols, 'tblTrainers', last_data_row=1 + len(trainers))
    return 1 + len(trainers)


def _add_inline_dropdown(ws, column_key: str, columns: list[str],
                         options: list[str], last_data_row: int,
                         title: str = '', hint: str = '') -> None:
    """Прикріпити drop-down зі статичним списком значень.

    Використовується для невеликих enum-полів (event_type, формат, статус).
    Excel-обмеження inline-list у formula1 -- 255 символів; для довших
    списків потрібен окремий sheet з reference-значеннями.
    """
    if not options:
        return
    col_letter = get_column_letter(columns.index(column_key) + 1)
    # Inline-list у formula1 має бути обгорнутий лапками й розділений комами.
    formula = '"' + ','.join(options) + '"'
    if len(formula) > 255:
        # Мовчки віддати файл, який Excel вважає пошкодженим, гірше, ніж
        # віддати його без цієї випадайки.
        logger.warning(
            'Drop-down для %r пропущено: список %s символів (ліміт Excel 255). '
            'Для довших списків потрібен reference-sheet.',
            column_key, len(formula),
        )
        return
    dv = DataValidation(
        type='list',
        formula1=formula,
        allow_blank=False,
        showDropDown=False,  # False у XML = ПОКАЗУВАТИ стрілочку
        errorStyle='stop',
        error=f'Оберіть значення зі списку: {", ".join(options)}',
        errorTitle='Невалідне значення',
        prompt=hint,
        promptTitle=title,
    )
    final_row = max(last_data_row, 1) + _DROPDOWN_BUFFER_ROWS
    dv.add(f'{col_letter}2:{col_letter}{final_row}')
    ws.add_data_validation(dv)


def _add_trainer_dropdown(ws, column_key: str, columns: list[str],
                          last_data_row: int, trainers_last_row: int) -> None:
    """Прикріпити data-validation drop-down з тренерами до вказаної
    колонки. range покриває існуючі рядки + буфер для додавання нових.
    """
    if trainers_last_row < 2:  # порожній список тренерів
        return
    col_letter = get_column_letter(columns.index(column_key) + 1)
    formula = f"='{_TRAINERS_SHEET_NAME}'!$A$2:$A${trainers_last_row}"
    dv = DataValidation(
        type='list',
        formula1=formula,
        allow_blank=True,
        # showDropDown у OOXML інвертоване: False = ПОКАЗУВАТИ стрілочку
        showDropDown=False,
        errorStyle='warning',
        error='Тренер з таким slug відсутній у sheet "Тренери".',
        errorTitle='Невідомий тренер',
        prompt='Оберіть тренера зі списку (натисніть стрілочку)',
        promptTitle='Тренер',
    )
    final_row = max(last_data_row, 1) + _DROPDOWN_BUFFER_ROWS
    dv.add(f'{col_letter}2:{col_letter}{final_row}')
    ws.add_data_validation(dv)


@dataclass
class CourseChange:
    slug: str
    action: str  # 'create' | 'update' | 'unchanged' | 'error'
    fields_changed: list[str] = field(default_factory=list)
    error: str | None = None


@dataclass
class CoursesImportPlan:
    courses: list[dict] = field(default_factory=list)   # parsed rows
    program_blocks: dict[str, list[dict]] = field(default_factory=dict)  # slug -> list
    faq: dict[str, list[dict]] = field(default_factory=dict)
    changes: list[CourseChange] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    # slug в файлі (для оцінки яких program_blocks/faq REPLACE)
    program_slugs_in_file: set[str] = field(default_factory=set)
    faq_slugs_in_file: set[str] = field(default_factory=set)

    @property
    def is_valid(self) -> bool:
        return not self.errors

    @property
    def counts(self) -> dict[str, int]:
        c = {'create': 0, 'update': 0, 'unchanged': 0, 'error': 0}
        for ch in self.changes:
            c[ch.action] = c.get(ch.action, 0) + 1
        return c


def export_courses_xlsx(active: str = 'all') -> io.BytesIO:
    """Згенерувати xlsx з 3 sheet: Courses / Program blocks / FAQ.

    Параметри:
      active: 'all' | 'true' | 'false' -- фільтр за полем is_active.
              Дефолтно 'all' (історична поведінка -- усі курси).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Курси'
    _style_header(ws, COURSE_COLS, COURSE_LABELS)

    # Тренер у клітинці — ПІБ (Ukrainian). На імпорті повертаємо в slug.
    trainer_name_by_id = {t.id: t.full_name for t in Trainer.query.all()}

    q = Course.query.order_by(Course.id)
    if active == 'true':
        q = q.filter(Course.is_active.is_(True))
    elif active == 'false':
        q = q.filter(Course.is_active.is_(False))
    courses = q.all()
    for row_idx, c in enumerate(courses, start=2):
        values = [
            c.id,
            c.slug,
            c.title or '',
            c.subtitle or '',
            c.short_description or '',
            c.description or '',
            EVENT_TYPE_LABEL.get(c.event_type, c.event_type or ''),
            float(c.base_price) if c.base_price is not None else 0,
            c.cpd_points,
            c.max_participants,
            trainer_name_by_id.get(c.trainer_id, '') if c.trainer_id else '',
            # Експортуємо ОСНОВНИЙ media-URL (не варіант) -> резолвиться назад
            # у реєстр за file_path на імпорті (_resolve_media_id).
            c.hero_media.url if c.hero_media else '',
            c.card_media.url if c.card_media else '',
            c.speaker_info or '',
            c.agenda or '',
            c.final_cta_text or '',
            _to_lines(c.target_audience),
            _to_lines(c.tags),
            bool(c.is_active),
            bool(c.is_featured),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = write_cell(ws, row_idx, col_idx, v)
            cell.alignment = WRAP

    courses_last_row = ws.max_row

    # ЗЕБРА перед enum-кольорами, щоб enum-fills перекрили її на своїх клітинках.
    _apply_zebra(ws, len(COURSE_COLS), first_data_row=2, last_data_row=courses_last_row)

    # ----- Кольори за значенням -----------------------------------------
    et_col = COURSE_COLS.index('event_type') + 1
    ia_col = COURSE_COLS.index('is_active') + 1
    if_col = COURSE_COLS.index('is_featured') + 1
    for row_idx, c in enumerate(courses, start=2):
        if c.event_type and c.event_type in EVENT_TYPE_FILLS:
            ws.cell(row=row_idx, column=et_col).fill = EVENT_TYPE_FILLS[c.event_type]
        ws.cell(row=row_idx, column=ia_col).fill = (
            BOOL_TRUE_FILL if c.is_active else BOOL_FALSE_FILL
        )
        if c.is_featured:
            ws.cell(row=row_idx, column=if_col).fill = BOOL_TRUE_FILL

    _set_column_widths(ws, COURSE_COLS, COURSE_WIDTHS)
    _apply_number_formats(ws, COURSE_COLS, courses_last_row)

    # Program blocks
    ws_p = wb.create_sheet('Блоки програми')
    _style_header(ws_p, PROGRAM_COLS, PROGRAM_LABELS)
    row_idx = 2
    for c in courses:
        for b in sorted(c.program_blocks, key=lambda x: x.sort_order or 0):
            ws_p.cell(row=row_idx, column=1, value=c.slug)
            ws_p.cell(row=row_idx, column=2, value=b.sort_order or 0)
            ws_p.cell(row=row_idx, column=3, value=b.heading or '').alignment = WRAP
            ws_p.cell(row=row_idx, column=4, value=_to_lines(b.items)).alignment = WRAP
            row_idx += 1
    program_last_row = ws_p.max_row
    _apply_zebra(ws_p, len(PROGRAM_COLS), first_data_row=2, last_data_row=program_last_row)
    _set_column_widths(ws_p, PROGRAM_COLS, PROGRAM_WIDTHS)
    _apply_number_formats(ws_p, PROGRAM_COLS, program_last_row)

    # FAQ
    ws_f = wb.create_sheet('FAQ')
    _style_header(ws_f, FAQ_COLS, FAQ_LABELS)
    row_idx = 2
    for c in courses:
        for item in (c.faq or []):
            if not isinstance(item, dict):
                continue
            ws_f.cell(row=row_idx, column=1, value=c.slug)
            ws_f.cell(row=row_idx, column=2, value=item.get('question') or '').alignment = WRAP
            ws_f.cell(row=row_idx, column=3, value=item.get('answer') or '').alignment = WRAP
            row_idx += 1
    faq_last_row = ws_f.max_row
    _apply_zebra(ws_f, len(FAQ_COLS), first_data_row=2, last_data_row=faq_last_row)
    _set_column_widths(ws_f, FAQ_COLS, FAQ_WIDTHS)

    # Reference sheet з тренерами (вже з Table) + drop-down у колонці trainer_slug.
    trainers_last_row = _add_trainers_sheet(wb)
    _add_trainer_dropdown(
        ws, 'trainer_slug', COURSE_COLS,
        last_data_row=courses_last_row,
        trainers_last_row=trainers_last_row,
    )

    # Drop-down для типу заходу.
    _add_inline_dropdown(
        ws, 'event_type', COURSE_COLS,
        options=[label for _key, label in Course.EVENT_TYPES],
        last_data_row=courses_last_row,
        title='Тип заходу',
        hint='Оберіть зі списку: Семінар, Вебінар, Курс, Майстер-клас, Конференція',
    )

    # Excel Tables (forматовані з зеброю + auto-filter).
    _apply_table_style(ws, COURSE_COLS, 'tblCourses', courses_last_row)
    _apply_table_style(ws_p, PROGRAM_COLS, 'tblProgramBlocks', program_last_row)
    _apply_table_style(ws_f, FAQ_COLS, 'tblFAQ', faq_last_row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _read_sheet(ws, columns: list[str], labels: dict[str, str] | None = None,
                optional: tuple[str, ...] = ()) -> list[dict]:
    """Прочитати sheet у list[dict].

    Заголовки приймаються або як internal key (англ., 'slug'), або як
    українські labels (з `labels`), для зворотньої сумісності зі старими
    xlsx-файлами.

    `optional` -- колонки, відсутність яких НЕ є помилкою: так у формат можна
    додати нове поле, не ламаючи імпорт файлів, експортованих раніше. Ключів
    відсутніх опційних колонок у рядку не буде взагалі (а не None), щоб
    виклик міг відрізнити "у файлі порожньо" від "колонки не було" і не
    занулив наявне значення.

    Заголовок і рядки читаємо одним проходом ітератора, БЕЗ ws.max_row: у
    read-only режимі він дорівнює None, якщо у файлі немає запису
    <dimension> (так зберігають Google Sheets і частина конвертерів), і
    порівняння з числом падало TypeError на цілком нормальному файлі.
    Побічно це виправляє й інше: перевірка колонок більше не пропускається
    для листа без рядків даних -- раніше лист із чужими заголовками тихо
    читався як "порожньо" замість помилки.
    """
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return []

    # accepted: будь-яка валідна назва заголовка -> internal key
    accepted: dict[str, str] = {}
    for key in columns:
        accepted[key] = key
        accepted[key.lower()] = key
        if labels and key in labels:
            ua = labels[key]
            accepted[ua] = key
            accepted[ua.lower()] = key

    col_idx: dict[str, int] = {}
    for i, hv in enumerate(header):
        if hv is None:
            continue
        key = accepted.get(str(hv).strip()) or accepted.get(str(hv).strip().lower())
        if key:
            col_idx[key] = i + 1

    missing = [c for c in columns if c not in col_idx and c not in optional]
    if missing:
        pretty = [(labels.get(k, k) if labels else k) for k in missing]
        raise ValueError(
            f'Sheet "{ws.title}": бракує колонок: {", ".join(pretty)}'
        )

    present = [c for c in columns if c in col_idx]
    rows = []
    for row in rows_iter:
        if not any(v is not None and str(v).strip() for v in row):
            continue  # повністю порожній рядок
        # Рядок може бути коротшим за заголовок (обрізані хвостові порожні
        # клітинки) -- беремо None замість IndexError.
        d = {
            name: (row[col_idx[name] - 1] if col_idx[name] <= len(row) else None)
            for name in present
        }
        rows.append(d)
    return rows


def parse_courses_xlsx(path: Path) -> CoursesImportPlan:
    plan = CoursesImportPlan()
    try:
        wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    except Exception as exc:
        plan.errors.append(f'Не вдалося відкрити xlsx: {exc}')
        return plan

    # ---- Courses sheet ----
    ws_c = _find_sheet(wb, 'courses')
    if ws_c is None:
        plan.errors.append('Відсутній sheet "Курси"')
        return plan

    try:
        rows = _read_sheet(ws_c, COURSE_COLS, COURSE_LABELS,
                           optional=OPTIONAL_COURSE_COLS)
    except ValueError as exc:
        plan.errors.append(str(exc))
        return plan

    _all_trainers = Trainer.query.all()
    trainer_id_by_slug = {t.slug: t.id for t in _all_trainers}
    trainer_id_by_name = {t.full_name: t.id for t in _all_trainers}
    existing_by_id = {c.id: c for c in Course.query.all()}
    existing_by_slug = {c.slug: c for c in existing_by_id.values()}

    seen_slugs = set()
    for line_no, raw in enumerate(rows, start=2):
        try:
            slug = _str(raw.get('slug'))
            if not slug:
                raise ValueError('порожній slug')
            if slug in seen_slugs:
                raise ValueError(f'дублюючий slug у файлі: {slug!r}')
            seen_slugs.add(slug)

            event_type_raw = _str(raw.get('event_type')) or 'course'
            # Приймаємо і англ. internal key ('course'), і українську назву
            # з drop-down ('Курс'). Нормалізуємо у key.
            event_type = EVENT_TYPE_KEY_BY_LABEL.get(event_type_raw, event_type_raw)
            if event_type not in VALID_EVENT_TYPES:
                allowed = sorted(VALID_EVENT_TYPES) + sorted(EVENT_TYPE_KEY_BY_LABEL.keys())
                raise ValueError(
                    f'event_type={event_type_raw!r} – допустимі: {allowed}'
                )

            # Колонка "Тренер" може містити або slug (старі файли), або ПІБ
            # (новий експорт + drop-down). Спершу шукаємо за slug, потім за
            # full_name -- так покриваємо обидва формати.
            trainer_raw = _str(raw.get('trainer_slug'))
            trainer_id = None
            if trainer_raw:
                trainer_id = (
                    trainer_id_by_slug.get(trainer_raw)
                    or trainer_id_by_name.get(trainer_raw)
                )
                if trainer_id is None:
                    raise ValueError(
                        f'тренера {trainer_raw!r} не знайдено '
                        f'(ні за slug, ні за ПІБ)'
                    )

            parsed = {
                'id': _int(raw.get('id')),
                'slug': slug,
                'title': _str(raw.get('title')) or '',
                'subtitle': _str(raw.get('subtitle')),
                'short_description': _str(raw.get('short_description')),
                'description': _str(raw.get('description')),
                'event_type': event_type,
                'base_price': _decimal(raw.get('base_price')) or Decimal(0),
                'cpd_points': _int(raw.get('cpd_points')),
                'max_participants': _int(raw.get('max_participants')),
                'trainer_id': trainer_id,
                'hero_image': _str(raw.get('hero_image')),
                'card_image': _str(raw.get('card_image')),
                'speaker_info': _str(raw.get('speaker_info')),
                'agenda': _str(raw.get('agenda')),
                'target_audience': _from_lines(raw.get('target_audience')),
                'tags': _from_lines(raw.get('tags')),
                'is_active': _bool(raw.get('is_active')),
                'is_featured': _bool(raw.get('is_featured')),
            }
            # Опційні колонки кладемо в parsed ЛИШЕ якщо вони були у файлі:
            # відсутність колонки має лишити поле як є, а не занулити його.
            for opt in OPTIONAL_COURSE_COLS:
                if opt in raw:
                    parsed[opt] = _str(raw.get(opt))

            if not parsed['title']:
                raise ValueError('порожній title')
            if parsed['base_price'] < 0:
                raise ValueError('base_price < 0')

            # знайти існуючий: id має пріоритет, потім slug
            existing = None
            if parsed['id'] is not None:
                existing = existing_by_id.get(parsed['id'])
                if existing is None:
                    raise ValueError(
                        f'id={parsed["id"]} не існує в БД '
                        f'(використайте порожній id для нового курсу)'
                    )
                if existing.slug != slug and slug in existing_by_slug:
                    raise ValueError(
                        f'slug={slug!r} вже зайнятий іншим курсом'
                    )
            else:
                existing = existing_by_slug.get(slug)

            plan.courses.append({'parsed': parsed, 'existing_id': existing.id if existing else None})

            if existing is None:
                plan.changes.append(CourseChange(slug=slug, action='create'))
            else:
                diff = _diff_course(existing, parsed, trainer_id_by_slug)
                if diff:
                    plan.changes.append(CourseChange(
                        slug=slug, action='update', fields_changed=diff,
                    ))
                else:
                    plan.changes.append(CourseChange(slug=slug, action='unchanged'))
        except Exception as exc:
            plan.errors.append(f'Рядок {line_no} (Courses): {exc}')
            plan.changes.append(CourseChange(
                slug=_str(raw.get('slug')) or f'#{line_no}',
                action='error',
                error=str(exc),
            ))

    # ---- Program blocks sheet ----
    ws_p = _find_sheet(wb, 'program_blocks')
    if ws_p is not None:
        try:
            p_rows = _read_sheet(ws_p, PROGRAM_COLS, PROGRAM_LABELS)
        except ValueError as exc:
            plan.errors.append(str(exc))
            p_rows = []
        for line_no, raw in enumerate(p_rows, start=2):
            try:
                slug = _str(raw.get('course_slug'))
                if not slug:
                    raise ValueError('порожній course_slug')
                heading = _str(raw.get('heading'))
                if not heading:
                    raise ValueError('порожній heading')
                sort_order = _int(raw.get('sort_order')) or 0
                items = _from_lines(raw.get('items'))
                plan.program_blocks.setdefault(slug, []).append({
                    'sort_order': sort_order,
                    'heading': heading,
                    'items': items,
                })
                plan.program_slugs_in_file.add(slug)
            except Exception as exc:
                plan.errors.append(f'Рядок {line_no} (Program blocks): {exc}')

    # ---- FAQ sheet ----
    ws_f = _find_sheet(wb, 'faq')
    if ws_f is not None:
        try:
            f_rows = _read_sheet(ws_f, FAQ_COLS, FAQ_LABELS)
        except ValueError as exc:
            plan.errors.append(str(exc))
            f_rows = []
        for line_no, raw in enumerate(f_rows, start=2):
            try:
                slug = _str(raw.get('course_slug'))
                if not slug:
                    raise ValueError('порожній course_slug')
                question = _str(raw.get('question'))
                if not question:
                    raise ValueError('порожнє question')
                answer = _str(raw.get('answer'))
                plan.faq.setdefault(slug, []).append({
                    'question': question, 'answer': answer or '',
                })
                plan.faq_slugs_in_file.add(slug)
            except Exception as exc:
                plan.errors.append(f'Рядок {line_no} (FAQ): {exc}')

    # перевірити, що course_slug у program/faq sheets існує у Courses
    # sheet або в БД (програмні блоки не повинні висіти без курсу)
    db_slugs = set(existing_by_slug.keys())
    file_slugs = {c['parsed']['slug'] for c in plan.courses}
    all_known = db_slugs | file_slugs
    for slug in plan.program_slugs_in_file:
        if slug not in all_known:
            plan.errors.append(
                f'Program blocks: course_slug={slug!r} не існує ні в xlsx, '
                f'ні в БД'
            )
    for slug in plan.faq_slugs_in_file:
        if slug not in all_known:
            plan.errors.append(
                f'FAQ: course_slug={slug!r} не існує ні в xlsx, ні в БД'
            )

    return plan


def _diff_course(existing: Course, parsed: dict, trainer_id_by_slug: dict) -> list[str]:
    """Повернути список імен змінених полів. Порівняння помилкостійке."""
    changed = []
    # hero_image/card_image тут НЕМАЄ свідомо: після переходу на медіа-реєстр
    # (фаза 6) у Course лишились тільки *_media_id, і getattr по старій назві
    # кидав AttributeError -- через це будь-який рядок з ІСНУЮЧИМ курсом
    # ставав помилкою і файл цілком відхилявся. Зображення порівнюємо нижче,
    # резолвленими id.
    fields = [
        'title', 'subtitle', 'short_description', 'description', 'event_type',
        'cpd_points', 'max_participants', 'trainer_id',
        'speaker_info', 'agenda', 'is_active', 'is_featured',
    ]
    for f in fields:
        if (getattr(existing, f) or None) != (parsed[f] or None) and not (
            (getattr(existing, f) in ('', None)) and (parsed[f] in ('', None))
        ):
            changed.append(f)
    if (existing.base_price or Decimal(0)) != parsed['base_price']:
        changed.append('base_price')
    if (existing.target_audience or []) != parsed['target_audience']:
        changed.append('target_audience')
    if (existing.tags or []) != parsed['tags']:
        changed.append('tags')
    # У файлі -- людиночитний URL; порівнюємо так само, як apply записує.
    for media_col, url_key in (('hero_media_id', 'hero_image'),
                               ('card_media_id', 'card_image')):
        if getattr(existing, media_col) != _resolve_media_id(parsed[url_key]):
            changed.append(url_key)
    for opt in OPTIONAL_COURSE_COLS:
        if opt in parsed and (getattr(existing, opt) or None) != (parsed[opt] or None):
            changed.append(opt)
    return changed


def apply_courses_plan(plan: CoursesImportPlan) -> dict:
    """Atomic upsert. Очікує plan.is_valid==True."""
    if not plan.is_valid:
        return {'ok': False, 'reason': 'plan has errors'}

    created = 0
    updated = 0
    blocks_touched = 0
    faq_touched = 0

    try:
        # 1) courses upsert
        for item in plan.courses:
            p = item['parsed']
            ex_id = item['existing_id']
            if ex_id is None:
                course = Course(slug=p['slug'])
                db.session.add(course)
                created += 1
            else:
                course = db.session.get(Course, ex_id)
                updated += 1

            course.title = p['title']
            course.slug = p['slug']
            course.subtitle = p['subtitle']
            course.short_description = p['short_description']
            course.description = p['description']
            course.event_type = p['event_type']
            course.base_price = p['base_price']
            course.cpd_points = p['cpd_points']
            course.max_participants = p['max_participants']
            course.trainer_id = p['trainer_id']
            course.hero_media_id = _resolve_media_id(p['hero_image'])
            course.card_media_id = _resolve_media_id(p['card_image'])
            course.speaker_info = p['speaker_info']
            course.agenda = p['agenda']
            course.target_audience = p['target_audience']
            course.tags = p['tags']
            course.is_active = p['is_active']
            course.is_featured = p['is_featured']
            for opt in OPTIONAL_COURSE_COLS:
                if opt in p:
                    setattr(course, opt, p[opt])

        db.session.flush()

        # 2) program blocks: REPLACE для курсів, чий slug згаданий у sheet
        slug_to_course = {c.slug: c for c in Course.query.all()}
        for slug in plan.program_slugs_in_file:
            course = slug_to_course.get(slug)
            if course is None:
                continue
            ProgramBlock.query.filter_by(course_id=course.id).delete()
            blocks = plan.program_blocks.get(slug, [])
            for b in blocks:
                db.session.add(ProgramBlock(
                    course_id=course.id,
                    heading=b['heading'],
                    items=b['items'],
                    sort_order=b['sort_order'],
                ))
                blocks_touched += 1

        # 3) faq: REPLACE як JSON-stored у Course.faq
        for slug in plan.faq_slugs_in_file:
            course = slug_to_course.get(slug)
            if course is None:
                continue
            faq_list = plan.faq.get(slug, [])
            course.faq = faq_list
            faq_touched += len(faq_list)

        db.session.commit()
        return {
            'ok': True,
            'created': created,
            'updated': updated,
            'blocks_touched': blocks_touched,
            'faq_touched': faq_touched,
        }
    except Exception as exc:
        db.session.rollback()
        logger.exception('apply_courses_plan failed')
        return {'ok': False, 'reason': str(exc)}


# ======================================================================
# COURSE INSTANCES (розклад)
# ======================================================================

INSTANCE_COLS = [
    'id', 'course_slug', 'start_date', 'end_date', 'event_format',
    'price', 'cpd_points', 'max_participants', 'trainer_slug',
    'location', 'online_link', 'status',
]

INSTANCE_LABELS = {
    'id': 'ID',
    'course_slug': 'Курс (slug)',
    'start_date': 'Початок',
    'end_date': 'Кінець',
    'event_format': 'Формат',
    'price': 'Ціна (грн)',
    'cpd_points': 'Бали БПР',
    'max_participants': 'Макс. учасників',
    'trainer_slug': 'Тренер',
    'location': 'Локація',
    'online_link': 'Онлайн-лінк',
    'status': 'Статус',
}


@dataclass
class InstanceChange:
    line_no: int
    course_slug: str
    start_date: str
    action: str  # 'create' | 'update' | 'unchanged' | 'error'
    fields_changed: list[str] = field(default_factory=list)
    error: str | None = None


@dataclass
class InstancesImportPlan:
    instances: list[dict] = field(default_factory=list)
    changes: list[InstanceChange] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors

    @property
    def counts(self) -> dict[str, int]:
        c = {'create': 0, 'update': 0, 'unchanged': 0, 'error': 0}
        for ch in self.changes:
            c[ch.action] = c.get(ch.action, 0) + 1
        return c


def export_instances_xlsx(
    year: int | None = None,
    upcoming_only: bool = False,
    status: str | None = None,
) -> io.BytesIO:
    """Експорт розкладу. Усі фільтри необов'язкові.

    Параметри:
      year: int -- лише проведення з start_date у вказаному році.
      upcoming_only: True -- лише з start_date >= зараз.
      status: 'draft'|'published'|'active'|'completed'|'cancelled' -- фільтр статусу.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Розклад'
    _style_header(ws, INSTANCE_COLS, INSTANCE_LABELS)

    course_slug_by_id = {c.id: c.slug for c in Course.query.all()}
    trainer_name_by_id = {t.id: t.full_name for t in Trainer.query.all()}

    q = CourseInstance.query.order_by(CourseInstance.start_date)
    if year:
        start = datetime(year, 1, 1, tzinfo=timezone.utc)
        end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        q = q.filter(
            CourseInstance.start_date >= start,
            CourseInstance.start_date < end,
        )
    if upcoming_only:
        q = q.filter(CourseInstance.start_date >= datetime.now(timezone.utc))
    if status:
        q = q.filter(CourseInstance.status == status)
    instances = q.all()
    for row_idx, i in enumerate(instances, start=2):
        values = [
            i.id,
            course_slug_by_id.get(i.course_id, ''),
            _to_kyiv_naive(i.start_date),
            _to_kyiv_naive(i.end_date),
            FORMAT_LABEL.get(i.event_format, i.event_format or ''),
            float(i.price) if i.price is not None else None,
            i.cpd_points,
            i.max_participants,
            trainer_name_by_id.get(i.trainer_id, '') if i.trainer_id else '',
            i.location or '',
            i.online_link or '',
            STATUS_LABEL.get(i.status, i.status or 'draft'),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = write_cell(ws, row_idx, col_idx, v)
            cell.alignment = WRAP

    instances_last_row = ws.max_row

    # ЗЕБРА до enum-кольорів, щоб ті перекрили її.
    _apply_zebra(ws, len(INSTANCE_COLS), first_data_row=2, last_data_row=instances_last_row)

    # ----- Кольори за значенням -----------------------------------------
    fmt_col = INSTANCE_COLS.index('event_format') + 1
    st_col = INSTANCE_COLS.index('status') + 1
    for row_idx, i in enumerate(instances, start=2):
        if i.event_format in EVENT_FORMAT_FILLS:
            ws.cell(row=row_idx, column=fmt_col).fill = EVENT_FORMAT_FILLS[i.event_format]
        if i.status in STATUS_FILLS:
            ws.cell(row=row_idx, column=st_col).fill = STATUS_FILLS[i.status]

    _set_column_widths(ws, INSTANCE_COLS, INSTANCE_WIDTHS)
    _apply_number_formats(ws, INSTANCE_COLS, instances_last_row)

    # Reference sheet з тренерами + drop-down у колонці trainer_slug розкладу.
    trainers_last_row = _add_trainers_sheet(wb)
    _add_trainer_dropdown(
        ws, 'trainer_slug', INSTANCE_COLS,
        last_data_row=instances_last_row,
        trainers_last_row=trainers_last_row,
    )

    # Drop-down для формату та статусу — українські labels.
    _add_inline_dropdown(
        ws, 'event_format', INSTANCE_COLS,
        options=[label for _key, label in CourseInstance.FORMATS],
        last_data_row=instances_last_row,
        title='Формат',
        hint='Оберіть формат: Онлайн / Офлайн / Гібрид',
    )
    _add_inline_dropdown(
        ws, 'status', INSTANCE_COLS,
        options=[label for _key, label in CourseInstance.STATUSES],
        last_data_row=instances_last_row,
        title='Статус',
        hint='Чернетка / Опубліковано / Активний / Завершено / Скасовано',
    )

    # Excel Table style.
    _apply_table_style(ws, INSTANCE_COLS, 'tblSchedule', instances_last_row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def parse_instances_xlsx(path: Path) -> InstancesImportPlan:
    plan = InstancesImportPlan()
    try:
        wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    except Exception as exc:
        plan.errors.append(f'Не вдалося відкрити xlsx: {exc}')
        return plan

    ws_i = _find_sheet(wb, 'instances')
    if ws_i is None:
        plan.errors.append('Відсутній sheet "Розклад"')
        return plan

    try:
        rows = _read_sheet(ws_i, INSTANCE_COLS, INSTANCE_LABELS)
    except ValueError as exc:
        plan.errors.append(str(exc))
        return plan

    course_id_by_slug = {c.slug: c.id for c in Course.query.all()}
    _all_trainers = Trainer.query.all()
    trainer_id_by_slug = {t.slug: t.id for t in _all_trainers}
    trainer_id_by_name = {t.full_name: t.id for t in _all_trainers}
    existing_by_id = {i.id: i for i in CourseInstance.query.all()}

    for line_no, raw in enumerate(rows, start=2):
        try:
            course_slug = _str(raw.get('course_slug'))
            if not course_slug:
                raise ValueError('порожній course_slug')
            course_id = course_id_by_slug.get(course_slug)
            if course_id is None:
                raise ValueError(f'course_slug={course_slug!r} не існує')

            start_date = _dt(raw.get('start_date'))
            if start_date is None:
                raise ValueError('порожня start_date')
            end_date = _dt(raw.get('end_date'))
            if end_date is not None and end_date <= start_date:
                raise ValueError(
                    f'end_date ({end_date.isoformat()}) має бути пізніше за '
                    f'start_date ({start_date.isoformat()})'
                )

            event_format_raw = _str(raw.get('event_format'))
            # Приймаємо як ('Онлайн','Офлайн','Гібрид'), так і ('online',
            # 'offline','hybrid') -- нормалізуємо у internal key.
            event_format = (
                FORMAT_KEY_BY_LABEL.get(event_format_raw, event_format_raw)
                if event_format_raw else None
            )
            if event_format and event_format not in VALID_FORMATS:
                allowed = sorted(VALID_FORMATS) + sorted(FORMAT_KEY_BY_LABEL.keys())
                raise ValueError(
                    f'event_format={event_format_raw!r} – допустимі: {allowed}'
                )

            online_link = _str(raw.get('online_link'))
            location = _str(raw.get('location')) or ''

            # Логічна узгодженість формат ↔ канал залишаємо warning-only:
            # порожній location у поточних seed-даних означає Київ за
            # замовчуванням; порожній online_link можна допилити в адмінці.
            # Hard-error лише на end_date < start_date (вище).

            status_raw = _str(raw.get('status')) or 'draft'
            status = STATUS_KEY_BY_LABEL.get(status_raw, status_raw)
            if status not in VALID_STATUSES:
                allowed = sorted(VALID_STATUSES) + sorted(STATUS_KEY_BY_LABEL.keys())
                raise ValueError(
                    f'status={status_raw!r} – допустимі: {allowed}'
                )

            # Колонка "Тренер" -- ПІБ (новий формат) або slug (старий).
            trainer_raw = _str(raw.get('trainer_slug'))
            trainer_id = None
            if trainer_raw:
                trainer_id = (
                    trainer_id_by_slug.get(trainer_raw)
                    or trainer_id_by_name.get(trainer_raw)
                )
                if trainer_id is None:
                    raise ValueError(
                        f'тренера {trainer_raw!r} не знайдено '
                        f'(ні за slug, ні за ПІБ)'
                    )

            parsed = {
                'id': _int(raw.get('id')),
                'course_id': course_id,
                'course_slug': course_slug,
                'start_date': start_date,
                'end_date': end_date,
                'event_format': event_format,
                'price': _decimal(raw.get('price')),
                'cpd_points': _int(raw.get('cpd_points')),
                'max_participants': _int(raw.get('max_participants')),
                'trainer_id': trainer_id,
                'location': location,
                'online_link': online_link,
                'status': status,
            }

            existing = None
            if parsed['id'] is not None:
                existing = existing_by_id.get(parsed['id'])
                if existing is None:
                    raise ValueError(
                        f'id={parsed["id"]} не існує '
                        f'(використайте порожній id для нового проведення)'
                    )

            plan.instances.append({'parsed': parsed, 'existing_id': existing.id if existing else None})

            sd = start_date.strftime('%Y-%m-%d')
            if existing is None:
                plan.changes.append(InstanceChange(
                    line_no=line_no, course_slug=course_slug,
                    start_date=sd, action='create',
                ))
            else:
                diff = _diff_instance(existing, parsed)
                if diff:
                    plan.changes.append(InstanceChange(
                        line_no=line_no, course_slug=course_slug,
                        start_date=sd, action='update',
                        fields_changed=diff,
                    ))
                else:
                    plan.changes.append(InstanceChange(
                        line_no=line_no, course_slug=course_slug,
                        start_date=sd, action='unchanged',
                    ))
        except Exception as exc:
            plan.errors.append(f'Рядок {line_no} (Instances): {exc}')
            plan.changes.append(InstanceChange(
                line_no=line_no,
                course_slug=_str(raw.get('course_slug')) or '',
                start_date=str(raw.get('start_date') or ''),
                action='error',
                error=str(exc),
            ))

    return plan


def _diff_instance(existing: CourseInstance, parsed: dict) -> list[str]:
    changed = []
    if existing.course_id != parsed['course_id']:
        changed.append('course_slug')
    # Дати -- через ensure_utc: SQLite віддає їх naive, а з файлу вони
    # приходять з київською tz. Порівняння naive з aware не падає, а просто
    # ЗАВЖДИ нерівне, тож на dev кожне наявне проведення показувалось як
    # "оновити". На PostgreSQL (prod) колонки timezone-aware і збігу немає.
    for f in ('start_date', 'end_date'):
        if ensure_utc(getattr(existing, f)) != ensure_utc(parsed[f]):
            changed.append(f)
    for f in ('event_format', 'cpd_points', 'max_participants', 'trainer_id',
              'online_link', 'status'):
        ev = getattr(existing, f)
        pv = parsed[f]
        if (ev or None) != (pv or None):
            changed.append(f)
    if (existing.location or '') != (parsed['location'] or ''):
        changed.append('location')
    ep = existing.price
    pp = parsed['price']
    if (ep is None) != (pp is None) or (ep is not None and pp is not None and ep != pp):
        changed.append('price')
    return changed


def apply_instances_plan(plan: InstancesImportPlan) -> dict:
    if not plan.is_valid:
        return {'ok': False, 'reason': 'plan has errors'}

    created = 0
    updated = 0

    try:
        for item in plan.instances:
            p = item['parsed']
            ex_id = item['existing_id']
            if ex_id is None:
                inst = CourseInstance(course_id=p['course_id'])
                db.session.add(inst)
                created += 1
            else:
                inst = db.session.get(CourseInstance, ex_id)
                updated += 1

            inst.course_id = p['course_id']
            inst.start_date = p['start_date']
            inst.end_date = p['end_date']
            inst.event_format = p['event_format']
            inst.price = p['price']
            inst.cpd_points = p['cpd_points']
            inst.max_participants = p['max_participants']
            inst.trainer_id = p['trainer_id']
            inst.location = p['location']
            inst.online_link = p['online_link']
            inst.status = p['status']

        db.session.commit()
        return {'ok': True, 'created': created, 'updated': updated}
    except Exception as exc:
        db.session.rollback()
        logger.exception('apply_instances_plan failed')
        return {'ok': False, 'reason': str(exc)}


# ======================================================================
# PARTICIPANTS (учасники заходів)
# ======================================================================

PARTICIPANT_COLS = [
    'reg_id', 'event', 'last_name', 'first_name', 'middle_name', 'email',
    'phone', 'participant_type', 'birth_date', 'education', 'workplace',
    'position', 'specializations', 'status', 'payment_status',
    'payment_amount', 'attended', 'cpd_points_awarded', 'experience_years',
    'license_number', 'admin_notes',
]

PARTICIPANT_LABELS = {
    'reg_id': 'ID реєстрації',
    'event': 'Захід',
    'last_name': 'Прізвище',
    'first_name': "Ім'я",
    'middle_name': 'По батькові',
    'email': 'Email',
    'phone': 'Телефон',
    'participant_type': 'Тип учасника',
    'birth_date': 'Дата народження',
    'education': 'Освіта',
    'workplace': 'Місце роботи / місто',
    'position': 'Посада',
    'specializations': 'Спеціалізації',
    'status': 'Статус',
    'payment_status': 'Оплата',
    'payment_amount': 'Сума (грн)',
    'attended': 'Присутній',
    'cpd_points_awarded': 'Бали БПР',
    'experience_years': 'Стаж (років)',
    'license_number': 'Ліцензія',
    'admin_notes': 'Нотатки',
    'promo_code': 'Промокод',
    'discount_amount': 'Знижка (грн)',
}

# Колонки ЛИШЕ на вивантаження: промокод і знижка -- довідкові, ними
# володіє promo_service, тож імпорт їх не читає (невідомі заголовки
# _read_sheet просто ігнорує, тож round-trip "вивантажив -> завантажив"
# лишається робочим).
PARTICIPANT_EXPORT_COLS = PARTICIPANT_COLS + ['promo_code', 'discount_amount']

PARTICIPANT_WIDTHS = {
    'reg_id': 12,
    'event': 46,
    'last_name': 20,
    'first_name': 18,
    'middle_name': 20,
    'email': 30,
    'phone': 18,
    'participant_type': 28,
    'birth_date': 16,
    'education': 40,
    'workplace': 34,
    'position': 26,
    'specializations': 40,
    'status': 16,
    'payment_status': 16,
    'payment_amount': 14,
    'attended': 12,
    'cpd_points_awarded': 12,
    'experience_years': 12,
    'license_number': 18,
    'admin_notes': 40,
    'promo_code': 18,
    'discount_amount': 14,
}

REG_STATUS_LABEL = dict(EventRegistration.STATUSES)
REG_STATUS_KEY_BY_LABEL = {v: k for k, v in REG_STATUS_LABEL.items()}
PAYMENT_STATUS_LABEL = dict(EventRegistration.PAYMENT_STATUSES)
PAYMENT_STATUS_KEY_BY_LABEL = {v: k for k, v in PAYMENT_STATUS_LABEL.items()}
PARTICIPANT_TYPE_LABEL = dict(MedicalProfile.PARTICIPANT_TYPES)
PARTICIPANT_TYPE_KEY_BY_LABEL = {v: k for k, v in PARTICIPANT_TYPE_LABEL.items()}
SPEC_LABEL_BY_CODE = dict(SPECIALIZATIONS)
SPEC_CODE_BY_LABEL = {v: k for k, v in SPECIALIZATIONS}

VALID_REG_STATUSES = set(REG_STATUS_LABEL.keys())
VALID_PAYMENT_STATUSES = set(PAYMENT_STATUS_LABEL.keys())
VALID_PARTICIPANT_TYPES = set(PARTICIPANT_TYPE_LABEL.keys())

REG_STATUS_FILLS = {
    'pending': _fill('FEF3C7'),     # yellow
    'confirmed': _fill('DBEAFE'),   # blue
    'completed': _fill('A7F3D0'),   # green
    'cancelled': _fill('FECACA'),   # red
}
PAYMENT_STATUS_FILLS = {
    'unpaid': _fill('F3F4F6'),      # gray
    'pending': _fill('FEF3C7'),     # yellow
    'paid': _fill('A7F3D0'),        # green
    'refunded': _fill('FECACA'),    # red
}

_EVENTS_SHEET_NAME = 'Заходи'
_SPEC_SHEET_NAME = 'Спеціалізації (довідник)'


def _date(v) -> date | None:
    """Прийняти date/datetime (openpyxl) або рядок (ISO / dd.mm.yyyy)."""
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    raise ValueError(f'неможливо розпарсити дату: {v!r}')


def _participant_event_label(instance) -> str:
    """Людиночитний ярлик заходу для колонки/довідника. Починається з
    '#<id>' -> id можна розпарсити навіть якщо назву трохи змінили.

    Єдина реалізація живе у participant_service.event_label (спільна з
    випадними списками форми)."""
    from app.services import participant_service
    return participant_service.event_label(instance, with_id=True)


def _resolve_event_to_id(label, label_to_id, instance_by_id):
    """Ярлик заходу -> instance_id. Спершу точний збіг, потім '#<id>'."""
    if not label:
        return None
    if label in label_to_id:
        return label_to_id[label]
    m = re.match(r'^\s*#(\d+)', str(label))
    if m:
        iid = int(m.group(1))
        if iid in instance_by_id:
            return iid
    return None


def _parse_spec_cell(raw):
    """Текст спеціалізацій (по рядку / через кому) -> (codes, unknown).

    Приймає і label ('Терапія'), і code ('therapy')."""
    codes, unknown, seen = [], [], set()
    for line in _from_lines(raw):
        for part in re.split(r'[;,]', line):
            p = part.strip()
            if not p:
                continue
            if p in SPEC_LABEL_BY_CODE:
                code = p
            elif p in SPEC_CODE_BY_LABEL:
                code = SPEC_CODE_BY_LABEL[p]
            else:
                unknown.append(p)
                continue
            if code not in seen:
                seen.add(code)
                codes.append(code)
    return codes, unknown


def _add_events_sheet(wb) -> int:
    """Reference-sheet із заходами. Ярлик у колонці A (для drop-down),
    id у B. Повертає номер останнього рядка з даними."""
    ws = wb.create_sheet(_EVENTS_SHEET_NAME)
    cols = ['event', 'id']
    _style_header(ws, cols, {'event': 'Захід (значення)', 'id': 'ID'})
    instances = (
        CourseInstance.query
        .options(joinedload(CourseInstance.course))
        .order_by(CourseInstance.start_date.desc().nullslast())
        .all()
    )
    for row_idx, inst in enumerate(instances, start=2):
        ws.cell(row=row_idx, column=1, value=_participant_event_label(inst)).alignment = WRAP
        ws.cell(row=row_idx, column=2, value=inst.id)
    _set_column_widths(ws, cols, {'event': 70, 'id': 8})
    _apply_zebra(ws, len(cols), first_data_row=2, last_data_row=1 + len(instances))
    _apply_table_style(ws, cols, 'tblEvents', last_data_row=1 + len(instances))
    return 1 + len(instances)


def _add_specializations_sheet(wb) -> None:
    """Reference-sheet (label + code) для ручного пошуку спеціалізацій."""
    ws = wb.create_sheet(_SPEC_SHEET_NAME)
    cols = ['label', 'code']
    _style_header(ws, cols, {'label': 'Спеціалізація', 'code': 'Код'})
    for row_idx, (code, label) in enumerate(SPECIALIZATIONS, start=2):
        ws.cell(row=row_idx, column=1, value=label).alignment = WRAP
        ws.cell(row=row_idx, column=2, value=code)
    _set_column_widths(ws, cols, {'label': 50, 'code': 30})
    _apply_zebra(ws, len(cols), first_data_row=2, last_data_row=1 + len(SPECIALIZATIONS))
    _apply_table_style(ws, cols, 'tblSpecs', last_data_row=1 + len(SPECIALIZATIONS))


def _add_ref_dropdown(ws, column_key, columns, last_data_row, sheet_name,
                      ref_last_row, title='', prompt=''):
    """Drop-down з reference-sheet (колонка A) на вказану колонку."""
    if ref_last_row < 2:
        return
    col_letter = get_column_letter(columns.index(column_key) + 1)
    formula = f"='{sheet_name}'!$A$2:$A${ref_last_row}"
    dv = DataValidation(
        type='list', formula1=formula, allow_blank=True,
        showDropDown=False, errorStyle='warning',
        error='Значення відсутнє у довіднику.', errorTitle='Невідоме значення',
        prompt=prompt, promptTitle=title,
    )
    final_row = max(last_data_row, 1) + _DROPDOWN_BUFFER_ROWS
    dv.add(f'{col_letter}2:{col_letter}{final_row}')
    ws.add_data_validation(dv)


@dataclass
class ParticipantChange:
    action: str  # 'create' | 'update' | 'unchanged' | 'error'
    name: str = ''
    event: str = ''
    fields_changed: list[str] = field(default_factory=list)
    error: str | None = None


@dataclass
class ParticipantsImportPlan:
    participants: list[dict] = field(default_factory=list)
    changes: list[ParticipantChange] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors

    @property
    def counts(self) -> dict[str, int]:
        c = {'create': 0, 'update': 0, 'unchanged': 0, 'error': 0}
        for ch in self.changes:
            c[ch.action] = c.get(ch.action, 0) + 1
        return c


def export_participants_xlsx(instance_id=None, blank=False) -> io.BytesIO:
    """Експорт учасників у xlsx (також слугує формою-шаблоном для додавання).

    instance_id: якщо задано -- лише учасники цього заходу.
    blank: True -- порожній шаблон (лише заголовки + dropdown-и, без даних).
    Reference-sheet «Заходи» + drop-down дозволяють додавати рядки для
    будь-якого заходу.

    Шаблон для заповнення (blank) містить лише редаговані колонки, а
    вивантаження даних -- ще й довідкові промокод/знижку: у шаблоні вони
    були б пасткою (заповнив -- нічого не сталось), бо знижками володіє
    promo_service, і імпорт їх не читає.
    """
    cols = PARTICIPANT_COLS if blank else PARTICIPANT_EXPORT_COLS
    wb = Workbook()
    ws = wb.active
    ws.title = 'Учасники'
    _style_header(ws, cols, PARTICIPANT_LABELS)

    from app.services import participant_service

    if blank:
        regs = []
    else:
        q = (
            EventRegistration.query
            .options(
                joinedload(EventRegistration.user).joinedload(User.medical_profile),
                joinedload(EventRegistration.instance).joinedload(CourseInstance.course),
            )
            .order_by(EventRegistration.created_at.desc())
        )
        if instance_id:
            q = q.filter(EventRegistration.instance_id == instance_id)
        regs = q.all()

    for row_idx, reg in enumerate(regs, start=2):
        user = reg.user
        profile = user.medical_profile if user else None
        raw_email = user.email if user else ''
        email = '' if participant_service.is_placeholder_email(raw_email) else (raw_email or '')
        spec_codes = (profile.specializations if profile else []) or []
        spec_text = '\n'.join(SPEC_LABEL_BY_CODE.get(c, c) for c in spec_codes)
        ptype = profile.participant_type if profile else None
        values = [
            reg.id,
            _participant_event_label(reg.instance) if reg.instance else '',
            (user.last_name if user else '') or '',
            (user.first_name if user else '') or '',
            (profile.middle_name if profile else '') or '',
            email,
            reg.phone or (profile.phone if profile else '') or '',
            PARTICIPANT_TYPE_LABEL.get(ptype, '') if ptype else '',
            profile.birth_date if profile else None,
            (profile.education if profile else '') or '',
            (profile.workplace if profile else '') or reg.workplace or '',
            (profile.position if profile else '') or '',
            spec_text,
            REG_STATUS_LABEL.get(reg.status, reg.status or ''),
            PAYMENT_STATUS_LABEL.get(reg.payment_status, reg.payment_status or ''),
            float(reg.payment_amount) if reg.payment_amount is not None else None,
            'Так' if reg.attended else 'Ні',
            reg.cpd_points_awarded,
            reg.experience_years,
            reg.license_number or '',
            reg.admin_notes or '',
            reg.promo_code.code if reg.promo_code else '',
            float(reg.discount_amount) if reg.discount_amount is not None else None,
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = write_cell(ws, row_idx, col_idx, v)
            cell.alignment = WRAP

    last_row = ws.max_row
    _apply_zebra(ws, len(cols), first_data_row=2, last_data_row=last_row)

    # Кольори за статусом / оплатою.
    st_col = PARTICIPANT_COLS.index('status') + 1
    pay_col = PARTICIPANT_COLS.index('payment_status') + 1
    for row_idx, reg in enumerate(regs, start=2):
        if reg.status in REG_STATUS_FILLS:
            ws.cell(row=row_idx, column=st_col).fill = REG_STATUS_FILLS[reg.status]
        if reg.payment_status in PAYMENT_STATUS_FILLS:
            ws.cell(row=row_idx, column=pay_col).fill = PAYMENT_STATUS_FILLS[reg.payment_status]

    _set_column_widths(ws, cols, PARTICIPANT_WIDTHS)
    _apply_number_formats(ws, cols, last_row)

    # Reference-sheets + drop-downs.
    events_last_row = _add_events_sheet(wb)
    _add_specializations_sheet(wb)
    _add_ref_dropdown(
        ws, 'event', PARTICIPANT_COLS, last_row,
        _EVENTS_SHEET_NAME, events_last_row,
        title='Захід', prompt='Оберіть захід зі списку (натисніть стрілочку)',
    )
    _add_inline_dropdown(
        ws, 'participant_type', PARTICIPANT_COLS,
        options=[label for _k, label in MedicalProfile.PARTICIPANT_TYPES],
        last_data_row=last_row, title='Тип учасника',
        hint='Лікар / Молодший спеціаліст / Інтерн / Студент',
    )
    _add_inline_dropdown(
        ws, 'status', PARTICIPANT_COLS,
        options=[label for _k, label in EventRegistration.STATUSES],
        last_data_row=last_row, title='Статус',
        hint='Очікує / Підтверджено / Скасовано / Завершено',
    )
    _add_inline_dropdown(
        ws, 'payment_status', PARTICIPANT_COLS,
        options=[label for _k, label in EventRegistration.PAYMENT_STATUSES],
        last_data_row=last_row, title='Оплата',
        hint='Не оплачено / Очікує оплати / Оплачено / Повернено',
    )
    _add_inline_dropdown(
        ws, 'attended', PARTICIPANT_COLS, options=['Так', 'Ні'],
        last_data_row=last_row, title='Присутній', hint='Так / Ні',
    )

    _apply_table_style(ws, cols, 'tblParticipants', last_row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


_PARTICIPANT_DIFF_LABELS = {
    'last_name': 'прізвище', 'first_name': "ім'я", 'middle_name': 'по батькові',
    'email': 'email', 'phone': 'телефон', 'participant_type': 'тип учасника',
    'birth_date': 'дата народж.', 'education': 'освіта', 'workplace': 'місце роботи',
    'position': 'посада', 'specializations': 'спеціалізації', 'status': 'статус',
    'payment_status': 'оплата', 'payment_amount': 'сума', 'attended': 'присутність',
    'cpd_points_awarded': 'бали БПР', 'experience_years': 'стаж',
    'license_number': 'ліцензія', 'admin_notes': 'нотатки',
}


def _diff_participant(reg, data):
    """Які поля зміняться при застосуванні data до наявної реєстрації.

    Дзеркалить семантику participant_service.upsert_participant: identity та
    профіль оновлюються лише непорожніми значеннями; реєстраційні поля --
    завжди (replace). Повертає список людиночитних назв змінених полів
    (порожній -> рядок 'без змін')."""
    user = reg.user
    profile = user.medical_profile if user else None
    changed = []

    def norm(v):
        return v if v not in ('', None) else None

    # Identity + профіль: оновлюються лише непорожнім вводом.
    soft = [
        ('last_name', user.last_name if user else None),
        ('first_name', user.first_name if user else None),
        ('middle_name', profile.middle_name if profile else None),
        ('participant_type', profile.participant_type if profile else None),
        ('birth_date', profile.birth_date if profile else None),
        ('education', profile.education if profile else None),
        ('workplace', profile.workplace if profile else None),
        ('position', profile.position if profile else None),
    ]
    for key, cur in soft:
        new = data.get(key)
        if new in ('', None):
            continue
        if norm(new) != norm(cur):
            changed.append(_PARTICIPANT_DIFF_LABELS[key])

    new_email = (data.get('email') or '').strip().lower()
    if new_email and user and new_email != (user.email or '').lower():
        changed.append(_PARTICIPANT_DIFF_LABELS['email'])

    new_specs = list(data.get('specializations') or [])
    if new_specs and new_specs != list((profile.specializations if profile else []) or []):
        changed.append(_PARTICIPANT_DIFF_LABELS['specializations'])

    # Реєстраційні поля -- replace-семантика (порівнюємо завжди).
    if norm(data.get('phone')) != norm(reg.phone):
        changed.append(_PARTICIPANT_DIFF_LABELS['phone'])
    if (data.get('status') or 'confirmed') != reg.status:
        changed.append(_PARTICIPANT_DIFF_LABELS['status'])
    if (data.get('payment_status') or 'unpaid') != reg.payment_status:
        changed.append(_PARTICIPANT_DIFF_LABELS['payment_status'])
    if bool(data.get('attended')) != bool(reg.attended):
        changed.append(_PARTICIPANT_DIFF_LABELS['attended'])
    if data.get('cpd_points_awarded') != reg.cpd_points_awarded:
        changed.append(_PARTICIPANT_DIFF_LABELS['cpd_points_awarded'])
    if data.get('experience_years') != reg.experience_years:
        changed.append(_PARTICIPANT_DIFF_LABELS['experience_years'])
    if norm(data.get('license_number')) != norm(reg.license_number):
        changed.append(_PARTICIPANT_DIFF_LABELS['license_number'])
    if norm(data.get('admin_notes')) != norm(reg.admin_notes):
        changed.append(_PARTICIPANT_DIFF_LABELS['admin_notes'])
    pa_new, pa_cur = data.get('payment_amount'), reg.payment_amount
    if (pa_new is None) != (pa_cur is None) or (
        pa_new is not None and pa_cur is not None and pa_new != pa_cur
    ):
        changed.append(_PARTICIPANT_DIFF_LABELS['payment_amount'])

    return changed


def parse_participants_xlsx(path: Path) -> ParticipantsImportPlan:
    plan = ParticipantsImportPlan()
    try:
        wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    except Exception as exc:
        plan.errors.append(f'Не вдалося відкрити xlsx: {exc}')
        return plan

    ws = _find_sheet(wb, 'participants')
    if ws is None:
        plan.errors.append('Відсутній sheet "Учасники"')
        return plan

    try:
        rows = _read_sheet(ws, PARTICIPANT_COLS, PARTICIPANT_LABELS)
    except ValueError as exc:
        plan.errors.append(str(exc))
        return plan

    instances = (
        CourseInstance.query.options(joinedload(CourseInstance.course)).all()
    )
    instance_by_id = {i.id: i for i in instances}
    label_to_id = {_participant_event_label(i): i.id for i in instances}

    # Preload проти N+1: реєстрації за reg_id, мапи email->User та
    # (user,instance)->активна реєстрація -- усе для визначення дії
    # (create/update/unchanged) без запитів у циклі.
    #
    # reg_id тягнемо теж: при звичайному циклі "вивантажив -> поправив ->
    # завантажив" кожен рядок має reg_id, і db.session.get у циклі давав
    # запит на кожного учасника.
    reg_ids_in_file = set()
    for r in rows:
        try:
            rid = _int(r.get('reg_id'))
        except ValueError:
            continue  # нечислове значення -- помилка рядка, не префетчу
        if rid:
            reg_ids_in_file.add(rid)
    regs_by_id = {}
    if reg_ids_in_file:
        regs_by_id = {
            r.id: r for r in EventRegistration.query
            .filter(EventRegistration.id.in_(reg_ids_in_file)).all()
        }

    emails_in_file = {
        (_str(r.get('email')) or '').strip().lower()
        for r in rows if _str(r.get('email'))
    }
    users_by_email = {}
    if emails_in_file:
        for u in User.query.filter(User.email.in_(emails_in_file)).all():
            users_by_email[u.email] = u
    active_reg = {}
    if users_by_email:
        uids = [u.id for u in users_by_email.values()]
        for r in (
            EventRegistration.query
            .filter(EventRegistration.user_id.in_(uids),
                    EventRegistration.status != 'cancelled')
            .all()
        ):
            active_reg[(r.user_id, r.instance_id)] = r

    for line_no, raw in enumerate(rows, start=2):
        try:
            reg_id = _int(raw.get('reg_id'))
            reg = regs_by_id.get(reg_id) if reg_id else None
            if reg_id and reg is None:
                raise ValueError(f'реєстрацію id={reg_id} не знайдено')

            last_name = _str(raw.get('last_name'))
            first_name = _str(raw.get('first_name'))
            phone = _str(raw.get('phone'))
            if not last_name:
                raise ValueError('порожнє Прізвище')
            if not first_name:
                raise ValueError("порожнє Ім'я")
            if not phone:
                raise ValueError('порожній Телефон')

            if reg is not None:
                instance_id = reg.instance_id
            else:
                event_label = _str(raw.get('event'))
                instance_id = _resolve_event_to_id(event_label, label_to_id, instance_by_id)
                if instance_id is None:
                    raise ValueError(f'захід {event_label!r} не знайдено')

            ptype_raw = _str(raw.get('participant_type'))
            ptype = PARTICIPANT_TYPE_KEY_BY_LABEL.get(ptype_raw, ptype_raw) if ptype_raw else None
            if ptype and ptype not in VALID_PARTICIPANT_TYPES:
                raise ValueError(f'тип учасника {ptype_raw!r} недопустимий')

            status_raw = _str(raw.get('status')) or 'Підтверджено'
            status = REG_STATUS_KEY_BY_LABEL.get(status_raw, status_raw)
            if status not in VALID_REG_STATUSES:
                raise ValueError(f'статус {status_raw!r} недопустимий')

            pay_raw = _str(raw.get('payment_status')) or 'Не оплачено'
            payment_status = PAYMENT_STATUS_KEY_BY_LABEL.get(pay_raw, pay_raw)
            if payment_status not in VALID_PAYMENT_STATUSES:
                raise ValueError(f'статус оплати {pay_raw!r} недопустимий')

            specs, unknown = _parse_spec_cell(raw.get('specializations'))
            if unknown:
                raise ValueError(f'невідомі спеціалізації: {", ".join(unknown)}')

            # Числові діапазони -- валідуємо тут, щоб дати чітку per-row
            # помилку замість падіння на DB CHECK-constraint у apply.
            payment_amount = _decimal(raw.get('payment_amount'))
            if payment_amount is not None and payment_amount < 0:
                raise ValueError('сума оплати не може бути від\'ємною')
            cpd = _int(raw.get('cpd_points_awarded'))
            if cpd is not None and cpd < 0:
                raise ValueError('бали БПР не можуть бути від\'ємними')
            experience = _int(raw.get('experience_years'))
            if experience is not None and not (0 <= experience <= 70):
                raise ValueError('стаж має бути в межах 0-70 років')
            birth = _date(raw.get('birth_date'))
            if birth is not None and (birth > date.today() or birth.year < 1900):
                raise ValueError('некоректна дата народження')

            data = {
                'instance_id': instance_id,
                'last_name': last_name,
                'first_name': first_name,
                'middle_name': _str(raw.get('middle_name')),
                'email': _str(raw.get('email')),
                'phone': phone,
                'participant_type': ptype,
                'birth_date': birth,
                'education': _str(raw.get('education')),
                'workplace': _str(raw.get('workplace')),
                'position': _str(raw.get('position')),
                'specializations': specs,
                'status': status,
                'payment_status': payment_status,
                'payment_amount': payment_amount,
                'attended': _bool(raw.get('attended')),
                'cpd_points_awarded': cpd,
                'experience_years': experience,
                'license_number': _str(raw.get('license_number')),
                'admin_notes': _str(raw.get('admin_notes')),
            }

            # Дія для preview: reg_id або наявна активна реєстрація за email
            # -> update (або unchanged, якщо нічого не змінюється); інакше create.
            existing_reg = reg
            if existing_reg is None:
                email = (data['email'] or '').strip().lower()
                if email:
                    u = users_by_email.get(email)
                    if u is not None:
                        existing_reg = active_reg.get((u.id, instance_id))

            if existing_reg is not None:
                diff = _diff_participant(existing_reg, data)
                action = 'update' if diff else 'unchanged'
            else:
                diff = []
                action = 'create'

            name = f'{last_name} {first_name}'.strip()
            event_lbl = (
                _participant_event_label(instance_by_id[instance_id])
                if instance_id in instance_by_id else ''
            )
            plan.participants.append({
                'data': data, 'reg_id': reg.id if reg else None, 'action': action,
            })
            plan.changes.append(ParticipantChange(
                action=action, name=name, event=event_lbl, fields_changed=diff,
            ))
        except Exception as exc:
            plan.errors.append(f'Рядок {line_no} (Учасники): {exc}')
            plan.changes.append(ParticipantChange(
                action='error',
                name=_str(raw.get('last_name')) or f'#{line_no}',
                event=_str(raw.get('event')) or '',
                error=str(exc),
            ))

    return plan


def apply_participants_plan(plan: ParticipantsImportPlan) -> dict:
    """Atomic upsert учасників. Очікує plan.is_valid==True."""
    if not plan.is_valid:
        return {'ok': False, 'reason': 'plan has errors'}

    from app.services import participant_service

    created = 0
    updated = 0
    skipped = 0
    try:
        for item in plan.participants:
            if item.get('action') == 'unchanged':
                skipped += 1
                continue
            reg_id = item['reg_id']
            reg = db.session.get(EventRegistration, reg_id) if reg_id else None
            _reg, was_created = participant_service.upsert_participant(
                item['data'], reg=reg, on_duplicate='update',
            )
            if was_created:
                created += 1
            else:
                updated += 1
        db.session.commit()
        return {'ok': True, 'created': created, 'updated': updated, 'skipped': skipped}
    except participant_service.ParticipantError as exc:
        db.session.rollback()
        return {'ok': False, 'reason': str(exc)}
    except Exception as exc:
        db.session.rollback()
        logger.exception('apply_participants_plan failed')
        return {'ok': False, 'reason': str(exc)}


# ==================== MM MEDIC MATERIALS TEMPLATE ====================

_MATERIALS_COLS = ['image', 'sku', 'name', 'available', 'quantity']
_MATERIALS_LABELS = {
    'image': 'Зображення',
    'sku': 'Артикул',
    'name': 'Назва',
    'available': 'Наявно',
    'quantity': 'Кількість',
}
_MATERIALS_WIDTHS = {'image': 9, 'sku': 20, 'name': 46, 'available': 12, 'quantity': 14}
_MATERIALS_THUMB_PX = 40


# Сумарний бюджет на ВСІ мініатюри одного експорту. Без нього каталог на
# сотню позицій міг тягнутись сотні секунд (до 8 с на позицію) і впертись у
# таймаут шлюзу -- заради декоративних картинок.
_MATERIALS_THUMB_BUDGET_SECONDS = 20.0
_MATERIALS_THUMB_MAX_BYTES = 2_000_000


def _download_thumb(url, max_px=_MATERIALS_THUMB_PX):
    """Завантажити зображення товару й повернути BytesIO з PNG-мініатюрою
    (max_px), або None (порожнє/не-http/збій/не зображення). Best-effort."""
    url = (url or '').strip()
    if not url.startswith(('http://', 'https://')):
        return None
    try:
        import requests
        from PIL import Image as PILImage

        # stream + порізний ліміт: раніше resp.content матеріалізував тіло
        # ЦІЛКОМ, і перевірка розміру після цього вже нічого не рятувала.
        with requests.get(url, timeout=(3.0, 5.0), stream=True) as resp:
            if not resp.ok:
                return None
            chunks, total = [], 0
            for chunk in resp.iter_content(64 * 1024):
                total += len(chunk)
                if total > _MATERIALS_THUMB_MAX_BYTES:
                    return None
                chunks.append(chunk)
        img = PILImage.open(io.BytesIO(b''.join(chunks)))
        img.thumbnail((max_px, max_px))
        out = io.BytesIO()
        img.convert('RGB').save(out, format='PNG')
        out.seek(0)
        return out
    except Exception:
        logger.info('materials thumb fetch failed: %s', url)
        return None


def export_materials_template_xlsx(catalog: list[dict]) -> io.BytesIO:
    """Шаблон для резервування витратних матеріалів MM Medic.

    `catalog` -- список dict з ключами sku, name, available, image (як віддає
    MM Medic /catalog). Перша колонка `Зображення` містить вбудовану мініатюру
    товару (не URL); зображення тягнуться з MM Medic best-effort. Колонка
    `Кількість` порожня: адмін вписує потрібні кількості; незаповнені рядки на
    імпорті ігноруються.
    """
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = 'Матеріали'
    _style_header(ws, _MATERIALS_COLS, _MATERIALS_LABELS)

    img_col = get_column_letter(_MATERIALS_COLS.index('image') + 1)
    sku_i = _MATERIALS_COLS.index('sku') + 1
    name_i = _MATERIALS_COLS.index('name') + 1
    avail_i = _MATERIALS_COLS.index('available') + 1

    deadline = time.monotonic() + _MATERIALS_THUMB_BUDGET_SECONDS
    skipped_thumbs = 0

    for row_idx, item in enumerate(catalog or [], start=2):
        ws.cell(row=row_idx, column=sku_i, value=item.get('sku') or '')
        ws.cell(row=row_idx, column=name_i, value=item.get('name') or '')
        ws.cell(row=row_idx, column=avail_i, value=item.get('available'))
        # quantity column left empty for the admin

        # Вичерпали бюджет -- решта рядків без картинок. Файл лишається
        # придатним до роботи, а адмін не чекає таймауту.
        if time.monotonic() >= deadline:
            skipped_thumbs += 1
            continue
        thumb = _download_thumb(item.get('image'))
        if thumb is not None:
            xi = XLImage(thumb)
            xi.width = _MATERIALS_THUMB_PX
            xi.height = _MATERIALS_THUMB_PX
            ws.add_image(xi, f'{img_col}{row_idx}')
            ws.row_dimensions[row_idx].height = 32

    if skipped_thumbs:
        logger.info('materials export: %s thumbnails skipped (time budget)',
                    skipped_thumbs)

    last_row = ws.max_row
    _set_column_widths(ws, _MATERIALS_COLS, _MATERIALS_WIDTHS)
    _apply_zebra(ws, len(_MATERIALS_COLS), first_data_row=2, last_data_row=last_row)
    if last_row >= 2:
        _apply_table_style(ws, _MATERIALS_COLS, 'tblMaterials', last_row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


_RESV_COLS = ['event', 'date', 'status', 'positions', 'reserved', 'actual']
_RESV_LABELS = {
    'event': 'Захід', 'date': 'Дата', 'status': 'Статус',
    'positions': 'Позицій', 'reserved': 'Зарезервовано', 'actual': 'Фактично',
}
_RESV_WIDTHS = {'event': 46, 'date': 14, 'status': 16, 'positions': 10,
                'reserved': 14, 'actual': 12}


def export_material_reservations_xlsx(reservations) -> io.BytesIO:
    """Огляд резервувань матеріалів -> xlsx (для експорту зі сторінки огляду)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Резервування'
    _style_header(ws, _RESV_COLS, _RESV_LABELS)

    for row_idx, r in enumerate(reservations, start=2):
        course = r.instance.course.title if (r.instance and r.instance.course) else '—'
        date = (r.instance.start_date.strftime('%d.%m.%Y')
                if (r.instance and r.instance.start_date) else '')
        reserved = sum((it.quantity_reserved or 0) for it in r.items)
        actual = sum((it.quantity_actual or 0) for it in r.items)
        values = [course, date, r.status_label, len(r.items), reserved, actual]
        for col_idx, v in enumerate(values, start=1):
            write_cell(ws, row_idx, col_idx, v)

    last_row = ws.max_row
    _set_column_widths(ws, _RESV_COLS, _RESV_WIDTHS)
    _apply_zebra(ws, len(_RESV_COLS), first_data_row=2, last_data_row=last_row)
    if last_row >= 2:
        _apply_table_style(ws, _RESV_COLS, 'tblResv', last_row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def parse_materials_xlsx(path: Path) -> dict[str, int]:
    """Прочитати заповнений шаблон -> {sku: quantity} лише для quantity > 0.

    Рядки з порожньою/нульовою/невалідною кількістю ігноруються (за вимогою).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Раніше тут стояло _find_sheet(wb, 'Матеріали') -- ключа 'Матеріали'
        # у SHEET_ALIASES немає, тож функція ЗАВЖДИ повертала None і код
        # мовчки читав активний лист. Варто було адміну лишити активним
        # інший лист (напр. власні нотатки), як імпорт повертав порожньо, а
        # повідомлення казало "Завантажено 0 позицій".
        ws = _find_sheet(wb, 'materials')
        if ws is None:
            if len(wb.sheetnames) == 1:
                ws = wb[wb.sheetnames[0]]  # файл зібрали вручну -- беремо єдиний
            else:
                raise ValueError(
                    'у файлі немає листа "Матеріали"; наявні листи: '
                    + ', '.join(wb.sheetnames)
                )
        rows = _read_sheet(ws, ['sku', 'quantity'], _MATERIALS_LABELS)
    finally:
        # read_only mode keeps the file handle open; must close or Windows
        # blocks the subsequent cleanup_upload() unlink.
        wb.close()

    result: dict[str, int] = {}
    for row in rows:
        sku = (str(row.get('sku')).strip() if row.get('sku') is not None else '')
        if not sku:
            continue
        try:
            qty = _int(row.get('quantity'))
        except ValueError:
            continue  # невалідне число -> ігнор
        if not qty or qty <= 0:
            continue
        result[sku] = qty
    return result
