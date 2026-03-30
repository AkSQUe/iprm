"""Event Excel import/export service."""

import logging
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload

from app.extensions import db
from app.models.event import Event
from app.models.trainer import Trainer
from app.services import event_service
from app.services.event_excel_config import (
    COLUMNS_CONFIG,
    bool_to_ua,
    get_export_columns,
    get_import_mapping,
    list_to_semicolon,
)

logger = logging.getLogger(__name__)

VALID_EVENT_TYPES = {code for code, _ in Event.EVENT_TYPES}
VALID_FORMATS = {code for code, _ in Event.FORMATS}
VALID_STATUSES = {code for code, _ in Event.STATUSES}

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='7055A4', end_color='7055A4', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

UA_MONTHS = {
    1: '\u0441\u0456\u0447', 2: '\u043b\u044e\u0442', 3: '\u0431\u0435\u0440',
    4: '\u043a\u0432\u0456', 5: '\u0442\u0440\u0430', 6: '\u0447\u0435\u0440',
    7: '\u043b\u0438\u043f', 8: '\u0441\u0435\u0440', 9: '\u0432\u0435\u0440',
    10: '\u0436\u043e\u0432', 11: '\u043b\u0438\u0441', 12: '\u0433\u0440\u0443',
}


def _build_export_filename():
    now = datetime.now()
    month = UA_MONTHS[now.month]
    return f"\u0417\u0430\u0445\u043e\u0434\u0438-\u0406\u041f\u0420\u041c-{now.day}-{month}-{now.year}.xlsx"


def export_events_to_xlsx():
    """Export all events to an xlsx BytesIO stream.

    Returns:
        tuple: (BytesIO stream, filename string)
    """
    events = (
        Event.query
        .options(joinedload(Event.trainer))
        .order_by(Event.created_at.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '\u0417\u0430\u0445\u043e\u0434\u0438'

    columns = get_export_columns()

    for col_idx, (_, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, event in enumerate(events, start=2):
        for col_idx, (field_key, _, _) in enumerate(columns, start=1):
            value = _extract_export_value(event, field_key)
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = 'A2'
    if events:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f'A1:{last_col}{len(events) + 1}'

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = _build_export_filename()
    return stream, filename


def _extract_export_value(event, field_key):
    """Extract a single field value from an Event for export."""
    if field_key == 'id':
        return event.id
    if field_key == 'trainer_name':
        return event.trainer.full_name if event.trainer else ''
    if field_key == 'target_audience':
        return list_to_semicolon(event.target_audience)
    if field_key == 'tags':
        return list_to_semicolon(event.tags)
    if field_key == 'is_featured':
        return bool_to_ua(event.is_featured)
    if field_key == 'is_active':
        return bool_to_ua(event.is_active)
    if field_key in ('start_date', 'end_date'):
        dt = getattr(event, field_key)
        return dt.strftime('%d.%m.%Y %H:%M') if dt else ''
    if field_key == 'price':
        return float(event.price) if event.price is not None else 0

    db_field = _get_db_field(field_key)
    value = getattr(event, db_field, None)
    return value if value is not None else ''


def _get_db_field(field_key):
    """Look up the db_field for a given field_key from config."""
    for fk, _, _, db_field, _, _ in COLUMNS_CONFIG:
        if fk == field_key:
            return db_field
    return field_key


def import_events_from_xlsx(file_stream, user_id):
    """Import events from an xlsx file stream.

    Args:
        file_stream: BytesIO or file-like object with xlsx data
        user_id: ID of the admin performing the import

    Returns:
        dict with keys: total_rows, created, updated, skipped, errors
    """
    stats = {
        'total_rows': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
    }

    try:
        wb = openpyxl.load_workbook(
            file_stream,
            data_only=True,
            keep_vba=False,
            keep_links=False,
        )
    except Exception as exc:
        stats['errors'].append(f'\u041d\u0435 \u0432\u0434\u0430\u043b\u043e\u0441\u044f \u0432\u0456\u0434\u043a\u0440\u0438\u0442\u0438 \u0444\u0430\u0439\u043b: {exc}')
        return stats

    ws = wb.active
    if ws.max_row is None or ws.max_row < 2:
        stats['errors'].append('\u0424\u0430\u0439\u043b \u043f\u043e\u0440\u043e\u0436\u043d\u0456\u0439 \u0430\u0431\u043e \u043d\u0435 \u043c\u0456\u0441\u0442\u0438\u0442\u044c \u0434\u0430\u043d\u0438\u0445')
        return stats

    if ws.max_row > 10000:
        stats['errors'].append('\u0424\u0430\u0439\u043b \u043f\u0435\u0440\u0435\u0432\u0438\u0449\u0443\u0454 \u043c\u0430\u043a\u0441\u0438\u043c\u0443\u043c 10 000 \u0440\u044f\u0434\u043a\u0456\u0432')
        return stats

    header_row, col_map = _find_headers(ws)
    if header_row is None:
        stats['errors'].append("\u0420\u044f\u0434\u043e\u043a \u0437\u0430\u0433\u043e\u043b\u043e\u0432\u043a\u0456\u0432 \u043d\u0435 \u0437\u043d\u0430\u0439\u0434\u0435\u043d\u043e (\u043f\u043e\u0442\u0440\u0456\u0431\u043d\u0430 \u043a\u043e\u043b\u043e\u043d\u043a\u0430 '\u041d\u0430\u0437\u0432\u0430')")
        return stats

    trainer_cache = _build_trainer_cache()

    for row_idx in range(header_row + 1, ws.max_row + 1):
        try:
            row_data = _read_row(ws, row_idx, col_map)
            if not row_data:
                continue

            stats['total_rows'] += 1
            _process_row(row_data, row_idx, user_id, trainer_cache, stats)
        except Exception as exc:
            stats['total_rows'] += 1
            stats['errors'].append(f'\u0420\u044f\u0434\u043e\u043a {row_idx}: {exc}')
            stats['skipped'] += 1

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        stats['errors'].append(f'\u041f\u043e\u043c\u0438\u043b\u043a\u0430 \u0431\u0430\u0437\u0438 \u0434\u0430\u043d\u0438\u0445: {exc}')
        stats['created'] = 0
        stats['updated'] = 0

    return stats


def _find_headers(ws):
    """Scan first 10 rows for the header row containing '\u041d\u0430\u0437\u0432\u0430'.

    Returns:
        tuple: (header_row_number, {col_index: (db_field, converter, default)})
    """
    import_mapping = get_import_mapping()

    for row_idx in range(1, min(11, (ws.max_row or 1) + 1)):
        for col_idx in range(1, (ws.max_column or 1) + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and str(cell_value).strip() == '\u041d\u0430\u0437\u0432\u0430':
                col_map = {}
                for ci in range(1, (ws.max_column or 1) + 1):
                    header = ws.cell(row=row_idx, column=ci).value
                    if header and str(header).strip() in import_mapping:
                        col_map[ci] = import_mapping[str(header).strip()]
                return row_idx, col_map
    return None, None


def _build_trainer_cache():
    """Build {full_name_lower: Trainer} lookup dict."""
    trainers = Trainer.query.filter_by(is_active=True).all()
    return {t.full_name.lower(): t for t in trainers}


def _read_row(ws, row_idx, col_map):
    """Read a single row into a dict of {db_field: converted_value}.

    Returns None if the row is entirely empty.
    """
    data = {}
    has_value = False

    for col_idx, (db_field, converter, default) in col_map.items():
        raw = ws.cell(row=row_idx, column=col_idx).value
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            data[db_field] = None
            continue

        has_value = True
        if isinstance(raw, str):
            raw = raw.strip()
        try:
            data[db_field] = converter(raw) if converter else raw
        except (ValueError, TypeError) as exc:
            header = ws.cell(row=1, column=col_idx).value or db_field
            raise ValueError(f"\u043a\u043e\u043b\u043e\u043d\u043a\u0430 '{header}': {exc}") from exc

    return data if has_value else None


def _process_row(row_data, row_idx, user_id, trainer_cache, stats):
    """Process a single row: update existing or create new event."""
    event_id = row_data.get('id')

    if event_id is not None:
        try:
            event_id = int(event_id)
        except (ValueError, TypeError):
            raise ValueError(f"\u041d\u0435\u043a\u043e\u0440\u0435\u043a\u0442\u043d\u0435 \u0437\u043d\u0430\u0447\u0435\u043d\u043d\u044f ID: '{event_id}'")
        event = db.session.get(Event, event_id)
        if not event:
            raise ValueError(f'\u0417\u0430\u0445\u0456\u0434 \u0437 ID {event_id} \u043d\u0435 \u0437\u043d\u0430\u0439\u0434\u0435\u043d\u043e')
        _update_event(event, row_data, trainer_cache, stats)
        stats['updated'] += 1
    else:
        title = row_data.get('title')
        if not title:
            raise ValueError('\u041d\u0430\u0437\u0432\u0430 \u043e\u0431\u043e\u0432\u044f\u0437\u043a\u043e\u0432\u0430 \u0434\u043b\u044f \u043d\u043e\u0432\u0438\u0445 \u0437\u0430\u0445\u043e\u0434\u0456\u0432')
        _create_event(row_data, user_id, trainer_cache, stats)
        stats['created'] += 1


def _validate_constrained_fields(row_data):
    """Validate event_type, event_format, status against allowed values."""
    event_type = row_data.get('event_type')
    if event_type and event_type not in VALID_EVENT_TYPES:
        raise ValueError(
            f"\u041d\u0435\u0432\u0456\u0440\u043d\u0438\u0439 \u0442\u0438\u043f \u0437\u0430\u0445\u043e\u0434\u0443 '{event_type}'. "
            f"\u0414\u043e\u0437\u0432\u043e\u043b\u0435\u043d\u0456: {', '.join(VALID_EVENT_TYPES)}"
        )

    event_format = row_data.get('event_format')
    if event_format and event_format not in VALID_FORMATS:
        raise ValueError(
            f"\u041d\u0435\u0432\u0456\u0440\u043d\u0438\u0439 \u0444\u043e\u0440\u043c\u0430\u0442 '{event_format}'. "
            f"\u0414\u043e\u0437\u0432\u043e\u043b\u0435\u043d\u0456: {', '.join(VALID_FORMATS)}"
        )

    status = row_data.get('status')
    if status and status not in VALID_STATUSES:
        raise ValueError(
            f"\u041d\u0435\u0432\u0456\u0440\u043d\u0438\u0439 \u0441\u0442\u0430\u0442\u0443\u0441 '{status}'. "
            f"\u0414\u043e\u0437\u0432\u043e\u043b\u0435\u043d\u0456: {', '.join(VALID_STATUSES)}"
        )


def _resolve_trainer(row_data, trainer_cache, stats):
    """Resolve trainer_name to trainer_id. Returns trainer_id or None."""
    trainer_name = row_data.get('trainer_name')
    if not trainer_name:
        return None

    trainer = trainer_cache.get(trainer_name.lower())
    if trainer:
        return trainer.id

    stats['errors'].append(
        f"\u0422\u0440\u0435\u043d\u0435\u0440 '{trainer_name}' \u043d\u0435 \u0437\u043d\u0430\u0439\u0434\u0435\u043d\u043e, \u043f\u0440\u0438\u0437\u043d\u0430\u0447\u0435\u043d\u043d\u044f \u043f\u0440\u043e\u043f\u0443\u0449\u0435\u043d\u043e"
    )
    return None


DIRECT_FIELDS = {
    'title', 'slug', 'subtitle', 'short_description', 'description',
    'event_type', 'event_format', 'status', 'start_date', 'end_date',
    'max_participants', 'price', 'location', 'online_link',
    'hero_image', 'card_image', 'cpd_points', 'target_audience', 'tags',
    'is_featured', 'is_active',
}


def _update_event(event, row_data, trainer_cache, stats):
    """Update an existing event with non-None values from row_data."""
    _validate_constrained_fields(row_data)

    trainer_id = _resolve_trainer(row_data, trainer_cache, stats)
    if row_data.get('trainer_name') is not None:
        event.trainer_id = trainer_id

    for field in DIRECT_FIELDS:
        value = row_data.get(field)
        if value is not None:
            setattr(event, field, value)


def _create_event(row_data, user_id, trainer_cache, stats):
    """Create a new event from row_data."""
    _validate_constrained_fields(row_data)

    title = row_data['title']
    slug = row_data.get('slug')
    if not slug:
        slug, error = event_service.generate_slug(title)
        if error:
            for counter in range(2, 102):
                candidate = f'{slug}-{counter}'
                if not Event.query.filter_by(slug=candidate).first():
                    slug = candidate
                    break
            else:
                raise ValueError(f"\u041d\u0435 \u0432\u0434\u0430\u043b\u043e\u0441\u044f \u0437\u0433\u0435\u043d\u0435\u0440\u0443\u0432\u0430\u0442\u0438 \u0443\u043d\u0456\u043a\u0430\u043b\u044c\u043d\u0438\u0439 slug \u0434\u043b\u044f '{title}'")

    event = Event(
        title=title,
        slug=slug,
        created_by=user_id,
    )

    trainer_id = _resolve_trainer(row_data, trainer_cache, stats)
    event.trainer_id = trainer_id

    for field in DIRECT_FIELDS:
        if field in ('title', 'slug'):
            continue
        value = row_data.get(field)
        if value is not None:
            setattr(event, field, value)

    if event.status is None:
        event.status = 'draft'
    if event.price is None:
        event.price = Decimal('0')
    if event.is_active is None:
        event.is_active = True

    db.session.add(event)
