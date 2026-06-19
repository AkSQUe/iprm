"""Рахунок на оплату участі: Excel (openpyxl) + PDF (WeasyPrint).

Excel будується openpyxl за версткою IPRM-рахунка; PDF рендериться WeasyPrint
з HTML-шаблона invoices/invoice.html (той самий стек, що й сертифікати) --
без зовнішніх залежностей на кшталт LibreOffice. Обидва документи будуються
з одних даних. Реквізити постачальника беруться з SiteSettings (банк, IBAN,
ЄДРПОУ, податковий статус).
"""
import io
import logging
from decimal import Decimal

from flask import current_app, render_template
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models.mixins import utcnow
from app.models.site_settings import SiteSettings
from app.services.number_to_words_ua import number_to_words_ua
from app.utils import ensure_utc

logger = logging.getLogger(__name__)

_MONTHS = ['', 'січня', 'лютого', 'березня', 'квітня', 'травня', 'червня',
           'липня', 'серпня', 'вересня', 'жовтня', 'листопада', 'грудня']

# Apple-inspired палітра: фіолетовий акцент + помаранчевий вторинний.
_PURPLE = '7055A4'
_PURPLE_LIGHT = 'EDE9F6'
_ORANGE_LIGHT = 'FEF0E6'
_ORANGE_DARK = 'D97A1F'
_INK = '1D1D1F'
_GRAY = '6E6E73'
_GRAY_LINE = 'E5E5EA'
_SECTION = 'F5F5F7'
_WHITE = 'FFFFFF'

_SIDE = Side(style='thin', color=_GRAY_LINE)
_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

_FILL_PURPLE = PatternFill('solid', fgColor=_PURPLE)
_FILL_PURPLE_LIGHT = PatternFill('solid', fgColor=_PURPLE_LIGHT)
_FILL_ORANGE_LIGHT = PatternFill('solid', fgColor=_ORANGE_LIGHT)
_FILL_SECTION = PatternFill('solid', fgColor=_SECTION)

_F_COMPANY = Font(name='Calibri', size=14, bold=True, color=_PURPLE)
_F_TITLE = Font(name='Calibri', size=20, bold=True, color=_PURPLE)
_F_LABEL = Font(name='Calibri', size=8.5, bold=True, color=_PURPLE)
_F_HEADER = Font(name='Calibri', size=10, bold=True, color=_WHITE)
_F_BOLD = Font(name='Calibri', size=10.5, bold=True, color=_INK)
_F_NORMAL = Font(name='Calibri', size=10, color=_INK)
_F_MUTED = Font(name='Calibri', size=9.5, color=_GRAY)
_F_MUTED_IT = Font(name='Calibri', size=9.5, italic=True, color=_GRAY)
_F_TOTAL = Font(name='Calibri', size=13, bold=True, color=_ORANGE_DARK)

_WRAP = Alignment(wrap_text=True, vertical='center')
_WRAP_TOP = Alignment(wrap_text=True, vertical='top')
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_RIGHT = Alignment(horizontal='right', vertical='center')

_COLS = ['A', 'B', 'C', 'D', 'E', 'F']
_WIDTHS = {'A': 5, 'B': 46, 'C': 9, 'D': 7, 'E': 13, 'F': 15}


class InvoiceError(Exception):
    """Помилка генерації/конвертації рахунка (показується користувачу)."""


def _date_phrase(dt):
    dt = ensure_utc(dt) or utcnow()
    return f'{dt.day} {_MONTHS[dt.month]} {dt.year} р.'


def _invoice_context(reg):
    settings = SiteSettings.get()
    course = reg.instance.course if reg.instance else None
    title = course.title if course else (reg.target_title or f'Реєстрація #{reg.id}')
    item_name = 'Участь у заході: ' + title
    if reg.instance and reg.instance.start_date:
        item_name += f' ({ensure_utc(reg.instance.start_date).strftime("%d.%m.%Y")})'
    payer = ''
    if reg.user:
        payer = (reg.user.full_name or '').strip() or reg.user.email or ''
    return {
        'settings': settings,
        'number': reg.id,
        'date': reg.created_at,
        'item_name': item_name,
        'amount': Decimal(reg.payment_amount or 0),
        'payer': payer or 'Фізична особа',
    }


def build_invoice_xlsx(reg) -> io.BytesIO:
    """Згенерувати Excel-рахунок (BytesIO) за версткою IPRM."""
    ctx = _invoice_context(reg)
    s = ctx['settings']
    amount = ctx['amount']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Рахунок'
    ws.sheet_view.showGridLines = False
    for col, w in _WIDTHS.items():
        ws.column_dimensions[col].width = w

    def band(row, text, *, font, fill=None, align=None, height=None):
        """Рядок-смуга на всю ширину A:F."""
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = text
        cell.font = font
        cell.alignment = align or _WRAP
        if fill is not None:
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = fill
        if height:
            ws.row_dimensions[row].height = height
        return cell

    r = 1

    # ----- Шапка: назва компанії + бейдж -----
    ws.merge_cells(f'A{r}:D{r}')
    ws[f'A{r}'].value = s.company_full_name or 'ІПРМ'
    ws[f'A{r}'].font = _F_COMPANY
    ws[f'A{r}'].alignment = _WRAP
    ws.merge_cells(f'E{r}:F{r}')
    ws[f'E{r}'].value = 'Документ для оплати'
    ws[f'E{r}'].font = _F_MUTED
    ws[f'E{r}'].alignment = _RIGHT
    ws.row_dimensions[r].height = 22
    r += 1
    band(r, f'код за ЄДРПОУ {s.edrpou}', font=_F_MUTED)
    r += 2

    # ----- Заголовок -----
    band(r, 'РАХУНОК НА ОПЛАТУ', font=_F_TITLE, height=26)
    r += 1
    band(r, f'№ {ctx["number"]} від {_date_phrase(ctx["date"])}', font=_F_MUTED)
    r += 2

    # ----- Попередження -----
    band(r, (
        'Увага! Оплата цього рахунку означає погодження з умовами надання послуг. '
        'Повідомлення про оплату є обов\'язковим. Послуга надається за фактом '
        'надходження коштів на рахунок Постачальника.'
    ), font=_F_MUTED, align=_WRAP_TOP, height=34)
    r += 2

    # ----- Реквізити для оплати (помаранчева картка) -----
    band(r, 'РЕКВІЗИТИ ДЛЯ ОПЛАТИ', font=_F_LABEL, fill=_FILL_ORANGE_LIGHT)
    r += 1
    for label, value in [
        ('Отримувач', s.company_full_name or 'ІПРМ'),
        ('Код ЄДРПОУ', s.edrpou or ''),
        ('Банк отримувача', s.bank_name or ''),
        ('Рахунок (IBAN)', s.bank_iban or ''),
    ]:
        ws.merge_cells(f'A{r}:B{r}')
        ws.merge_cells(f'C{r}:F{r}')
        ws[f'A{r}'].value = label
        ws[f'A{r}'].font = _F_MUTED
        ws[f'A{r}'].alignment = _WRAP
        ws[f'C{r}'].value = value
        ws[f'C{r}'].font = _F_BOLD
        ws[f'C{r}'].alignment = _WRAP
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = _FILL_ORANGE_LIGHT
        r += 1
    r += 1

    # ----- Постачальник (сіра картка) -----
    band(r, 'ПОСТАЧАЛЬНИК', font=_F_LABEL, fill=_FILL_SECTION)
    r += 1
    for text, font in [
        (s.company_full_name or 'ІПРМ', _F_BOLD),
        (f'п/р {s.bank_iban} у банку {s.bank_name}', _F_MUTED),
        (f'код за ЄДРПОУ {s.edrpou}', _F_MUTED),
        (s.tax_status or '', _F_MUTED_IT),
    ]:
        if not text:
            continue
        band(r, text, font=font, fill=_FILL_SECTION)
        r += 1
    r += 1

    # ----- Платник (фіолетова картка) -----
    band(r, 'ПЛАТНИК', font=_F_LABEL, fill=_FILL_PURPLE_LIGHT)
    r += 1
    band(r, ctx['payer'], font=_F_BOLD, fill=_FILL_PURPLE_LIGHT)
    r += 2

    # ----- Таблиця послуг -----
    headers = ['№', 'Найменування робіт, послуг', 'Кіл-ть', 'Од.', 'Ціна', 'Сума']
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx + 1, value=h)
        cell.font = _F_HEADER
        cell.fill = _FILL_PURPLE
        cell.alignment = _CENTER if idx != 1 else _WRAP
        cell.border = _BORDER
    ws.row_dimensions[r].height = 20
    r += 1

    row_vals = [1, ctx['item_name'], 1, 'посл', float(amount), float(amount)]
    for idx, v in enumerate(row_vals):
        cell = ws.cell(row=r, column=idx + 1, value=v)
        cell.font = _F_NORMAL
        cell.border = _BORDER
        if idx in (0, 2, 3):
            cell.alignment = _CENTER
        elif idx in (4, 5):
            cell.alignment = _RIGHT
            cell.number_format = '#,##0.00'
        else:
            cell.alignment = _WRAP
    r += 1

    # Разом (помаранчевий акцент)
    ws.merge_cells(f'A{r}:E{r}')
    tc = ws[f'A{r}']
    tc.value = 'Разом до сплати:'
    tc.font = _F_BOLD
    tc.alignment = _RIGHT
    sum_cell = ws.cell(row=r, column=6, value=float(amount))
    sum_cell.font = _F_TOTAL
    sum_cell.alignment = _RIGHT
    sum_cell.number_format = '#,##0.00 ₴'
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = _FILL_ORANGE_LIGHT
        ws.cell(row=r, column=c).border = _BORDER
    ws.row_dimensions[r].height = 22
    r += 2

    # ----- Сума прописом (фіолетова картка) -----
    band(r, 'СУМА ПРОПИСОМ', font=_F_LABEL, fill=_FILL_PURPLE_LIGHT)
    r += 1
    words = number_to_words_ua(amount, 'UAH')
    words = words[0].upper() + words[1:] if words else words
    band(r, words, font=_F_MUTED_IT, fill=_FILL_PURPLE_LIGHT, align=_WRAP_TOP, height=30)
    r += 2

    # ----- Підпис -----
    band(r, 'Виписав(ла): ______________________________', font=_F_NORMAL)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _money(value):
    """Грошовий формат у стилі '1 500,00' (пробіл-тисячі, кома-копійки)."""
    s = f'{Decimal(value):,.2f}'  # 1,500.00
    return s.replace(',', ' ').replace('.', ',')


def _invoice_template_ctx(reg):
    """Підготувати дані для HTML-шаблона рахунка."""
    ctx = _invoice_context(reg)
    s = ctx['settings']
    amount = ctx['amount']
    words = number_to_words_ua(amount, 'UAH')
    words = (words[0].upper() + words[1:]) if words else words
    return {
        'number': ctx['number'],
        'date_phrase': _date_phrase(ctx['date']),
        'supplier': {
            'name': s.company_full_name or 'ІПРМ',
            'iban': s.bank_iban or '',
            'bank': s.bank_name or '',
            'edrpou': s.edrpou or '',
            'tax_status': s.tax_status or '',
        },
        'payer': ctx['payer'],
        'item_name': ctx['item_name'],
        'amount_str': _money(amount),
        'amount_words': words,
    }


def render_invoice_pdf(reg) -> bytes:
    """Згенерувати PDF-рахунок через WeasyPrint (HTML -> PDF, без LibreOffice).

    Кидає InvoiceError, якщо рендер не вдався."""
    try:
        from weasyprint import HTML
    except Exception as exc:  # бібліотека не встановлена / системні залежності
        logger.exception('WeasyPrint unavailable for invoice reg=%s', reg.id)
        raise InvoiceError(f'PDF-рендер недоступний: {exc}')
    try:
        html = render_template('invoices/invoice.html', **_invoice_template_ctx(reg))
        return HTML(string=html, base_url=current_app.static_folder).write_pdf()
    except InvoiceError:
        raise
    except Exception as exc:
        logger.exception('Invoice PDF render failed for reg=%s', reg.id)
        raise InvoiceError(f'Помилка генерації PDF-рахунка: {exc}')


def invoice_filename(reg, ext):
    return f'rahunok-{reg.id}.{ext}'
