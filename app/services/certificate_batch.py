"""Масова генерація сертифікатів з xlsx (dry-run прев'ю + фоновий батч).

Потік:
  1. create_job(file)  -- зберігає завантажений xlsx у тимчасову теку job-а;
  2. parse_workbook()  -- розбирає рядки зі статусом OK/помилка (для прев'ю);
  3. start_job()       -- фоновий потік рендерить PDF валідних рядків,
     стрімить їх у ZIP на диску, оновлює status.json (прогрес);
  4. read_status()/ZIP -- сторінка статусу опитує прогрес і віддає архів.

Стан job-а тримаємо на диску (status.json) -- працює між gunicorn-воркерами
(потік в одному воркері, статус читають усі). Сертифікати в БД не пишемо.
"""
import csv
import io
import json
import logging
import os
import re
import tempfile
import threading
import time
import uuid
import zipfile
from datetime import datetime

from app.models.mixins import utcnow

logger = logging.getLogger(__name__)

# Колонки шаблону (порядок фіксований).
COLUMNS = [
    'ПІБ учасника',
    'Назва заходу',
    'Дата проведення (ДД.ММ.РРРР)',
    'Бали БПР',
    'ПІБ лектора',
    'Спеціальності',
    'Номер заходу БПР (7 цифр)',
    'Номер учасника (6 цифр)',
]
COLUMN_WIDTHS = [28, 42, 26, 10, 24, 30, 24, 22]
_NCOLS = len(COLUMNS)

MAX_ROWS = 2000  # запобіжник проти величезних файлів
_JOB_ID_RE = re.compile(r'^[0-9a-f]{32}$')

# ---- Транслітерація (укр -> латиниця) для імен файлів ----
_TRANSLIT = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'h', 'ґ': 'g', 'д': 'd', 'е': 'e',
    'є': 'ie', 'ж': 'zh', 'з': 'z', 'и': 'y', 'і': 'i', 'ї': 'i', 'й': 'i',
    'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
    'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch',
    'ш': 'sh', 'щ': 'shch', 'ь': '', 'ю': 'iu', 'я': 'ia', "'": '', 'ʼ': '',
}


def transliterate(text):
    """Укр-текст -> безпечне для файлової системи латинське ім'я."""
    out = []
    for ch in (text or ''):
        low = ch.lower()
        if low in _TRANSLIT:
            rep = _TRANSLIT[low]
            out.append(rep.capitalize() if ch.isupper() and rep else rep)
        elif ch.isalnum():
            out.append(ch)
        elif ch in ' _-':
            out.append('_')
        # інше -- відкидаємо
    name = re.sub(r'_+', '_', ''.join(out)).strip('_')
    return name or 'cert'


# ---- Файловий job-store ----
def _jobs_root():
    root = os.path.join(tempfile.gettempdir(), 'iprm-cert-jobs')
    os.makedirs(root, exist_ok=True)
    return root


def job_dir(job_id):
    if not _JOB_ID_RE.match(job_id or ''):
        raise ValueError('invalid job id')
    return os.path.join(_jobs_root(), job_id)


def _prune_old(max_age_seconds=12 * 3600):
    root = _jobs_root()
    now = time.time()
    for name in os.listdir(root):
        p = os.path.join(root, name)
        try:
            if os.path.isdir(p) and now - os.path.getmtime(p) > max_age_seconds:
                for f in os.listdir(p):
                    os.remove(os.path.join(p, f))
                os.rmdir(p)
        except OSError:
            pass


def create_job(file_storage):
    """Зберегти завантажений xlsx у нову job-теку. Повертає job_id."""
    _prune_old()
    job_id = uuid.uuid4().hex
    d = job_dir(job_id)
    os.makedirs(d, exist_ok=True)
    file_storage.save(os.path.join(d, 'source.xlsx'))
    _write_status(job_id, {'status': 'created', 'total': 0, 'done': 0,
                           'created': utcnow().isoformat()})
    return job_id


def _status_path(job_id):
    return os.path.join(job_dir(job_id), 'status.json')


def _write_status(job_id, data):
    """Атомарний запис status.json (temp + rename)."""
    path = _status_path(job_id)
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as fh:
        json.dump(data, fh, ensure_ascii=False)
    os.replace(tmp, path)


def read_status(job_id):
    try:
        with open(_status_path(job_id), encoding='utf-8') as fh:
            return json.load(fh)
    except (OSError, ValueError):
        return None


def zip_path(job_id):
    return os.path.join(job_dir(job_id), 'certificates.zip')


# ---- Парсинг ----
def _parse_date(val):
    if val in (None, ''):
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%d.%m.%y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_workbook(job_id):
    """Розібрати source.xlsx job-а у список рядків зі статусом.

    Кожен елемент: dict(row, name, title, date, date_str, cpd, lecturer,
    event, part, status='ok'|'error', problems=[...]).
    """
    from openpyxl import load_workbook
    path = os.path.join(job_dir(job_id), 'source.xlsx')
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    seen_part = {}
    for idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not raw or all(c is None or str(c).strip() == '' for c in raw):
            continue
        if len(rows) >= MAX_ROWS:
            rows.append({'row': idx, 'name': '', 'title': '', 'date': None,
                         'date_str': '', 'cpd': None, 'lecturer': None,
                         'specialties': None, 'event': '', 'part': '',
                         'status': 'error',
                         'problems': [f'перевищено ліміт {MAX_ROWS} рядків']})
            break
        cells = list(raw) + [None] * (_NCOLS - len(raw))
        (name, title, date_raw, cpd_raw, lecturer,
         spec_raw, event_raw, part_raw) = cells[:_NCOLS]
        name = str(name).strip() if name is not None else ''
        title = str(title).strip() if title is not None else ''
        dt = _parse_date(date_raw)
        event_num = str(event_raw).strip() if event_raw is not None else ''
        part_num = str(part_raw).strip() if part_raw is not None else ''
        problems = []
        if not name:
            problems.append('немає ПІБ')
        if not title:
            problems.append('немає назви заходу')
        if dt is None:
            problems.append('некоректна дата')
        if not event_num:
            problems.append('немає номера заходу')
        if not part_num:
            problems.append('немає номера учасника')
        cpd = None
        if cpd_raw not in (None, ''):
            try:
                cpd = int(float(str(cpd_raw).strip().replace(',', '.')))
            except ValueError:
                problems.append('бали БПР не число')
        # дубль номера учасника в межах файлу
        if part_num:
            key = (event_num, part_num)
            if key in seen_part:
                problems.append(f'дубль номера учасника (рядок {seen_part[key]})')
            else:
                seen_part[key] = idx
        rows.append({
            'row': idx, 'name': name, 'title': title, 'date': dt,
            'date_str': dt.strftime('%d.%m.%Y') if dt else (str(date_raw).strip() if date_raw else ''),
            'cpd': cpd, 'lecturer': str(lecturer).strip() if lecturer else None,
            'specialties': str(spec_raw).strip() if spec_raw else None,
            'event': event_num, 'part': part_num,
            'status': 'error' if problems else 'ok',
            'problems': problems,
        })
    return rows


# ---- Фоновий рендер ----
def start_job(job_id, provider):
    """Запустити фонову генерацію (daemon-потік)."""
    from flask import current_app
    app = current_app._get_current_object()
    t = threading.Thread(target=_run, args=(app, job_id, provider), daemon=True)
    t.start()


def _run(app, job_id, provider):
    with app.app_context():
        try:
            from weasyprint.text.fonts import FontConfiguration
            from app.services import certificate_service as cs
            from app.models.certificate import Certificate

            valid = [r for r in parse_workbook(job_id) if r['status'] == 'ok']
            total = len(valid)
            _write_status(job_id, {'status': 'running', 'total': total, 'done': 0,
                                   'created': utcnow().isoformat()})
            font_config = FontConfiguration()
            seen_names = {}
            summary = [['Файл', 'Номер', 'ПІБ', 'Захід', 'Дата']]

            with zipfile.ZipFile(zip_path(job_id), 'w', zipfile.ZIP_DEFLATED) as zf:
                for i, r in enumerate(valid, start=1):
                    number = Certificate.format_number(
                        r['date'].year, provider, r['event'], r['part'],
                    )
                    pdf = cs.render_adhoc_pdf(
                        number=number, recipient_name=r['name'],
                        event_title=r['title'], event_date=r['date'],
                        cpd_points=r['cpd'], lecturer_name=r['lecturer'],
                        specialties=r['specialties'],
                        issued_at=r['date'], font_config=font_config,
                    )
                    fname = f"{transliterate(r['name'])}_{number}.pdf"
                    if fname in seen_names:
                        seen_names[fname] += 1
                        fname = f"{fname[:-4]}_{seen_names[fname]}.pdf"
                    else:
                        seen_names[fname] = 1
                    zf.writestr(fname, pdf)
                    summary.append([fname, number, r['name'], r['title'], r['date_str']])
                    _write_status(job_id, {'status': 'running', 'total': total,
                                           'done': i, 'created': utcnow().isoformat()})

                # Підсумкова таблиця (utf-8-sig -> коректно в Excel).
                buf = io.StringIO()
                csv.writer(buf, delimiter=';').writerows(summary)
                zf.writestr('_summary.csv', '﻿' + buf.getvalue())

            _write_status(job_id, {'status': 'done', 'total': total, 'done': total,
                                   'finished': utcnow().isoformat()})
            logger.info('Cert batch %s done: %d certificates', job_id, total)
        except Exception as exc:
            logger.exception('Cert batch %s failed', job_id)
            _write_status(job_id, {'status': 'failed', 'error': str(exc)[:300]})
