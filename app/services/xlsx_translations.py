"""XLSX-канал перекладів контенту: експорт і розбір файлу.

Окремий від xlsx_io канал і окремий файл -- свідомо. Наявний імпорт курсів
пише в канонічні українські колонки і робить REPLACE блоків програми та
FAQ; канал перекладів не має доступу до жодної канонічної колонки взагалі,
тому затерти оригінал ним неможливо конструктивно, а не за домовленістю.

Формат однаковий для всіх сутностей -- рядок на один перекладний фрагмент:

    Тип | ID | Об'єкт | Поле | Ключ | Українська | Російська | English

`Ключ` -- це uid одиниці перекладу (`title`, `faq:9c1f2a3b4d5e`), стабільний
між експортом і імпортом і не залежний від позиції фрагмента в структурі.
`Українська` служить не лише довідкою: на імпорті значення звіряється з
поточним у БД, і розбіжність означає, що джерело змінилось після експорту
(рядок піде в конфлікти, а не застосується поверх іншого тексту).
"""
import io
import logging
from dataclasses import dataclass, field as dc_field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.i18n import PREFIXED_LANGUAGES
from app.services import translation_registry as registry
from app.services.xlsx_io import (
    HEADER_FILL, HEADER_FONT, WRAP,
    _apply_table_style, _apply_zebra, _set_column_widths, _style_header,
)

logger = logging.getLogger(__name__)

LANG_LABELS = {'ru': 'Російська', 'en': 'English'}

# Колонки-ключі та довідка; мовні колонки додаються за PREFIXED_LANGUAGES,
# тож нова мова -- це один рядок у config.LANGUAGES.
BASE_COLS = ['entity', 'id', 'object', 'field', 'key', 'uk']

BASE_LABELS = {
    'entity': 'Тип',
    'id': 'ID',
    'object': "Об'єкт",
    'field': 'Поле',
    'key': 'Ключ',
    'uk': 'Українська (оригінал)',
}

WIDTHS = {
    'entity': 16, 'id': 8, 'object': 38, 'field': 22, 'key': 22, 'uk': 60,
}
LANG_WIDTH = 60

# Колонки, які перекладач не редагує. Лист захищаємо, а мовні клітинки
# розблоковуємо -- так правки фізично можливі лише там, де треба.
LOCKED_COLS = set(BASE_COLS)


def translation_cols():
    return BASE_COLS + list(PREFIXED_LANGUAGES)


def translation_labels():
    return {**BASE_LABELS, **{lang: LANG_LABELS.get(lang, lang.upper())
                              for lang in PREFIXED_LANGUAGES}}


# ----------------------------------------------------------------------
# Що саме потрапляє в кожен файл
# ----------------------------------------------------------------------

def _course_sheets():
    from app.models.course import Course
    courses = Course.query.order_by(Course.title).all()
    yield 'Курси', 'course', [(c, c.title) for c in courses]
    yield 'Блоки програми', 'program_block', [
        (b, f'{c.title} -> {b.heading}')
        for c in courses
        for b in sorted(c.program_blocks, key=lambda b: b.sort_order or 0)
    ]
    yield 'Тарифи курсів', 'course_tariff', [
        (t, f'{c.title} -> {t.name}')
        for c in courses for t in c.default_tariffs
    ]


def _trainer_sheets():
    from app.models.trainer import Trainer
    trainers = Trainer.query.order_by(Trainer.full_name).all()
    yield 'Тренери', 'trainer', [(t, t.full_name) for t in trainers]


def _instance_sheets():
    from app.models.city import City
    from app.models.course_instance import CourseInstance

    instances = (
        CourseInstance.query
        .order_by(CourseInstance.start_date.desc().nullslast())
        .all()
    )

    def label(inst, tariff):
        course = inst.course.title if inst.course else f'#{inst.course_id}'
        when = inst.start_date.strftime('%d.%m.%Y') if inst.start_date else 'без дати'
        return f'{course} ({when}) -> {tariff.name}'

    yield 'Тарифи проведень', 'instance_tariff', [
        (t, label(inst, t)) for inst in instances for t in inst.tariffs
    ]
    # Локації розкладу перекладаються довідником, а не полем проведення --
    # див. app/services/city_glossary.py.
    yield 'Локації', 'city', [
        (c, c.name) for c in City.query.order_by(City.name).all()
    ]


SCOPES = {
    'courses': ('Переклади курсів', _course_sheets),
    'trainers': ('Переклади тренерів', _trainer_sheets),
    'instances': ('Переклади розкладу', _instance_sheets),
}


# ----------------------------------------------------------------------
# Експорт
# ----------------------------------------------------------------------

def collect_rows(entity, pairs, only_untranslated=False):
    """[(об'єкт, підпис)] -> рядки файлу.

    Одиниці з порожнім джерелом пропускаємо: перекладати нічого, а рядок
    лише засмічував би файл.
    """
    rows = []
    for obj, label in pairs:
        for unit in registry.units(obj):
            if not (unit.source or '').strip():
                continue
            values = {
                lang: registry.stored_value(obj, lang, unit)
                for lang in PREFIXED_LANGUAGES
            }
            if only_untranslated and all(values.values()):
                continue
            rows.append({
                'entity': entity,
                'id': obj.id,
                'object': label,
                'field': unit.label,
                'key': unit.uid,
                'uk': unit.source,
                **values,
            })
    return rows


def _write_sheet(wb, title, rows):
    cols = translation_cols()
    ws = wb.create_sheet(title)
    _style_header(ws, cols, translation_labels())

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key))
            if key in ('object', 'uk') or key in PREFIXED_LANGUAGES:
                cell.alignment = WRAP
            if key not in LOCKED_COLS:
                # Захист листа блокує все; відкриваємо рівно мовні клітинки.
                cell.protection = Protection(locked=False)

    last_row = max(ws.max_row, 1)
    _apply_zebra(ws, len(cols), first_data_row=2, last_data_row=last_row)
    _set_column_widths(ws, cols, {**WIDTHS, **{l: LANG_WIDTH for l in PREFIXED_LANGUAGES}})
    _apply_table_style(ws, cols, f'tbl{title.replace(" ", "")}', last_row)

    # Фільтрувати й сортувати можна, редагувати -- лише мовні колонки.
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.selectLockedCells = False
    return ws


def _write_help_sheet(wb, scope_title):
    ws = wb.create_sheet('Довідка', 0)
    lines = [
        (scope_title, True),
        ('', False),
        ('Заповнюйте лише мовні колонки. Решта колонок заблокована:', False),
        ('вони потрібні, щоб імпорт знайшов, куди покласти переклад.', False),
        ('', False),
        ('Порожня комірка = не чіпати наявний переклад.', False),
        ('Щоб ВИДАЛИТИ переклад, зробіть це в адмінці -- порожній файл', False),
        ('нічого не стирає, інакше частково заповнений файл витер би', False),
        ('готову роботу.', False),
        ('', False),
        ('Колонку "Українська (оригінал)" не редагуйте: на імпорті вона', False),
        ('звіряється з поточним текстом на сайті. Якщо оригінал змінили', False),
        ('після експорту, рядок потрапить у конфлікти і НЕ застосується --', False),
        ('щоб переклад не ліг поверх іншого тексту.', False),
        ('', False),
        ('Рядки, яких немає у файлі, не змінюються. Можна вантажити', False),
        ('файл частинами.', False),
    ]
    for idx, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=idx, column=1, value=text)
        cell.alignment = Alignment(vertical='top')
        if bold:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        elif text:
            cell.font = Font(size=11)
    ws.column_dimensions[get_column_letter(1)].width = 78
    return ws


def export_translations_xlsx(scope, only_untranslated=False):
    """Файл перекладів для розділу адмінки ('courses'|'trainers'|'instances')."""
    if scope not in SCOPES:
        raise ValueError(f'невідомий розділ перекладів: {scope!r}')
    title, sheets_provider = SCOPES[scope]

    wb = Workbook()
    wb.remove(wb.active)

    total = 0
    for sheet_title, entity, pairs in sheets_provider():
        rows = collect_rows(entity, pairs, only_untranslated=only_untranslated)
        _write_sheet(wb, sheet_title, rows)
        total += len(rows)

    _write_help_sheet(wb, title)
    logger.info('Exported %s translation rows for scope=%s (todo=%s)',
                total, scope, only_untranslated)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ----------------------------------------------------------------------
# Розбір завантаженого файлу
# ----------------------------------------------------------------------

HELP_SHEET = 'Довідка'

# Дії над рядком. Застосовуються лише 'add' і 'update'; решта -- показуються
# у прев'ю. Проблема в окремому рядку НЕ відхиляє весь файл: у перекладача
# цілком нормально мати застарілі рядки зі старого експорту.
ACTIONS = ('add', 'update', 'unchanged', 'conflict', 'error')


@dataclass
class TranslationChange:
    sheet: str
    line_no: int
    entity: str
    obj_id: int | None
    object_label: str
    field_label: str
    key: str
    lang: str
    old: str
    new: str
    action: str
    error: str | None = None


@dataclass
class TranslationsImportPlan:
    changes: list = dc_field(default_factory=list)
    errors: list = dc_field(default_factory=list)   # помилки рівня файлу

    @property
    def is_valid(self):
        """Файл придатний до застосування. Проблемні РЯДКИ не блокують --
        вони просто не застосовуються."""
        return not self.errors

    @property
    def counts(self):
        c = {a: 0 for a in ACTIONS}
        for ch in self.changes:
            c[ch.action] = c.get(ch.action, 0) + 1
        return c

    @property
    def applicable(self):
        return [ch for ch in self.changes if ch.action in ('add', 'update')]


def _read_translation_sheet(ws):
    """Рядки листа як list[(номер_рядка, dict)].

    Ключові колонки обов'язкові; мовні -- ні: якщо перекладач лишив у файлі
    тільки російську, англійську просто не імпортуємо.
    """
    if ws.max_row < 2:
        return [], []

    labels = translation_labels()
    accepted = {}
    for key in translation_cols():
        accepted[key.lower()] = key
        accepted[labels[key].strip().lower()] = key

    col_idx = {}
    for i, cell in enumerate(ws[1], start=1):
        raw = cell.value
        if raw is None:
            continue
        key = accepted.get(str(raw).strip().lower())
        if key:
            col_idx[key] = i

    missing = [k for k in BASE_COLS if k not in col_idx]
    if missing:
        pretty = ', '.join(labels[k] for k in missing)
        return [], [f'Лист "{ws.title}": бракує колонок: {pretty}']

    rows = []
    for line_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        rows.append((line_no, {
            key: row[idx - 1] for key, idx in col_idx.items()
        }))
    return rows, []


def _text(value):
    return '' if value is None else str(value).strip()


def parse_translations_xlsx(path):
    """Розібрати файл у план змін. БД не змінює."""
    plan = TranslationsImportPlan()
    try:
        wb = load_workbook(filename=str(Path(path)), read_only=True, data_only=True)
    except Exception as exc:
        plan.errors.append(f'Не вдалося прочитати файл: {exc}')
        return plan

    entities = registry.entity_registry()
    obj_cache = {}
    units_cache = {}

    def load_obj(entity, obj_id):
        cache_key = (entity, obj_id)
        if cache_key not in obj_cache:
            meta = entities.get(entity)
            obj_cache[cache_key] = (
                db.session.get(meta['model'], obj_id) if meta else None
            )
        return obj_cache[cache_key]

    def load_units(entity, obj):
        cache_key = (entity, obj.id)
        if cache_key not in units_cache:
            units_cache[cache_key] = {u.uid: u for u in registry.units(obj)}
        return units_cache[cache_key]

    sheets = [name for name in wb.sheetnames if name != HELP_SHEET]
    if not sheets:
        plan.errors.append('У файлі немає жодного листа з перекладами.')
        return plan

    for sheet_name in sheets:
        rows, sheet_errors = _read_translation_sheet(wb[sheet_name])
        plan.errors.extend(sheet_errors)

        for line_no, raw in rows:
            entity = _text(raw.get('entity'))
            key = _text(raw.get('key'))
            object_label = _text(raw.get('object'))
            field_label = _text(raw.get('field'))
            uk_in_file = _text(raw.get('uk'))

            def problem(message, obj_id=None):
                plan.changes.append(TranslationChange(
                    sheet=sheet_name, line_no=line_no, entity=entity,
                    obj_id=obj_id, object_label=object_label,
                    field_label=field_label, key=key, lang='',
                    old='', new='', action='error', error=message,
                ))

            try:
                obj_id = int(raw.get('id'))
            except (TypeError, ValueError):
                problem('порожній або нечисловий ID')
                continue

            if entity not in entities:
                problem(f'невідомий тип "{entity}"', obj_id)
                continue
            if not key:
                problem('порожній ключ', obj_id)
                continue

            obj = load_obj(entity, obj_id)
            if obj is None:
                problem(f'{entity} #{obj_id} не знайдено', obj_id)
                continue

            unit = load_units(entity, obj).get(key)
            if unit is None:
                problem('фрагмент більше не існує (текст переписали або видалили)',
                        obj_id)
                continue

            # Головна перевірка каналу: у файлі має бути ТОЙ САМИЙ оригінал,
            # інакше переклад ліг би поверх іншого тексту.
            if uk_in_file and uk_in_file != (unit.source or '').strip():
                plan.changes.append(TranslationChange(
                    sheet=sheet_name, line_no=line_no, entity=entity,
                    obj_id=obj_id, object_label=object_label,
                    field_label=field_label, key=key, lang='',
                    old=(unit.source or ''), new=uk_in_file, action='conflict',
                    error='оригінал змінився після експорту',
                ))
                continue

            for lang in PREFIXED_LANGUAGES:
                if lang not in raw:
                    continue
                new = _text(raw.get(lang))
                if not new:
                    continue  # порожньо = не чіпати
                old = registry.stored_value(obj, lang, unit)
                if new == old:
                    action = 'unchanged'
                else:
                    action = 'update' if old else 'add'
                plan.changes.append(TranslationChange(
                    sheet=sheet_name, line_no=line_no, entity=entity,
                    obj_id=obj_id, object_label=object_label,
                    field_label=field_label, key=key, lang=lang,
                    old=old, new=new, action=action,
                ))

    wb.close()
    return plan


def apply_translations_plan(plan):
    """Записати застосовні зміни. Торкається ВИКЛЮЧНО translations."""
    if not plan.is_valid:
        return {'ok': False, 'reason': 'plan has errors'}

    entities = registry.entity_registry()
    # (entity, id, lang) -> {uid: текст}
    grouped = {}
    for ch in plan.applicable:
        grouped.setdefault((ch.entity, ch.obj_id, ch.lang), {})[ch.key] = ch.new

    applied = 0
    try:
        for (entity, obj_id, lang), new_values in grouped.items():
            obj = db.session.get(entities[entity]['model'], obj_id)
            if obj is None:
                continue
            units = registry.units(obj)
            touched_fields = {
                u.field for u in units if u.uid in new_values
            }
            # apply_units перезбирає мапу оверрайдів поля ПОВНІСТЮ, тож
            # спершу засіваємо поточними значеннями всіх одиниць зачеплених
            # полів -- інакше частковий файл стер би фрагменти, яких у ньому
            # просто не було.
            values = {
                u.uid: registry.stored_value(obj, lang, u)
                for u in units if u.field in touched_fields
            }
            values.update(new_values)
            registry.apply_units(obj, lang, values)
            applied += len(new_values)

        db.session.commit()
        return {'ok': True, 'applied': applied, 'objects': len(grouped)}
    except Exception as exc:
        db.session.rollback()
        logger.exception('apply_translations_plan failed')
        return {'ok': False, 'reason': str(exc)}
