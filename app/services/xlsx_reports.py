"""Звітні xlsx-вивантаження списків адмінки (read-only).

На відміну від `xlsx_io`, який тримає ШАБЛОНИ для редагування (export + parse
у парі), тут -- звіти: те, що менеджер вивантажує з поточного зрізу сторінки й
більше не завантажує назад. Тому ані drop-down-ів, ані reference-аркушів; лише
таблиця з колонками сторінки та аркуш «Фільтри», що пояснює зріз.

Низькорівневі помічники (стилі, зебра, ширини, безпечний запис комірки)
лишаються в `xlsx_io` -- вигляд файлів має бути спільним.
"""
import io

from openpyxl import Workbook

from app.services.xlsx_io import (
    FORMAT_LABEL,
    PARTICIPANT_TYPE_LABEL,
    PAYMENT_STATUS_FILLS,
    PAYMENT_STATUS_LABEL,
    REG_STATUS_FILLS,
    REG_STATUS_LABEL,
    STATUS_FILLS,
    STATUS_LABEL,
    WRAP,
    _apply_number_formats,
    _apply_table_style,
    _apply_zebra,
    _set_column_widths,
    _style_header,
    _to_kyiv_naive,
    write_cell,
)

_REFERRAL_COLS = ['date', 'referrer', 'referrer_type', 'code', 'buyer',
                  'buyer_email', 'course', 'event_date', 'points', 'status']
_REFERRAL_LABELS = {
    'date': 'Дата нарахування', 'referrer': 'Реферер',
    'referrer_type': 'Тип реферера', 'code': 'Код', 'buyer': 'Учасник',
    'buyer_email': 'Email учасника', 'course': 'Захід',
    'event_date': 'Дата заходу', 'points': 'Бали', 'status': 'Статус',
}
_REFERRAL_WIDTHS = {'date': 16, 'referrer': 28, 'referrer_type': 14, 'code': 12,
                    'buyer': 28, 'buyer_email': 30, 'course': 40,
                    'event_date': 14, 'points': 8, 'status': 14}


def export_referral_rewards_xlsx(rewards, referrer_map,
                                 applied_filters=None) -> io.BytesIO:
    """Реєстр реферальних нарахувань -> xlsx (для звірки/виплат балів).

    rewards -- список ReferralReward (з підвантаженими registration/instance/
    course/user); referrer_map -- {code: {kind, name}} для імен рефереров;
    applied_filters -- пари (назва, значення) для аркуша «Фільтри».
    """
    rows = []
    for rw in rewards:
        ref = referrer_map.get(rw.referral_code) or {}
        reg = rw.registration
        user = reg.user if reg else None
        inst = reg.instance if reg else None
        course = inst.course if inst else None
        rows.append([
            rw.created_at.strftime('%d.%m.%Y') if rw.created_at else '',
            ref.get('name') or rw.referral_code,
            'Тренер' if ref.get('kind') == 'trainer' else (
                'Учасник' if ref.get('kind') == 'user' else '—'),
            rw.referral_code,
            (f'{user.first_name} {user.last_name}'.strip() if user else '—'),
            (user.email if user else ''),
            (course.title if course else '—'),
            (inst.start_date.strftime('%d.%m.%Y') if (inst and inst.start_date) else ''),
            rw.points,
            rw.status_label,
        ])

    return build_list_xlsx(
        'Реферали', _REFERRAL_COLS, _REFERRAL_LABELS, _REFERRAL_WIDTHS, rows,
        'tblReferral', applied_filters=applied_filters,
    )


_REGISTRATION_COLS = [
    'reg_id', 'event', 'event_date', 'location', 'trainer', 'place_number',
    'last_name', 'first_name', 'email', 'phone', 'workplace',
    'referrer', 'referrer_type', 'status', 'payment_status', 'payment_method',
    'payment_amount', 'promo_code', 'discount_amount', 'attended',
    'cpd_points_awarded', 'certificate', 'created_at',
]
_REGISTRATION_LABELS = {
    'reg_id': 'ID реєстрації', 'event': 'Захід', 'event_date': 'Дата заходу',
    'location': 'Місце', 'trainer': 'Тренер', 'place_number': 'Місце №',
    'last_name': 'Прізвище', 'first_name': "Ім'я", 'email': 'Email',
    'phone': 'Телефон', 'workplace': 'Місце роботи / місто',
    'referrer': 'Реферер', 'referrer_type': 'Тип реферера',
    'status': 'Статус', 'payment_status': 'Оплата',
    'payment_method': 'Спосіб оплати', 'payment_amount': 'Сума (грн)',
    'promo_code': 'Промокод', 'discount_amount': 'Знижка (грн)',
    'attended': 'Присутній', 'cpd_points_awarded': 'Бали БПР',
    'certificate': 'Сертифікат', 'created_at': 'Дата реєстрації',
}
_REGISTRATION_WIDTHS = {
    'reg_id': 12, 'event': 46, 'event_date': 14, 'location': 18, 'trainer': 26,
    'place_number': 10, 'last_name': 20, 'first_name': 18, 'email': 30,
    'phone': 18, 'workplace': 34, 'referrer': 26, 'referrer_type': 14,
    'status': 16, 'payment_status': 16, 'payment_method': 22,
    'payment_amount': 14, 'promo_code': 16, 'discount_amount': 14,
    'attended': 12, 'cpd_points_awarded': 12, 'certificate': 22,
    'created_at': 18,
}
_FILTERS_SHEET = 'Фільтри'


def _add_filters_sheet(wb, applied_filters, table_name) -> None:
    """Аркуш «Фільтри»: зріз даних, який лежить у файлі.

    Без нього вивантаження за фільтром неможливо відрізнити від повного:
    два файли з однаковими колонками, різним вмістом і без жодної підказки,
    чим саме вони відрізняються.
    """
    ws = wb.create_sheet(_FILTERS_SHEET)
    cols = ['filter', 'value']
    _style_header(ws, cols, {'filter': 'Фільтр', 'value': 'Значення'})
    for row_idx, (name, value) in enumerate(applied_filters, start=2):
        write_cell(ws, row_idx, 1, name)
        write_cell(ws, row_idx, 2, value).alignment = WRAP
    last_row = ws.max_row
    _set_column_widths(ws, cols, {'filter': 30, 'value': 60})
    _apply_zebra(ws, len(cols), first_data_row=2, last_data_row=last_row)
    _apply_table_style(ws, cols, table_name, last_row)


def build_list_xlsx(sheet_title, cols, labels, widths, rows, table_name,
                    applied_filters=None, row_fills=None) -> io.BytesIO:
    """Список адмінки -> xlsx: аркуш-таблиця + (за потреби) аркуш «Фільтри».

    Спільний білдер для read-only вивантажень зі сторінок-списків (реєстрації,
    користувачі): вони відрізняються лише набором колонок, тож стилі, зебра,
    ширини, формати чисел і таблиця-автофільтр живуть тут, а не копіюються
    в кожен експорт.

    rows       -- список списків значень у порядку `cols`;
    row_fills  -- паралельний `rows` список {col_key: PatternFill} (кольори
                  статусів); None -- без заливок;
    applied_filters -- пари (назва, значення) для аркуша «Фільтри».

    Це саме ЗВІТИ: імпорт назад їх не читає, тож drop-down-ів і reference-
    аркушів тут немає -- для редагування є xlsx-шаблони (`export_*_xlsx`
    з парою `parse_*_xlsx`).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    _style_header(ws, cols, labels)

    for row_idx, values in enumerate(rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            write_cell(ws, row_idx, col_idx, value).alignment = WRAP

    last_row = ws.max_row
    _apply_zebra(ws, len(cols), first_data_row=2, last_data_row=last_row)

    # Заливки статусів -- ПІСЛЯ зебри, щоб колір статусу її перекривав.
    for row_idx, fills in enumerate(row_fills or [], start=2):
        for key, fill in (fills or {}).items():
            ws.cell(row=row_idx, column=cols.index(key) + 1).fill = fill

    _set_column_widths(ws, cols, widths)
    _apply_number_formats(ws, cols, last_row)
    _apply_table_style(ws, cols, table_name, last_row)

    if applied_filters:
        _add_filters_sheet(wb, applied_filters, f'{table_name}Filters')

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_registrations_xlsx(regs, referrer_map=None,
                              applied_filters=None) -> io.BytesIO:
    """Реєстр реєстрацій (сторінка /admin/registrations) -> xlsx.

    regs -- уже відфільтрований і відсортований список EventRegistration із
    підвантаженими user / instance+course+trainer / promo_code / certificate.
    referrer_map -- {code: {kind, name}} для колонки «Реферер».
    applied_filters -- пари (назва, значення) для аркуша «Фільтри».

    Це вивантаження ЛИШЕ на читання (звіт), тож колонки дзеркалять таблицю
    адмінки, а не форму імпорту: редагувати учасників далі належить через
    xlsx-шаблон `export_participants_xlsx`, який імпорт уміє читати назад.
    """
    referrer_map = referrer_map or {}
    rows, row_fills = [], []
    for reg in regs:
        user = reg.user
        inst = reg.instance
        course = inst.course if inst else None
        # Ефективний тренер: тренер заходу, інакше -- тренер курсу (той самий
        # fallback, що й у фільтрі за тренером на сторінці).
        trainer = (inst.trainer if inst and inst.trainer else
                   (course.trainer if course else None))
        ref = referrer_map.get(reg.referral_code) or {}
        ref_kind = ref.get('kind')
        cert = reg.certificate
        rows.append([
            reg.id,
            (course.title if course else ''),
            _to_kyiv_naive(inst.start_date).date() if (inst and inst.start_date) else None,
            (inst.location if inst else '') or '',
            (trainer.full_name if trainer else ''),
            reg.place_number,
            (user.last_name if user else '') or '',
            (user.first_name if user else '') or '',
            (user.email if user else '') or '',
            reg.phone or '',
            reg.workplace or '',
            ref.get('name') or (reg.referral_code or ''),
            'Тренер' if ref_kind == 'trainer' else ('Учасник' if ref_kind == 'user' else ''),
            REG_STATUS_LABEL.get(reg.status, reg.status or ''),
            PAYMENT_STATUS_LABEL.get(reg.payment_status, reg.payment_status or ''),
            reg.payment_method_label or '',
            float(reg.payment_amount) if reg.payment_amount is not None else None,
            reg.promo_code.code if reg.promo_code else '',
            float(reg.discount_amount) if reg.discount_amount is not None else None,
            'Так' if reg.attended else 'Ні',
            reg.cpd_points_awarded,
            # Відкликаний сертифікат лишається в реєстрі, але з поміткою --
            # інакше звіт стверджував би, що документ чинний.
            (f'{cert.number} (відкликано)' if cert.revoked else cert.number) if cert else '',
            _to_kyiv_naive(reg.created_at),
        ])
        # Кольори за статусом / оплатою -- як в експорті учасників.
        fills = {}
        if reg.status in REG_STATUS_FILLS:
            fills['status'] = REG_STATUS_FILLS[reg.status]
        if reg.payment_status in PAYMENT_STATUS_FILLS:
            fills['payment_status'] = PAYMENT_STATUS_FILLS[reg.payment_status]
        row_fills.append(fills)

    return build_list_xlsx(
        'Реєстрації', _REGISTRATION_COLS, _REGISTRATION_LABELS,
        _REGISTRATION_WIDTHS, rows, 'tblRegistrations',
        applied_filters=applied_filters, row_fills=row_fills,
    )


_USER_COLS = [
    'user_id', 'last_name', 'first_name', 'middle_name', 'email', 'phone',
    'participant_type', 'workplace', 'position', 'specializations',
    'registrations', 'is_admin', 'is_active', 'email_confirmed',
    'email_opt_out', 'referral_code', 'last_login_at', 'created_at',
]
_USER_LABELS = {
    'user_id': 'ID', 'last_name': 'Прізвище', 'first_name': "Ім'я",
    'middle_name': 'По батькові', 'email': 'Email', 'phone': 'Телефон',
    'participant_type': 'Тип учасника', 'workplace': 'Місце роботи / місто',
    'position': 'Посада', 'specializations': 'Спеціалізації',
    'registrations': 'Реєстрацій', 'is_admin': 'Адмін',
    'is_active': 'Активний', 'email_confirmed': 'Email підтверджено',
    'email_opt_out': 'Відписка від розсилок', 'referral_code': 'Реф. код',
    'last_login_at': 'Останній вхід', 'created_at': 'Дата реєстрації',
}
_USER_WIDTHS = {
    'user_id': 8, 'last_name': 20, 'first_name': 18, 'middle_name': 20,
    'email': 32, 'phone': 18, 'participant_type': 26, 'workplace': 34,
    'position': 26, 'specializations': 40, 'registrations': 12,
    'is_admin': 10, 'is_active': 12, 'email_confirmed': 18,
    'email_opt_out': 20, 'referral_code': 14, 'last_login_at': 18,
    'created_at': 18,
}


def export_users_xlsx(users, applied_filters=None) -> io.BytesIO:
    """Реєстр користувачів (сторінка /admin/users) -> xlsx.

    users -- уже відфільтрований список User із підвантаженим medical_profile;
    кількість реєстрацій береться з `registration_count` (роут кладе туди
    агрегат одним запитом, тож тут немає N+1).
    """
    rows = []
    for user in users:
        profile = user.medical_profile
        ptype = profile.participant_type if profile else None
        rows.append([
            user.id,
            user.last_name or '',
            user.first_name or '',
            (profile.middle_name if profile else '') or '',
            user.email or '',
            (profile.phone if profile else '') or '',
            PARTICIPANT_TYPE_LABEL.get(ptype, '') if ptype else '',
            (profile.workplace if profile else '') or '',
            (profile.position if profile else '') or '',
            '\n'.join(user.specialization_labels or []),
            user.registration_count,
            'Так' if user.is_admin else 'Ні',
            'Так' if user.is_active else 'Ні',
            'Так' if user.email_confirmed else 'Ні',
            'Так' if user.email_opt_out else 'Ні',
            user.referral_code or '',
            _to_kyiv_naive(user.last_login_at),
            _to_kyiv_naive(user.created_at),
        ])

    return build_list_xlsx(
        'Користувачі', _USER_COLS, _USER_LABELS, _USER_WIDTHS, rows,
        'tblUsers', applied_filters=applied_filters,
    )


_CERT_COLS = [
    'number', 'recipient', 'email', 'event', 'event_date', 'cpd_points',
    'lecturer', 'event_place', 'issued_at', 'issued_by', 'state',
]
_CERT_LABELS = {
    'number': 'Номер', 'recipient': 'Учасник', 'email': 'Email',
    'event': 'Захід', 'event_date': 'Дата заходу', 'cpd_points': 'Бали БПР',
    'lecturer': 'Лектор', 'event_place': 'Місце', 'issued_at': 'Видано',
    'issued_by': 'Ким видано', 'state': 'Стан',
}
_CERT_WIDTHS = {
    'number': 28, 'recipient': 28, 'email': 30, 'event': 46, 'event_date': 14,
    'cpd_points': 10, 'lecturer': 26, 'event_place': 20, 'issued_at': 18,
    'issued_by': 28, 'state': 14,
}


def export_certificates_xlsx(certs, applied_filters=None) -> io.BytesIO:
    """Реєстр сертифікатів (сторінка /admin/certificates) -> xlsx.

    Колонки -- знімки з самого сертифіката (ПІБ, назва заходу, бали), бо саме
    вони надруковані в PDF; поточні дані курсу могли відтоді змінитись.
    """
    rows, row_fills = [], []
    for cert in certs:
        rows.append([
            cert.number or '',
            cert.recipient_name or '',
            (cert.user.email if cert.user else '') or '',
            cert.event_title or '',
            _to_kyiv_naive(cert.event_date).date() if cert.event_date else None,
            cert.cpd_points,
            cert.lecturer_name or '',
            cert.event_place or '',
            _to_kyiv_naive(cert.issued_at),
            (cert.issued_by.email if cert.issued_by else '') or '',
            'Відкликано' if cert.revoked else 'Дійсний',
        ])
        row_fills.append({
            'state': REG_STATUS_FILLS['cancelled'] if cert.revoked
            else REG_STATUS_FILLS['completed'],
        })

    return build_list_xlsx(
        'Сертифікати', _CERT_COLS, _CERT_LABELS, _CERT_WIDTHS, rows,
        'tblCertificates', applied_filters=applied_filters, row_fills=row_fills,
    )


_B2B_COLS = ['created_at', 'last_name', 'first_name', 'phone', 'email',
             'team_size', 'status', 'admin_notes']
_B2B_LABELS = {
    'created_at': 'Отримано', 'last_name': 'Прізвище', 'first_name': "Ім'я",
    'phone': 'Телефон', 'email': 'Email', 'team_size': 'Фахівців',
    'status': 'Статус', 'admin_notes': 'Нотатки',
}
_B2B_WIDTHS = {
    'created_at': 18, 'last_name': 20, 'first_name': 18, 'phone': 18,
    'email': 30, 'team_size': 14, 'status': 18, 'admin_notes': 46,
}


def export_b2b_requests_xlsx(requests, applied_filters=None) -> io.BytesIO:
    """Заявки на корпоративне навчання (/admin/b2b-requests) -> xlsx."""
    rows = [
        [
            _to_kyiv_naive(req.created_at),
            req.last_name or '',
            req.first_name or '',
            req.phone or '',
            req.email or '',
            req.team_size_label or '',
            req.status_label or '',
            req.admin_notes or '',
        ]
        for req in requests
    ]
    return build_list_xlsx(
        'B2B-заявки', _B2B_COLS, _B2B_LABELS, _B2B_WIDTHS, rows,
        'tblB2BRequests', applied_filters=applied_filters,
    )


_CREQ_COLS = ['created_at', 'course', 'email', 'phone', 'status', 'message',
              'admin_notes', 'updated_at']
_CREQ_LABELS = {
    'created_at': 'Отримано', 'course': 'Курс', 'email': 'Email',
    'phone': 'Телефон', 'status': 'Статус', 'message': 'Повідомлення',
    'admin_notes': 'Нотатки', 'updated_at': 'Оновлено',
}
_CREQ_WIDTHS = {
    'created_at': 18, 'course': 46, 'email': 30, 'phone': 18, 'status': 18,
    'message': 50, 'admin_notes': 40, 'updated_at': 18,
}


def export_course_requests_xlsx(requests, applied_filters=None) -> io.BytesIO:
    """Запити на проведення курсів (/admin/course-requests) -> xlsx."""
    rows = [
        [
            _to_kyiv_naive(req.created_at),
            (req.course.title if req.course else ''),
            req.email or '',
            req.phone or '',
            req.status_label or '',
            req.message or '',
            req.admin_notes or '',
            _to_kyiv_naive(req.updated_at),
        ]
        for req in requests
    ]
    return build_list_xlsx(
        'Запити на курси', _CREQ_COLS, _CREQ_LABELS, _CREQ_WIDTHS, rows,
        'tblCourseRequests', applied_filters=applied_filters,
    )


_PROMO_COLS = ['code', 'description', 'discount', 'scope', 'used_count',
               'max_uses', 'per_user_limit', 'valid_from', 'valid_until',
               'state', 'created_at']
_PROMO_LABELS = {
    'code': 'Код', 'description': 'Опис', 'discount': 'Знижка',
    'scope': 'Дія', 'used_count': 'Використань', 'max_uses': 'Ліміт',
    'per_user_limit': 'На людину', 'valid_from': 'Діє з',
    'valid_until': 'Діє до', 'state': 'Стан', 'created_at': 'Створено',
}
_PROMO_WIDTHS = {
    'code': 22, 'description': 40, 'discount': 14, 'scope': 46,
    'used_count': 12, 'max_uses': 10, 'per_user_limit': 12,
    'valid_from': 14, 'valid_until': 14, 'state': 16, 'created_at': 18,
}


def export_promo_codes_xlsx(promos, applied_filters=None) -> io.BytesIO:
    """Промокоди (/admin/promo-codes) -> xlsx для звірки знижок."""
    rows = [
        [
            promo.code or '',
            promo.description or '',
            promo.discount_label or '',
            promo.scope_label or '',
            promo.used_count or 0,
            promo.max_uses,
            promo.per_user_limit,
            _to_kyiv_naive(promo.valid_from).date() if promo.valid_from else None,
            _to_kyiv_naive(promo.valid_until).date() if promo.valid_until else None,
            promo.status_label or '',
            _to_kyiv_naive(promo.created_at),
        ]
        for promo in promos
    ]
    return build_list_xlsx(
        'Промокоди', _PROMO_COLS, _PROMO_LABELS, _PROMO_WIDTHS, rows,
        'tblPromoCodes', applied_filters=applied_filters,
    )


_EMAIL_LOG_COLS = ['created_at', 'to_email', 'subject', 'template', 'trigger',
                   'status', 'retry_count', 'sent_at', 'error_message']
_EMAIL_LOG_LABELS = {
    'created_at': 'Створено', 'to_email': 'Кому', 'subject': 'Тема',
    'template': 'Шаблон', 'trigger': 'Тригер', 'status': 'Статус',
    'retry_count': 'Спроб', 'sent_at': 'Відправлено',
    'error_message': 'Помилка',
}
_EMAIL_LOG_WIDTHS = {
    'created_at': 18, 'to_email': 30, 'subject': 50, 'template': 26,
    'trigger': 20, 'status': 16, 'retry_count': 10, 'sent_at': 18,
    'error_message': 50,
}


def export_email_logs_xlsx(logs, applied_filters=None) -> io.BytesIO:
    """Журнал листів (/admin/notifications/log) -> xlsx для розбору доставки."""
    rows = [
        [
            _to_kyiv_naive(log.created_at),
            log.to_email or '',
            log.subject or '',
            log.template_name or '',
            log.trigger_label or '',
            log.status_label or '',
            log.retry_count or 0,
            _to_kyiv_naive(log.sent_at),
            log.error_message or '',
        ]
        for log in logs
    ]
    return build_list_xlsx(
        'Журнал листів', _EMAIL_LOG_COLS, _EMAIL_LOG_LABELS, _EMAIL_LOG_WIDTHS,
        rows, 'tblEmailLog', applied_filters=applied_filters,
    )


_ERRLOG_COLS = ['created_at', 'error_code', 'error_type', 'error_message',
                'url', 'method', 'user', 'ip_address', 'state', 'resolved_at']
_ERRLOG_LABELS = {
    'created_at': 'Коли', 'error_code': 'Код', 'error_type': 'Тип',
    'error_message': 'Повідомлення', 'url': 'URL', 'method': 'Метод',
    'user': 'Користувач', 'ip_address': 'IP', 'state': 'Стан',
    'resolved_at': 'Вирішено',
}
_ERRLOG_WIDTHS = {
    'created_at': 18, 'error_code': 8, 'error_type': 26, 'error_message': 60,
    'url': 50, 'method': 10, 'user': 30, 'ip_address': 18, 'state': 14,
    'resolved_at': 18,
}


def export_error_logs_xlsx(logs, applied_filters=None) -> io.BytesIO:
    """Журнал помилок (/admin/error-logs) -> xlsx для розбору інцидентів."""
    rows = [
        [
            _to_kyiv_naive(log.created_at),
            log.error_code,
            log.error_type or '',
            log.error_message or '',
            log.url or '',
            log.method or '',
            (log.user.email if log.user else ''),
            log.ip_address or '',
            'Вирішено' if log.resolved else 'Відкрито',
            _to_kyiv_naive(log.resolved_at),
        ]
        for log in logs
    ]
    return build_list_xlsx(
        'Помилки', _ERRLOG_COLS, _ERRLOG_LABELS, _ERRLOG_WIDTHS, rows,
        'tblErrorLog', applied_filters=applied_filters,
    )


_INST_REPORT_COLS = ['id', 'course', 'start_date', 'end_date', 'event_format',
                     'location', 'trainer', 'price', 'cpd_points',
                     'max_participants', 'registrations', 'occupied',
                     'seats_left', 'status']
_INST_REPORT_LABELS = {
    'id': 'ID', 'course': 'Курс', 'start_date': 'Початок', 'end_date': 'Кінець',
    'event_format': 'Формат', 'location': 'Місце', 'trainer': 'Тренер',
    'price': 'Ціна', 'cpd_points': 'Бали БПР', 'max_participants': 'Місць',
    'registrations': 'Реєстрацій', 'occupied': 'Оплачено місць',
    'seats_left': 'Вільно', 'status': 'Статус',
}
_INST_REPORT_WIDTHS = {
    'id': 8, 'course': 46, 'start_date': 18, 'end_date': 18,
    'event_format': 14, 'location': 20, 'trainer': 26, 'price': 14,
    'cpd_points': 10, 'max_participants': 10, 'registrations': 12,
    'occupied': 14, 'seats_left': 10, 'status': 16,
}


def export_instances_report_xlsx(instances, reg_counts, occupied_map=None,
                                 applied_filters=None) -> io.BytesIO:
    """Проведення (/admin/instances) -> xlsx-звіт із завантаженістю.

    reg_counts -- {instance_id: активних реєстрацій}; occupied_map --
    {instance_id: ОПЛАЧЕНИХ місць} (лише вони тримають місце, див.
    services.seating). «Вільно» рахуємо від ефективної місткості
    (проведення, інакше курсу), NULL -- без обмежень; перевищення видно
    як «Оплачено місць» > «Місць».
    """
    occupied_map = occupied_map or {}
    rows, row_fills = [], []
    for inst in instances:
        course = inst.course
        trainer = inst.trainer or (course.trainer if course else None)
        taken = reg_counts.get(inst.id, 0)
        occupied = occupied_map.get(inst.id, 0)
        capacity = inst.max_participants
        if capacity is None and course is not None:
            capacity = course.max_participants
        rows.append([
            inst.id,
            (course.title if course else ''),
            _to_kyiv_naive(inst.start_date),
            _to_kyiv_naive(inst.end_date),
            FORMAT_LABEL.get(inst.event_format, inst.event_format or ''),
            inst.location or '',
            (trainer.full_name if trainer else ''),
            float(inst.effective_price) if inst.effective_price else None,
            inst.effective_cpd_points,
            capacity,
            taken,
            occupied,
            max(capacity - occupied, 0) if capacity is not None else None,
            STATUS_LABEL.get(inst.status, inst.status or ''),
        ])
        row_fills.append(
            {'status': STATUS_FILLS[inst.status]}
            if inst.status in STATUS_FILLS else {}
        )

    return build_list_xlsx(
        'Проведення', _INST_REPORT_COLS, _INST_REPORT_LABELS,
        _INST_REPORT_WIDTHS, rows, 'tblInstancesReport',
        applied_filters=applied_filters, row_fills=row_fills,
    )
