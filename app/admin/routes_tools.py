"""Адмін-інструменти: генератор сертифікатів з xlsx.

Потік: завантаження -> прев'ю (dry-run з OK/помилками) -> фонова генерація
(сторінка статусу з прогресом) -> ZIP з PDF. Логіка батчу -- у
app/services/certificate_batch.py. Сертифікати в БД не зберігаються.
"""
import io
import logging

from flask import (
    abort, flash, jsonify, redirect, render_template, request, send_file, url_for,
)
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.admin import admin_bp
from app.rbac import permission_required
from app.models.site_settings import SiteSettings
from app.models.trainer import Trainer
from app.services import certificate_batch as batch

logger = logging.getLogger(__name__)
audit_logger = logging.getLogger('audit')


def _provider():
    return (SiteSettings.get().bpr_provider_number or '').strip()


@admin_bp.route('/tools/certificate-generator')
@permission_required('cert_generator.view')
def tool_certificate_generator():
    return render_template(
        'admin/tools_certificate_generator.html',
        columns=batch.COLUMNS,
        provider_number=_provider(),
    )


@admin_bp.route('/tools/certificate-generator/template')
@permission_required('cert_generator.view')
def tool_certificate_generator_template():
    """Порожній xlsx-шаблон із заголовками + приклад-рядок."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Сертифікати'
    ws.append(batch.COLUMNS)
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='7437C9')
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
    ws.append([
        'Учасник', 'Шевченко Тарас Григорович', 'семінар',
        'Сучасні протоколи PRP-терапії', '15.05.2026', 'м. Київ', 10,
        'Абрамович Є.В.', 'усі лікарські спеціальності', '1028974', '1',
    ])
    # Приклад лектора: ПІБ у давальному відмінку, тип заходу в родовому,
    # номер у діапазоні 1xxxxx (окремий для лекторів), поле "ПІБ лектора" порожнє.
    ws.append([
        'Тренер', 'Абрамовичу Євгену Володимировичу', 'семінару',
        'Сучасні протоколи PRP-терапії', '15.05.2026', 'м. Київ', 20,
        '', 'усі лікарські спеціальності', '1028974', '100001',
    ])
    for i, w in enumerate(batch.COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # Випадаючий список типу особи (Учасник/Тренер) у першій колонці.
    type_col = get_column_letter(batch.COLUMNS.index('Тип особи (Учасник/Тренер)') + 1)
    dv_type = DataValidation(
        type='list', formula1='"{}"'.format(','.join(batch.PERSON_TYPES)),
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    dv_type.add(f'{type_col}2:{type_col}1000')

    # Випадаючий список лекторів у колонці "ПІБ лектора": ПІБ тренерів
    # тримаємо на прихованому аркуші, на колонку вішаємо list-валідацію.
    names = [t.full_name for t in Trainer.query
             .filter_by(is_active=True).order_by(Trainer.full_name).all()]
    if names:
        ref = wb.create_sheet('lectors')
        for i, nm in enumerate(names, start=1):
            ref.cell(row=i, column=1, value=nm)
        ref.sheet_state = 'hidden'
        col = get_column_letter(batch.COLUMNS.index('ПІБ лектора') + 1)
        dv = DataValidation(
            type='list', formula1=f'=lectors!$A$1:$A${len(names)}',
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(f'{col}2:{col}1000')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='certificate-generator-template.xlsx',
    )


@admin_bp.route('/tools/certificate-generator/preview', methods=['POST'])
@permission_required('cert_generator.manage')
def tool_certificate_generator_preview():
    """Розібрати завантажений xlsx і показати таблицю OK/помилки (dry-run)."""
    if not _provider():
        flash('Спочатку задайте реєстраційний номер провайдера БПР '
              '(Налаштування сайту)', 'error')
        return redirect(url_for('admin.tool_certificate_generator'))

    file = request.files.get('xlsx')
    if not file or not file.filename:
        flash('Файл не вибрано', 'error')
        return redirect(url_for('admin.tool_certificate_generator'))
    if not file.filename.lower().endswith('.xlsx'):
        flash('Потрібен файл формату .xlsx', 'error')
        return redirect(url_for('admin.tool_certificate_generator'))

    try:
        job_id = batch.create_job(file)
        rows = batch.parse_workbook(job_id)
    except Exception:
        logger.exception('certificate-generator: failed to parse xlsx')
        flash('Не вдалося прочитати xlsx-файл', 'error')
        return redirect(url_for('admin.tool_certificate_generator'))

    ok = sum(1 for r in rows if r['status'] == 'ok')
    return render_template(
        'admin/tools_certificate_preview.html',
        rows=rows, job_id=job_id,
        ok_count=ok, err_count=len(rows) - ok,
        provider_number=_provider(),
    )


@admin_bp.route('/tools/certificate-generator/generate', methods=['POST'])
@permission_required('cert_generator.manage')
def tool_certificate_generator_run():
    """Запустити фонову генерацію для job-а з прев'ю."""
    provider = _provider()
    if not provider:
        flash('Не задано номер провайдера БПР', 'error')
        return redirect(url_for('admin.tool_certificate_generator'))

    job_id = request.form.get('job_id', '')
    try:
        rows = batch.parse_workbook(job_id)
    except Exception:
        flash('Сесію генерації не знайдено. Завантажте файл повторно.', 'error')
        return redirect(url_for('admin.tool_certificate_generator'))

    if not any(r['status'] == 'ok' for r in rows):
        flash('Немає валідних рядків для генерації', 'error')
        return redirect(url_for('admin.tool_certificate_generator'))

    batch.start_job(job_id, provider)
    audit_logger.info('Admin %s started cert batch %s', current_user.email, job_id)
    return redirect(url_for('admin.tool_certificate_generator_job', job_id=job_id))


@admin_bp.route('/tools/certificate-generator/job/<job_id>')
@permission_required('cert_generator.view')
def tool_certificate_generator_job(job_id):
    status = batch.read_status(job_id)
    if status is None:
        flash('Завдання не знайдено', 'error')
        return redirect(url_for('admin.tool_certificate_generator'))
    return render_template('admin/tools_certificate_status.html',
                           job_id=job_id, status=status)


@admin_bp.route('/tools/certificate-generator/job/<job_id>/status')
@permission_required('cert_generator.view')
def tool_certificate_generator_job_status(job_id):
    status = batch.read_status(job_id)
    if status is None:
        return jsonify({'status': 'unknown'}), 404
    return jsonify(status)


@admin_bp.route('/tools/certificate-generator/job/<job_id>/download')
@permission_required('cert_generator.view')
def tool_certificate_generator_job_download(job_id):
    status = batch.read_status(job_id)
    if status is None or status.get('status') != 'done':
        flash('Архів ще не готовий', 'error')
        return redirect(url_for('admin.tool_certificate_generator_job', job_id=job_id))
    try:
        path = batch.zip_path(job_id)
    except ValueError:
        abort(404)
    return send_file(path, mimetype='application/zip',
                     as_attachment=True, download_name='certificates.zip')
